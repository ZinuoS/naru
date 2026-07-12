"""Unit tests for src/naru/artifact.py."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from naru import artifact

VALID_MANIFEST = """\
name: test_pipeline
version: v1
sheet: Sheet1
target_table: final_test
key: [id]
"""

VALID_FINGERPRINT = """\
{
  "sheet": "Sheet1",
  "header_row": 1,
  "columns": [
    {"name": "id", "type": "integer", "strictness": "strict"}
  ]
}
"""

VALID_SCHEMA = """\
from pydantic import BaseModel


class SourceRow(BaseModel):
    id: int


class TargetRow(BaseModel):
    id: int
"""

VALID_TRANSFORM = """\
def transform(df):
    return df
"""

VALID_VALIDATIONS = """\
row_count:
  min: 1
"""

VALID_CHANGELOG = "# Changelog\n\nv1: initial.\n"


def _write_valid_artifact(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    (root / "manifest.yaml").write_text(VALID_MANIFEST)
    (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (root / "schema.py").write_text(VALID_SCHEMA)
    (root / "transform.py").write_text(VALID_TRANSFORM)
    (root / "validations.yaml").write_text(VALID_VALIDATIONS)
    (root / "CHANGELOG.md").write_text(VALID_CHANGELOG)

    golden = root / "golden"
    golden.mkdir(exist_ok=True)
    wb = Workbook()
    wb.active.cell(row=1, column=1, value="id")
    wb.active.cell(row=2, column=1, value=1)
    wb.save(golden / "input_sample.xlsx")
    pd.DataFrame({"id": [1]}).to_parquet(golden / "expected_output.parquet", index=False)


@pytest.fixture
def valid_artifact_dir(tmp_path: Path) -> Path:
    root = tmp_path / "artifact"
    _write_valid_artifact(root)
    return root


class TestLoadArtifact:
    def test_valid_artifact_loads_successfully(self, valid_artifact_dir: Path) -> None:
        result = artifact.load_artifact(valid_artifact_dir)
        assert result.manifest.name == "test_pipeline"
        assert result.manifest.key == ["id"]
        assert result.fingerprint.sheet == "Sheet1"
        assert result.source_row.__name__ == "SourceRow"
        assert result.target_row.__name__ == "TargetRow"
        assert callable(result.transform)
        df = pd.DataFrame({"id": [1]})
        pd.testing.assert_frame_equal(result.transform(df), df)

    def test_directory_not_found_raises(self, tmp_path: Path) -> None:
        with pytest.raises(artifact.ArtifactLoadError, match="artifact directory not found"):
            artifact.load_artifact(tmp_path / "does_not_exist")

    def test_missing_manifest_raises_naming_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "manifest.yaml").unlink()
        with pytest.raises(
            artifact.ArtifactLoadError, match="manifest.yaml: required file missing"
        ):
            artifact.load_artifact(valid_artifact_dir)

    def test_missing_fingerprint_raises_naming_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "fingerprint.json").unlink()
        with pytest.raises(
            artifact.ArtifactLoadError, match="fingerprint.json: required file missing"
        ):
            artifact.load_artifact(valid_artifact_dir)

    def test_missing_transform_raises_naming_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "transform.py").unlink()
        with pytest.raises(artifact.ArtifactLoadError, match="transform.py: required file missing"):
            artifact.load_artifact(valid_artifact_dir)

    def test_missing_validations_raises_naming_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "validations.yaml").unlink()
        with pytest.raises(
            artifact.ArtifactLoadError, match="validations.yaml: required file missing"
        ):
            artifact.load_artifact(valid_artifact_dir)

    def test_missing_changelog_raises_naming_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "CHANGELOG.md").unlink()
        with pytest.raises(artifact.ArtifactLoadError, match="CHANGELOG.md: required file missing"):
            artifact.load_artifact(valid_artifact_dir)

    def test_missing_golden_dir_raises(self, valid_artifact_dir: Path) -> None:
        import shutil

        shutil.rmtree(valid_artifact_dir / "golden")
        with pytest.raises(artifact.ArtifactLoadError, match="golden: required directory missing"):
            artifact.load_artifact(valid_artifact_dir)

    def test_missing_golden_input_sample_raises(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "golden" / "input_sample.xlsx").unlink()
        with pytest.raises(
            artifact.ArtifactLoadError, match="input_sample.xlsx: required file missing"
        ):
            artifact.load_artifact(valid_artifact_dir)

    def test_missing_golden_expected_output_raises(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "golden" / "expected_output.parquet").unlink()
        with pytest.raises(
            artifact.ArtifactLoadError, match="expected_output.parquet: required file missing"
        ):
            artifact.load_artifact(valid_artifact_dir)

    def test_manifest_missing_field_names_field(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "manifest.yaml").write_text("name: test_pipeline\nversion: v1\n")
        with pytest.raises(artifact.ArtifactLoadError, match=r"manifest\.yaml: sheet:"):
            artifact.load_artifact(valid_artifact_dir)

    def test_manifest_invalid_yaml_raises(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "manifest.yaml").write_text("name: [unclosed\n")
        with pytest.raises(artifact.ArtifactLoadError, match="invalid YAML"):
            artifact.load_artifact(valid_artifact_dir)

    def test_fingerprint_invalid_json_raises(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "fingerprint.json").write_text("{not valid json")
        with pytest.raises(artifact.ArtifactLoadError, match="invalid JSON"):
            artifact.load_artifact(valid_artifact_dir)

    def test_fingerprint_bad_strictness_value_names_field(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "fingerprint.json").write_text(
            """
            {
              "sheet": "Sheet1",
              "header_row": 1,
              "columns": [
                {"name": "id", "type": "integer", "strictness": "bogus"}
              ]
            }
            """
        )
        with pytest.raises(artifact.ArtifactLoadError, match=r"columns\.0\.strictness"):
            artifact.load_artifact(valid_artifact_dir)

    def test_validations_invalid_yaml_raises(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "validations.yaml").write_text("row_count: [unclosed\n")
        with pytest.raises(artifact.ArtifactLoadError, match="invalid YAML"):
            artifact.load_artifact(valid_artifact_dir)

    def test_schema_missing_target_row_names_class(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "schema.py").write_text(
            "from pydantic import BaseModel\n\n\nclass SourceRow(BaseModel):\n    id: int\n"
        )
        with pytest.raises(artifact.ArtifactLoadError, match="missing required class 'TargetRow'"):
            artifact.load_artifact(valid_artifact_dir)

    def test_schema_class_not_basemodel_names_class(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "schema.py").write_text(
            "from pydantic import BaseModel\n\n\n"
            "class SourceRow(BaseModel):\n    id: int\n\n\n"
            "class TargetRow:\n    pass\n"
        )
        with pytest.raises(
            artifact.ArtifactLoadError, match="'TargetRow' must be a pydantic BaseModel"
        ):
            artifact.load_artifact(valid_artifact_dir)

    def test_schema_import_error_names_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "schema.py").write_text("this is not valid python (((\n")
        with pytest.raises(artifact.ArtifactLoadError, match="failed to import"):
            artifact.load_artifact(valid_artifact_dir)

    def test_transform_syntax_error_names_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "transform.py").write_text("def transform(df:\n    return df\n")
        with pytest.raises(artifact.ArtifactLoadError, match=r"transform\.py: syntax error"):
            artifact.load_artifact(valid_artifact_dir)

    def test_transform_missing_function_names_file(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "transform.py").write_text("x = 1\n")
        with pytest.raises(
            artifact.ArtifactLoadError, match="missing required function 'transform'"
        ):
            artifact.load_artifact(valid_artifact_dir)

    def test_transform_function_not_callable_raises(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "transform.py").write_text("transform = 5\n")
        with pytest.raises(artifact.ArtifactLoadError, match="'transform' must be callable"):
            artifact.load_artifact(valid_artifact_dir)

    def test_module_level_import_is_rejected(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "transform.py").write_text(
            "import os\n\n\ndef transform(df):\n    return df\n"
        )
        with pytest.raises(artifact.ArtifactLoadError, match=r"transform\.py: failed to execute"):
            artifact.load_artifact(valid_artifact_dir)

    def test_disallowed_builtin_is_rejected(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "transform.py").write_text(
            "open('/etc/passwd')\n\n\ndef transform(df):\n    return df\n"
        )
        with pytest.raises(artifact.ArtifactLoadError, match=r"transform\.py: failed to execute"):
            artifact.load_artifact(valid_artifact_dir)

    def test_allowed_builtins_and_ops_are_usable(self, valid_artifact_dir: Path) -> None:
        (valid_artifact_dir / "transform.py").write_text(
            "def transform(df):\n"
            "    assert len(df.columns) > 0\n"
            "    return ops.tag_verification(df, 'TO_VERIFY')\n"
        )
        result = artifact.load_artifact(valid_artifact_dir)
        out = result.transform(pd.DataFrame({"id": [1]}))
        assert out["_verification"].tolist() == ["TO_VERIFY"]

    def test_import_buried_in_function_body_is_not_caught_at_load_time(
        self, valid_artifact_dir: Path
    ) -> None:
        # Documents the known gap from docs/adr/0003-transform-loading.md:
        # module-level code runs at load time, but a function body's
        # statements only execute when the function is later called.
        (valid_artifact_dir / "transform.py").write_text(
            "def transform(df):\n    import os\n    return df\n"
        )
        result = artifact.load_artifact(valid_artifact_dir)
        assert callable(result.transform)
        # __import__ isn't in the builtins allowlist, so this raises
        # ImportError only once transform() actually runs -- not at load
        # time, since function bodies don't execute at definition time.
        with pytest.raises(ImportError):
            result.transform(pd.DataFrame({"id": [1]}))


class TestLowLevelHelpers:
    """Direct tests of the reusable file-loading helpers' own not-found
    guards, which load_artifact()'s upfront REQUIRED_FILES check makes
    unreachable through the public API alone.
    """

    def test_load_yaml_model_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(artifact.ArtifactLoadError, match="file not found"):
            artifact._load_yaml_model(tmp_path / "missing.yaml", artifact.Manifest)

    def test_load_json_model_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(artifact.ArtifactLoadError, match="file not found"):
            artifact._load_json_model(tmp_path / "missing.json", artifact.Fingerprint)

    def test_load_python_module_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(artifact.ArtifactLoadError, match="file not found"):
            artifact._load_python_module(tmp_path / "missing.py", "mod")

    def test_load_transform_missing_raises(self, tmp_path: Path) -> None:
        with pytest.raises(artifact.ArtifactLoadError, match="file not found"):
            artifact._load_transform(tmp_path / "missing.py")
