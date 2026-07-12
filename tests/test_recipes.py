"""Integration tests for the recipes/ shipped with this repo (spec.md
§2.6/§3): proves each one loads and actually runs against real data
produced by the real pipelines/mirrors, not just a synthetic table.
"""

import sqlite3
from pathlib import Path

from openpyxl import Workbook

from naru import mapping, mirror, query, runtime

REPO_ROOT = Path(__file__).parent.parent
RECIPES_DIR = REPO_ROOT / "recipes"
UST_ARTIFACT = REPO_ROOT / "pipelines" / "ust_auction_results" / "v1"
UST_FIXTURE = REPO_ROOT / "tests" / "fixtures" / "ust_lite.xlsx"


class TestAllRecipesLoad:
    def test_every_recipe_loads_without_error(self) -> None:
        recipes = query.list_recipes(RECIPES_DIR)
        assert len(recipes) == 3
        assert {r.name for r in recipes} == {
            "auction_tail",
            "weak_demand_auctions",
            "positions_by_counterparty",
        }


class TestUstRecipesAgainstRealPipeline:
    def _run_ust_pipeline(self, tmp_path: Path) -> Path:
        db_path = tmp_path / "naru.sqlite"
        runtime.run(
            artifact_path=UST_ARTIFACT,
            input_path=UST_FIXTURE,
            db_path=db_path,
            raw_dir=tmp_path / "raw",
        )
        return db_path

    def test_auction_tail_runs_against_real_output(self, tmp_path: Path) -> None:
        db_path = self._run_ust_pipeline(tmp_path)
        recipe = query.find_recipe(RECIPES_DIR, "auction_tail")
        conn = sqlite3.connect(db_path)
        result = query.run_recipe(conn, recipe, {"security": "4-Week Bill"})
        conn.close()

        assert result.columns == ["auction_date", "security_term", "high_yield", "yield_change"]
        assert len(result.rows) > 1
        assert all(row["security_term"] == "4-Week Bill" for row in result.rows)
        # first row for this security has no prior auction to diff against
        assert result.rows[0]["yield_change"] is None
        # every subsequent row must have a numeric change
        assert all(row["yield_change"] is not None for row in result.rows[1:])

    def test_weak_demand_auctions_runs_against_real_output(self, tmp_path: Path) -> None:
        db_path = self._run_ust_pipeline(tmp_path)
        recipe = query.find_recipe(RECIPES_DIR, "weak_demand_auctions")
        conn = sqlite3.connect(db_path)
        result = query.run_recipe(conn, recipe, {"max_bid_to_cover": "2.6"})
        conn.close()

        assert result.columns == ["auction_date", "security_term", "cusip", "bid_to_cover"]
        assert len(result.rows) > 0
        assert all(float(row["bid_to_cover"]) <= 2.6 for row in result.rows)  # type: ignore[arg-type]


class TestMirrorRecipeAgainstRealMirror:
    def test_positions_by_counterparty_runs_against_real_mirror_output(
        self, tmp_path: Path
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        ws.cell(row=1, column=1, value="Deal ID")
        ws.cell(row=1, column=2, value="Cpn (%)")
        ws.cell(row=1, column=3, value="As Of")
        for i, (deal_id, cpn) in enumerate([("D1", 2.5), ("D2", 3.25)], start=2):
            ws.cell(row=i, column=1, value=deal_id)
            ws.cell(row=i, column=2, value=cpn)
            ws.cell(row=i, column=3, value="2024-01-31")
        source_file = tmp_path / "client_statement.xlsx"
        wb.save(source_file)

        artifact_dir = tmp_path / "mapping_artifact"
        artifact_dir.mkdir()
        m = mapping.Mapping(
            target="warehouse.positions",
            key=["deal_id", "as_of"],
            on_duplicate="fail",
            columns=[
                mapping.ColumnMapping(
                    source="Deal ID", target="deal_id", transform="", basis="exact", approved=True
                ),
                mapping.ColumnMapping(
                    source="Cpn (%)",
                    target="coupon_rate",
                    transform="coerce_numeric(scale=0.01)",
                    basis="synonym",
                    approved=True,
                ),
                mapping.ColumnMapping(
                    source="As Of", target="as_of", transform="", basis="exact", approved=True
                ),
                mapping.ColumnMapping(
                    source="As Of",
                    target="counterparty",
                    transform="",
                    basis="llm",
                    evidence="test fixture: reuse As Of column to populate a required field",
                    approved=True,
                ),
            ],
            unmapped_source_columns="warn",
        )
        # deliberately reuse "As Of" as a stand-in counterparty value so this
        # fixture doesn't need a Broker column just to exercise the recipe.
        (artifact_dir / "mapping.yaml").write_text(mapping.to_yaml(m))
        (artifact_dir / "fingerprint.json").write_text(
            '{"sheet": "Statement", "header_row": 1, "columns": ['
            '{"name": "Deal ID", "type": "string", "strictness": "strict"}, '
            '{"name": "Cpn (%)", "type": "float", "strictness": "strict"}, '
            '{"name": "As Of", "type": "string", "strictness": "strict"}]}'
        )
        (artifact_dir / "schema.py").write_text(
            "from pydantic import BaseModel\n\n\n"
            "class TargetRow(BaseModel):\n"
            "    deal_id: str\n"
            "    coupon_rate: float\n"
            "    as_of: str\n"
            "    counterparty: str\n"
        )

        db_path = tmp_path / "naru.sqlite"
        mirror.mirror(artifact_dir, source_file, db_path, tmp_path / "raw", dry_run=False)

        recipe = query.find_recipe(RECIPES_DIR, "positions_by_counterparty")
        conn = sqlite3.connect(db_path)
        result = query.run_recipe(conn, recipe, {"counterparty": "2024-01-31"})
        conn.close()

        assert result.columns == ["deal_id", "coupon_rate", "as_of", "counterparty"]
        assert {row["deal_id"] for row in result.rows} == {"D1", "D2"}
