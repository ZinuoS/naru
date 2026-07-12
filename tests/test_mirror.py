"""Unit tests for src/naru/mirror.py."""

import datetime as dt
import json
import shutil
import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from naru import mapping, mirror
from naru.runtime import FingerprintDriftError

TARGET_ROW_SCHEMA = """\
from pydantic import BaseModel


class TargetRow(BaseModel):
    deal_id: str
    coupon_rate: float
    as_of: str
    trade_date: str | None = None
"""

FINGERPRINT = {
    "sheet": "Statement",
    "header_row": 1,
    "columns": [
        {"name": "Deal ID", "type": "string", "strictness": "strict"},
        {"name": "Cpn (%)", "type": "float", "strictness": "strict"},
        {"name": "As Of", "type": "string", "strictness": "strict"},
        {"name": "Notes", "type": "string", "strictness": "strict"},
    ],
}

VALID_MAPPING = mapping.Mapping(
    target="warehouse.positions",
    key=["deal_id", "as_of"],
    on_duplicate="fail",
    columns=[
        mapping.ColumnMapping(
            source="Deal ID", target="deal_id", transform="", basis="exact", approved=True
        ),
        mapping.ColumnMapping(
            source="Cpn (%)",
            target="coupon_rate",
            transform="coerce_numeric(scale=0.01)",
            basis="synonym",
            approved=True,
        ),
        mapping.ColumnMapping(
            source="As Of", target="as_of", transform="", basis="exact", approved=True
        ),
    ],
    unmapped_source_columns="warn",
)


def _write_source_file(path: Path, rows: list[tuple[str, float, str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.cell(row=1, column=1, value="Deal ID")
    ws.cell(row=1, column=2, value="Cpn (%)")
    ws.cell(row=1, column=3, value="As Of")
    ws.cell(row=1, column=4, value="Notes")
    for i, (deal_id, cpn, as_of, notes) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=deal_id)
        ws.cell(row=i, column=2, value=cpn)
        ws.cell(row=i, column=3, value=as_of)
        ws.cell(row=i, column=4, value=notes)
    wb.save(path)


def _write_mapping_artifact(
    root: Path,
    mapping_obj: mapping.Mapping = VALID_MAPPING,
    fingerprint: dict[str, object] = FINGERPRINT,
    target_row_schema: str = TARGET_ROW_SCHEMA,
) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    (root / "mapping.yaml").write_text(mapping.to_yaml(mapping_obj))
    (root / "fingerprint.json").write_text(json.dumps(fingerprint))
    (root / "schema.py").write_text(target_row_schema)
    return root


DEFAULT_ROWS: list[tuple[str, float, str, str]] = [
    ("D1", 2.5, "2024-01-01", "x"),
    ("D2", 3.25, "2024-01-01", "y"),
]


@pytest.fixture
def artifact_dir(tmp_path: Path) -> Path:
    return _write_mapping_artifact(tmp_path / "mapping_artifact")


@pytest.fixture
def source_file(tmp_path: Path) -> Path:
    path = tmp_path / "client_statement.xlsx"
    _write_source_file(path, DEFAULT_ROWS)
    return path


class TestTableNameForTarget:
    def test_dots_become_underscores(self) -> None:
        assert mirror._table_name_for_target("warehouse.positions") == "warehouse_positions"

    def test_no_dot_passes_through(self) -> None:
        assert mirror._table_name_for_target("positions") == "positions"


class TestMirrorDryRun:
    def test_writes_no_rows(self, artifact_dir: Path, source_file: Path, tmp_path: Path) -> None:
        db_path = tmp_path / "naru.sqlite"
        result = mirror.mirror(artifact_dir, source_file, db_path, tmp_path / "raw", dry_run=True)
        assert result.dry_run is True
        assert result.run_id is None
        assert result.row_ids == []

        conn = sqlite3.connect(db_path)
        count = conn.execute("SELECT COUNT(*) FROM warehouse_positions").fetchone()[0]
        assert count == 0
        conn.close()

    def test_summary_reports_rows_in_and_out(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        result = mirror.mirror(
            artifact_dir, source_file, tmp_path / "naru.sqlite", tmp_path / "raw", dry_run=True
        )
        assert result.summary.rows_in == 2
        assert result.summary.rows_out == 2

    def test_numeric_checks_reflect_pre_and_post_transform_sums(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        result = mirror.mirror(
            artifact_dir, source_file, tmp_path / "naru.sqlite", tmp_path / "raw", dry_run=True
        )
        checks = {c.target_column: c for c in result.summary.numeric_checks}
        assert checks["coupon_rate"].source_column == "Cpn (%)"
        assert checks["coupon_rate"].source_sum == pytest.approx(5.75)
        assert checks["coupon_rate"].target_sum == pytest.approx(0.0575)

    def test_unmapped_source_column_reported_under_warn_policy(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        result = mirror.mirror(
            artifact_dir, source_file, tmp_path / "naru.sqlite", tmp_path / "raw", dry_run=True
        )
        assert result.summary.unmapped_source_columns == ["Notes"]
        assert result.summary.unmapped_source_columns_action == "warn"

    def test_target_columns_not_populated_reported(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        result = mirror.mirror(
            artifact_dir, source_file, tmp_path / "naru.sqlite", tmp_path / "raw", dry_run=True
        )
        assert result.summary.target_columns_not_populated == ["trade_date"]

    def test_render_contains_key_facts_aligned(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        result = mirror.mirror(
            artifact_dir, source_file, tmp_path / "naru.sqlite", tmp_path / "raw", dry_run=True
        )
        text = result.summary.render()
        assert "DRY RUN" in text
        assert "Rows in file   : 2" in text
        assert "trade_date" in text
        assert '"Notes"' in text


class TestReconciliationSummaryRenderEmptyCases:
    def test_render_shows_placeholders_when_nothing_to_report(self) -> None:
        summary = mirror.ReconciliationSummary(
            source_file="f.xlsx",
            target="t",
            dry_run=True,
            rows_in=1,
            rows_out=1,
            numeric_checks=[],
            unmapped_source_columns=[],
            unmapped_source_columns_action="warn",
            target_columns_not_populated=[],
        )
        text = summary.render()
        assert "(no numeric target columns in this mapping)" in text
        assert "(none -- every source column is mapped)" in text
        assert "(none -- every target column is populated)" in text


class TestMirrorCommit:
    def test_writes_rows_with_lineage(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        db_path = tmp_path / "naru.sqlite"
        raw_dir = tmp_path / "raw"
        result = mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=False)
        assert result.dry_run is False
        assert result.run_id is not None
        assert result.row_ids == [1, 2]

        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT deal_id, coupon_rate, as_of, trade_date FROM warehouse_positions "
            "ORDER BY deal_id"
        ).fetchall()
        assert [dict(r) for r in rows] == [
            {"deal_id": "D1", "coupon_rate": 0.025, "as_of": "2024-01-01", "trade_date": None},
            {"deal_id": "D2", "coupon_rate": 0.0325, "as_of": "2024-01-01", "trade_date": None},
        ]

        lineage = conn.execute(
            "SELECT final_table, source_row_start, source_row_end, pipeline_version "
            "FROM meta_lineage ORDER BY row_id"
        ).fetchall()
        assert tuple(lineage[0]) == ("warehouse_positions", 2, 2, "mirror")
        assert tuple(lineage[1]) == ("warehouse_positions", 3, 3, "mirror")
        conn.close()

    def test_registers_raw_file_with_correct_hash(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        import hashlib

        db_path = tmp_path / "naru.sqlite"
        raw_dir = tmp_path / "raw"
        mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=False)
        expected_sha256 = hashlib.sha256(source_file.read_bytes()).hexdigest()
        assert (raw_dir / expected_sha256).exists()

        conn = sqlite3.connect(db_path)
        row = conn.execute(
            "SELECT sha256, original_name FROM raw_files WHERE sha256 = ?", (expected_sha256,)
        ).fetchone()
        assert row == (expected_sha256, "client_statement.xlsx")
        conn.close()


class TestMirrorDuplicateKey:
    def test_recommitting_same_file_aborts_on_existing_keys(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        db_path = tmp_path / "naru.sqlite"
        raw_dir = tmp_path / "raw"
        mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=False)

        with pytest.raises(mirror.MirrorDuplicateKeyError) as exc_info:
            mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=False)
        exc = exc_info.value
        assert sorted(exc.colliding_with_existing) == [
            ("D1", "2024-01-01"),
            ("D2", "2024-01-01"),
        ]
        assert exc.colliding_within_batch == []

    def test_duplicate_abort_writes_nothing_additional(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        db_path = tmp_path / "naru.sqlite"
        raw_dir = tmp_path / "raw"
        mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=False)
        with pytest.raises(mirror.MirrorDuplicateKeyError):
            mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=False)

        conn = sqlite3.connect(db_path)
        count = conn.execute("SELECT COUNT(*) FROM warehouse_positions").fetchone()[0]
        assert count == 2
        conn.close()

    def test_duplicate_keys_within_the_same_batch_abort(
        self, artifact_dir: Path, tmp_path: Path
    ) -> None:
        source_path = tmp_path / "dup_within_batch.xlsx"
        _write_source_file(
            source_path,
            [("D1", 2.5, "2024-01-01", "x"), ("D1", 9.75, "2024-01-01", "z")],
        )
        with pytest.raises(mirror.MirrorDuplicateKeyError) as exc_info:
            mirror.mirror(
                artifact_dir,
                source_path,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=False,
            )
        exc = exc_info.value
        assert exc.colliding_with_existing == []
        assert exc.colliding_within_batch == [("D1", "2024-01-01")]

    def test_duplicate_check_also_runs_in_dry_run_mode(
        self, artifact_dir: Path, source_file: Path, tmp_path: Path
    ) -> None:
        db_path = tmp_path / "naru.sqlite"
        raw_dir = tmp_path / "raw"
        mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=False)
        with pytest.raises(mirror.MirrorDuplicateKeyError):
            mirror.mirror(artifact_dir, source_file, db_path, raw_dir, dry_run=True)


class TestMirrorFingerprintDrift:
    def test_renamed_header_column_raises_fingerprint_drift(self, tmp_path: Path) -> None:
        artifact_dir = _write_mapping_artifact(tmp_path / "mapping_artifact")
        source_path = tmp_path / "drifted.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        ws.cell(row=1, column=1, value="Deal ID")
        ws.cell(row=1, column=2, value="Coupon Rate")  # renamed from "Cpn (%)"
        ws.cell(row=1, column=3, value="As Of")
        ws.cell(row=1, column=4, value="Notes")
        ws.cell(row=2, column=1, value="D1")
        ws.cell(row=2, column=2, value=2.5)
        ws.cell(row=2, column=3, value="2024-01-01")
        ws.cell(row=2, column=4, value="x")
        wb.save(source_path)

        with pytest.raises(FingerprintDriftError, match="Coupon Rate"):
            mirror.mirror(
                artifact_dir,
                source_path,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=True,
            )


class TestMirrorUnmappedSourceColumnsFailPolicy:
    def test_fail_policy_aborts_before_writing(self, source_file: Path, tmp_path: Path) -> None:
        strict_mapping = VALID_MAPPING.model_copy(update={"unmapped_source_columns": "fail"})
        artifact_dir = _write_mapping_artifact(
            tmp_path / "mapping_artifact", mapping_obj=strict_mapping
        )
        with pytest.raises(mirror.MirrorError, match="Notes"):
            mirror.mirror(
                artifact_dir,
                source_file,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=True,
            )


class TestMirrorMissingMappedColumn:
    def test_mapped_source_column_absent_from_file_raises(self, tmp_path: Path) -> None:
        bad_mapping = mapping.Mapping(
            target="warehouse.positions",
            key=["deal_id"],
            on_duplicate="fail",
            columns=[
                mapping.ColumnMapping(
                    source="Deal Identifier",  # doesn't exist in the file
                    target="deal_id",
                    transform="",
                    basis="exact",
                    approved=True,
                ),
            ],
            unmapped_source_columns="warn",
        )
        artifact_dir = _write_mapping_artifact(
            tmp_path / "mapping_artifact", mapping_obj=bad_mapping
        )
        source_path = tmp_path / "client_statement.xlsx"
        _write_source_file(source_path, DEFAULT_ROWS)

        with pytest.raises(mirror.MirrorError, match="Deal Identifier"):
            mirror.mirror(
                artifact_dir,
                source_path,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=True,
            )


WAREHOUSE_FIXTURE = Path(__file__).parent / "fixtures" / "warehouse_workbook.xlsx"
FIXED_NOW = dt.datetime(2026, 1, 1, 12, 0, 0, tzinfo=dt.UTC)

EXCEL_SOURCE_FINGERPRINT = {
    "sheet": "Statement",
    "header_row": 1,
    "columns": [
        {"name": "Deal ID", "type": "string", "strictness": "strict"},
        {"name": "Cpn (%)", "type": "float", "strictness": "strict"},
        {"name": "As Of", "type": "string", "strictness": "strict"},
        {"name": "Counterparty", "type": "string", "strictness": "strict"},
    ],
}

WAREHOUSE_FINGERPRINT = {
    "sheet": "Positions",
    "header_row": 1,
    "columns": [
        {"name": "Deal ID", "type": "string", "strictness": "strict"},
        {"name": "Coupon Rate", "type": "float", "strictness": "strict"},
        {"name": "As Of", "type": "string", "strictness": "strict"},
        {"name": "Counterparty", "type": "string", "strictness": "strict"},
    ],
}

EXCEL_TARGET_ROW_SCHEMA = """\
from pydantic import BaseModel


class TargetRow(BaseModel):
    deal_id: str
    coupon_rate: float
    as_of: str
    counterparty: str
"""

EXCEL_MAPPING = mapping.Mapping(
    target="warehouse.positions",
    key=["deal_id", "as_of"],
    on_duplicate="fail",
    columns=[
        mapping.ColumnMapping(
            source="Deal ID", target="deal_id", transform="", basis="exact", approved=True
        ),
        mapping.ColumnMapping(
            source="Cpn (%)",
            target="coupon_rate",
            transform="coerce_numeric(scale=0.01)",
            basis="synonym",
            approved=True,
        ),
        mapping.ColumnMapping(
            source="As Of", target="as_of", transform="", basis="exact", approved=True
        ),
        mapping.ColumnMapping(
            source="Counterparty",
            target="counterparty",
            transform="",
            basis="exact",
            approved=True,
        ),
    ],
    unmapped_source_columns="warn",
    excel_target=mapping.ExcelTarget(
        path="warehouse_workbook.xlsx",
        first_data_col="A",
        last_data_col="D",
        column_order=["deal_id", "coupon_rate", "as_of", "counterparty"],
    ),
)


def _write_excel_source_file(path: Path, rows: list[tuple[str, float, str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.cell(row=1, column=1, value="Deal ID")
    ws.cell(row=1, column=2, value="Cpn (%)")
    ws.cell(row=1, column=3, value="As Of")
    ws.cell(row=1, column=4, value="Counterparty")
    for i, (deal_id, cpn, as_of, counterparty) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=deal_id)
        ws.cell(row=i, column=2, value=cpn)
        ws.cell(row=i, column=3, value=as_of)
        ws.cell(row=i, column=4, value=counterparty)
    wb.save(path)


def _write_excel_mapping_artifact(
    root: Path,
    mapping_obj: mapping.Mapping = EXCEL_MAPPING,
    warehouse_fingerprint: dict[str, object] = WAREHOUSE_FINGERPRINT,
) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    (root / "mapping.yaml").write_text(mapping.to_yaml(mapping_obj))
    (root / "fingerprint.json").write_text(json.dumps(EXCEL_SOURCE_FINGERPRINT))
    (root / "schema.py").write_text(EXCEL_TARGET_ROW_SCHEMA)
    (root / "warehouse_fingerprint.json").write_text(json.dumps(warehouse_fingerprint))
    shutil.copy2(WAREHOUSE_FIXTURE, root / "warehouse_workbook.xlsx")
    return root


EXCEL_DEFAULT_ROWS: list[tuple[str, float, str, str]] = [
    ("D100", 2.75, "2024-01-31", "Gamma Securities"),
    ("D101", 3.55, "2024-01-31", "Alpha Bank"),
]


@pytest.fixture
def excel_artifact_dir(tmp_path: Path) -> Path:
    return _write_excel_mapping_artifact(tmp_path / "mapping_artifact")


@pytest.fixture
def excel_source_file(tmp_path: Path) -> Path:
    path = tmp_path / "client_statement.xlsx"
    _write_excel_source_file(path, EXCEL_DEFAULT_ROWS)
    return path


class TestMirrorExcelDryRun:
    def test_leaves_warehouse_file_byte_identical(
        self, excel_artifact_dir: Path, excel_source_file: Path, tmp_path: Path
    ) -> None:
        warehouse_path = excel_artifact_dir / "warehouse_workbook.xlsx"
        before = warehouse_path.read_bytes()
        result = mirror.mirror(
            excel_artifact_dir,
            excel_source_file,
            tmp_path / "naru.sqlite",
            tmp_path / "raw",
            dry_run=True,
        )
        assert result.backup_path is None
        assert warehouse_path.read_bytes() == before

    def test_summary_reports_target_as_file_path(
        self, excel_artifact_dir: Path, excel_source_file: Path, tmp_path: Path
    ) -> None:
        result = mirror.mirror(
            excel_artifact_dir,
            excel_source_file,
            tmp_path / "naru.sqlite",
            tmp_path / "raw",
            dry_run=True,
        )
        assert result.summary.target == "warehouse_workbook.xlsx"
        assert result.summary.rows_in == 2
        assert result.summary.rows_out == 2
        assert result.summary.unmapped_source_columns == []
        assert result.summary.target_columns_not_populated == []


class TestMirrorExcelCommit:
    def test_creates_timestamped_backup_matching_pristine_file(
        self, excel_artifact_dir: Path, excel_source_file: Path, tmp_path: Path
    ) -> None:
        result = mirror.mirror(
            excel_artifact_dir,
            excel_source_file,
            tmp_path / "naru.sqlite",
            tmp_path / "raw",
            dry_run=False,
            now=FIXED_NOW,
        )
        assert result.backup_path is not None
        assert result.backup_path.name == "warehouse_workbook.20260101T120000Z.bak.xlsx"
        assert result.backup_path.exists()
        assert result.backup_path.read_bytes() == WAREHOUSE_FIXTURE.read_bytes()

    def test_appends_new_rows_at_correct_position_only_in_region(
        self, excel_artifact_dir: Path, excel_source_file: Path, tmp_path: Path
    ) -> None:
        mirror.mirror(
            excel_artifact_dir,
            excel_source_file,
            tmp_path / "naru.sqlite",
            tmp_path / "raw",
            dry_run=False,
            now=FIXED_NOW,
        )
        wb = load_workbook(excel_artifact_dir / "warehouse_workbook.xlsx")
        ws = wb["Positions"]
        assert ws.max_row == 5
        assert [ws.cell(row=4, column=c).value for c in range(1, 5)] == [
            "D100",
            0.0275,
            "2024-01-31",
            "Gamma Securities",
        ]
        assert [ws.cell(row=5, column=c).value for c in range(1, 5)] == [
            "D101",
            0.0355,
            "2024-01-31",
            "Alpha Bank",
        ]
        # columns outside the declared region (E: Notional, F: Total)
        # must never be touched for the new rows.
        assert ws.cell(row=4, column=5).value is None
        assert ws.cell(row=4, column=6).value is None
        assert ws.cell(row=5, column=5).value is None
        assert ws.cell(row=5, column=6).value is None

    def test_stray_value_outside_region_in_a_trailing_row_does_not_fool_append_position(
        self, excel_artifact_dir: Path, tmp_path: Path
    ) -> None:
        # A row with a value only outside the declared region (column F,
        # outside A:D) must not be mistaken for existing region data --
        # the append position is judged by the region's own columns only.
        warehouse_path = excel_artifact_dir / "warehouse_workbook.xlsx"
        wb = load_workbook(warehouse_path)
        wb["Positions"]["F4"] = "stray leftover value"
        wb.save(warehouse_path)

        source_path = tmp_path / "client_statement.xlsx"
        _write_excel_source_file(source_path, [("D300", 4.5, "2024-02-29", "Delta Partners")])
        mirror.mirror(
            excel_artifact_dir,
            source_path,
            tmp_path / "naru.sqlite",
            tmp_path / "raw",
            dry_run=False,
            now=FIXED_NOW,
        )

        result_wb = load_workbook(warehouse_path)
        result_ws = result_wb["Positions"]
        assert [result_ws.cell(row=4, column=c).value for c in range(1, 5)] == [
            "D300",
            0.045,
            "2024-02-29",
            "Delta Partners",
        ]
        # the pre-existing stray value outside the region must survive.
        assert result_ws.cell(row=4, column=6).value == "stray leftover value"

    def test_preserves_every_pre_existing_cell_value_formula_and_number_format(
        self, excel_artifact_dir: Path, excel_source_file: Path, tmp_path: Path
    ) -> None:
        """THE test: reads back every pre-existing cell via openpyxl after
        a commit-mode mirror and asserts values, formulas, and number
        formats are unchanged. Deliberately does NOT compare raw file
        bytes -- the container isn't stable across an openpyxl load/save
        round trip even when nothing meaningful changed.
        """
        before_wb = load_workbook(WAREHOUSE_FIXTURE)  # formulas as formula strings
        before_ws = before_wb["Positions"]
        before_cells = {
            (row, col): (
                before_ws.cell(row=row, column=col).value,
                before_ws.cell(row=row, column=col).number_format,
            )
            for row in range(1, before_ws.max_row + 1)
            for col in range(1, before_ws.max_column + 1)
        }
        before_notes_ws = before_wb["Notes"]
        before_notes_cells = {
            (row, col): before_notes_ws.cell(row=row, column=col).value
            for row in range(1, before_notes_ws.max_row + 1)
            for col in range(1, before_notes_ws.max_column + 1)
        }
        before_cf = [
            (str(cf.sqref), [(r.type, r.operator, r.formula) for r in cf.rules])
            for cf in before_ws.conditional_formatting
        ]
        before_names = {name: dn.attr_text for name, dn in before_wb.defined_names.items()}
        before_sheetnames = before_wb.sheetnames

        mirror.mirror(
            excel_artifact_dir,
            excel_source_file,
            tmp_path / "naru.sqlite",
            tmp_path / "raw",
            dry_run=False,
            now=FIXED_NOW,
        )

        after_wb = load_workbook(excel_artifact_dir / "warehouse_workbook.xlsx")
        after_ws = after_wb["Positions"]
        for (row, col), (value, number_format) in before_cells.items():
            assert after_ws.cell(row=row, column=col).value == value, f"cell ({row},{col}) value"
            assert after_ws.cell(row=row, column=col).number_format == number_format, (
                f"cell ({row},{col}) number_format"
            )

        after_notes_ws = after_wb["Notes"]
        for (row, col), value in before_notes_cells.items():
            assert after_notes_ws.cell(row=row, column=col).value == value

        after_cf = [
            (str(cf.sqref), [(r.type, r.operator, r.formula) for r in cf.rules])
            for cf in after_ws.conditional_formatting
        ]
        assert after_cf == before_cf

        after_names = {name: dn.attr_text for name, dn in after_wb.defined_names.items()}
        assert after_names == before_names

        assert after_wb.sheetnames == before_sheetnames


class TestMirrorExcelDuplicateKey:
    def test_key_colliding_with_existing_warehouse_row_aborts(
        self, excel_artifact_dir: Path, tmp_path: Path
    ) -> None:
        # W1/2023-12-31 is one of warehouse_workbook.xlsx's existing rows.
        source_path = tmp_path / "client_statement.xlsx"
        _write_excel_source_file(source_path, [("W1", 5.25, "2023-12-31", "Someone")])
        with pytest.raises(mirror.MirrorDuplicateKeyError) as exc_info:
            mirror.mirror(
                excel_artifact_dir,
                source_path,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=False,
                now=FIXED_NOW,
            )
        assert exc_info.value.colliding_with_existing == [("W1", "2023-12-31")]

    def test_duplicate_within_batch_aborts(self, excel_artifact_dir: Path, tmp_path: Path) -> None:
        source_path = tmp_path / "client_statement.xlsx"
        _write_excel_source_file(
            source_path,
            [
                ("D200", 2.25, "2024-01-31", "X"),
                ("D200", 9.75, "2024-01-31", "Y"),
            ],
        )
        with pytest.raises(mirror.MirrorDuplicateKeyError) as exc_info:
            mirror.mirror(
                excel_artifact_dir,
                source_path,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=False,
                now=FIXED_NOW,
            )
        assert exc_info.value.colliding_within_batch == [("D200", "2024-01-31")]

    def test_duplicate_abort_writes_no_backup(
        self, excel_artifact_dir: Path, tmp_path: Path
    ) -> None:
        source_path = tmp_path / "client_statement.xlsx"
        _write_excel_source_file(source_path, [("W1", 5.25, "2023-12-31", "Someone")])
        with pytest.raises(mirror.MirrorDuplicateKeyError):
            mirror.mirror(
                excel_artifact_dir,
                source_path,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=False,
                now=FIXED_NOW,
            )
        backups = list(excel_artifact_dir.glob("*.bak.xlsx"))
        assert backups == []


class TestMirrorExcelWarehouseFingerprintDrift:
    def test_renamed_warehouse_header_raises_fingerprint_drift(
        self, excel_source_file: Path, tmp_path: Path
    ) -> None:
        root = _write_excel_mapping_artifact(tmp_path / "mapping_artifact")
        warehouse_path = root / "warehouse_workbook.xlsx"
        wb = load_workbook(warehouse_path)
        wb["Positions"]["B1"] = "Coupon %"  # renamed from "Coupon Rate"
        wb.save(warehouse_path)

        with pytest.raises(FingerprintDriftError, match="Coupon"):
            mirror.mirror(
                root,
                excel_source_file,
                tmp_path / "naru.sqlite",
                tmp_path / "raw",
                dry_run=True,
            )
