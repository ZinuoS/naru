"""Unit tests for src/naru/profiler.py."""

import datetime as dt
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from naru import profiler

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"
GOLDEN_PROFILES_DIR = Path(__file__).resolve().parent / "golden" / "profiles"

ALL_FIXTURE_NAMES = [
    "ust_lite",
    "h1_merged_headers",
    "h2_two_tables",
    "h3_date_ambiguity",
    "h4_units_in_header",
    "h5_buried_header",
]


class TestSheetDimensions:
    def test_ust_lite_dimensions(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "ust_lite.xlsx")
        dims, n_rows, n_cols = profiler.sheet_dimensions(wb["Results"])
        assert dims == "A1:G46"
        assert n_rows == 46
        assert n_cols == 7

    def test_h5_buried_header_dimensions(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h5_buried_header.xlsx")
        dims, n_rows, n_cols = profiler.sheet_dimensions(wb["Positions"])
        assert dims == "A1:D39"
        assert n_rows == 39
        assert n_cols == 4


class TestMergedCellRanges:
    def test_h1_merged_headers_has_two_ranges(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h1_merged_headers.xlsx")
        result = profiler.merged_cell_ranges(wb["Trades"])
        assert result == ["B1:C1", "D1:E1"]

    def test_ust_lite_has_one_range(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "ust_lite.xlsx")
        result = profiler.merged_cell_ranges(wb["Results"])
        assert result == ["F3:G3"]

    def test_no_merged_cells_returns_empty_list(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h3_date_ambiguity.xlsx")
        result = profiler.merged_cell_ranges(wb["Trades"])
        assert result == []


class TestCellType:
    def test_classifies_each_coarse_type(self) -> None:
        assert profiler.cell_type(None) == "empty"
        assert profiler.cell_type(True) == "boolean"
        assert profiler.cell_type(5) == "integer"
        assert profiler.cell_type(5.5) == "float"
        assert profiler.cell_type(dt.date(2020, 1, 1)) == "date"
        assert profiler.cell_type(dt.datetime(2020, 1, 1, 12, 0)) == "date"
        assert profiler.cell_type("x") == "string"

    def test_bool_is_not_classified_as_integer(self) -> None:
        # bool is a subclass of int in Python; must be checked first.
        assert profiler.cell_type(False) == "boolean"

    def test_unrecognized_type_falls_through_to_other(self) -> None:
        assert profiler.cell_type(object()) == "other"


class TestNonStringFraction:
    def test_all_empty_row_returns_zero(self) -> None:
        assert profiler._non_string_fraction({"empty": 3}, n_cols=3) == 0.0

    def test_mixed_row(self) -> None:
        assert profiler._non_string_fraction({"string": 1, "integer": 1}, n_cols=2) == 0.5


class TestDetectHeaderRows:
    def test_simple_header_above_data(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Amount")
        ws.cell(row=2, column=1, value="a")
        ws.cell(row=2, column=2, value=1.5)
        result = profiler.detect_header_rows(ws)
        assert result == [profiler.HeaderCandidate(row=1, confidence=0.75)]

    def test_empty_sheet_returns_no_candidates(self) -> None:
        wb = Workbook()
        result = profiler.detect_header_rows(wb.active)
        assert result == []

    def test_single_row_sheet_returns_no_candidates(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="only header, no data below")
        result = profiler.detect_header_rows(ws)
        assert result == []

    def test_no_row_ever_qualifies_returns_no_candidates(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=1)
        ws.cell(row=1, column=2, value=2)
        ws.cell(row=2, column=1, value=3)
        ws.cell(row=2, column=2, value=4)
        result = profiler.detect_header_rows(ws)
        assert result == []

    def test_sparse_row_below_density_threshold_is_skipped(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="banner text only")
        ws.cell(row=2, column=1, value="Name")
        ws.cell(row=2, column=2, value="Amount")
        ws.cell(row=2, column=3, value="Date")
        ws.cell(row=3, column=1, value="a")
        ws.cell(row=3, column=2, value=1.5)
        ws.cell(row=3, column=3, value="b")
        result = profiler.detect_header_rows(ws)
        # row 1 is only 1/3 columns populated (below the density gate), so
        # it's skipped even though the lone cell is text.
        assert result == [profiler.HeaderCandidate(row=2, confidence=0.6667)]

    def test_ust_lite_finds_row_3_past_two_banner_rows(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "ust_lite.xlsx")
        result = profiler.detect_header_rows(wb["Results"])
        assert len(result) == 1
        assert result[0].row == 3

    def test_h1_merged_headers_finds_row_2_not_the_merged_group_row(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h1_merged_headers.xlsx")
        result = profiler.detect_header_rows(wb["Trades"])
        assert len(result) == 1
        assert result[0].row == 2

    def test_h2_two_tables_finds_only_the_first_tables_header(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h2_two_tables.xlsx")
        result = profiler.detect_header_rows(wb["Book"])
        assert len(result) == 1
        assert result[0].row == 1

    def test_h3_date_ambiguity_finds_row_1(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h3_date_ambiguity.xlsx")
        result = profiler.detect_header_rows(wb["Trades"])
        assert len(result) == 1
        assert result[0].row == 1

    def test_h4_units_in_header_finds_row_1(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h4_units_in_header.xlsx")
        result = profiler.detect_header_rows(wb["Blotter"])
        assert len(result) == 1
        assert result[0].row == 1

    def test_h5_buried_header_finds_row_7_past_six_banner_rows(self) -> None:
        wb = load_workbook(FIXTURES_DIR / "h5_buried_header.xlsx")
        result = profiler.detect_header_rows(wb["Positions"])
        assert len(result) == 1
        assert result[0].row == 7


class TestDetectColumnSmells:
    def test_percent_string(self) -> None:
        assert profiler._detect_column_smells(["2.5%"]) == ["percent_string"]

    def test_thousands_separator(self) -> None:
        assert profiler._detect_column_smells(["38,000"]) == ["thousands_separator"]

    def test_parens_negative(self) -> None:
        assert profiler._detect_column_smells(["(1,234)"]) == ["parens_negative"]

    def test_date_serial_suspect(self) -> None:
        assert profiler._detect_column_smells([43480]) == ["date_serial_suspect"]

    def test_bool_is_not_flagged_as_date_serial(self) -> None:
        assert profiler._detect_column_smells([True]) == []

    def test_plain_values_have_no_smells(self) -> None:
        assert profiler._detect_column_smells(["hello", 5, 5.5]) == []

    def test_integer_outside_serial_range_not_flagged(self) -> None:
        assert profiler._detect_column_smells([5]) == []

    def test_multiple_smells_combine_sorted(self) -> None:
        result = profiler._detect_column_smells(["2.5%", "38,000", "(1,234)", 43480])
        assert result == [
            "date_serial_suspect",
            "parens_negative",
            "percent_string",
            "thousands_separator",
        ]


class TestInferColumnType:
    def test_single_type(self) -> None:
        assert profiler._infer_column_type([1, 2, None]) == "integer"

    def test_mixed_types(self) -> None:
        assert profiler._infer_column_type([1, "a"]) == "mixed"

    def test_all_null(self) -> None:
        assert profiler._infer_column_type([None, None]) == "empty"


class TestProfileColumns:
    def test_basic_column_stats(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=2, column=1, value="a")
        ws.cell(row=3, column=1, value="a")
        ws.cell(row=4, column=1, value=None)
        columns = profiler.profile_columns(ws, header_row=1)
        assert len(columns) == 1
        col = columns[0]
        assert col.header_text == "Name"
        assert col.inferred_type == "string"
        assert col.null_rate == pytest.approx(1 / 3, abs=1e-4)
        assert col.cardinality == 1
        assert col.samples == ["a"]

    def test_samples_capped_at_five_in_first_seen_order(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="X")
        for i, v in enumerate(["f", "e", "d", "c", "b", "a"]):
            ws.cell(row=2 + i, column=1, value=v)
        columns = profiler.profile_columns(ws, header_row=1)
        assert columns[0].samples == ["f", "e", "d", "c", "b"]
        assert columns[0].cardinality == 6

    def test_no_header_row_data_gives_none_header_text(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=None)
        ws.cell(row=2, column=1, value="a")
        columns = profiler.profile_columns(ws, header_row=1)
        assert columns[0].header_text is None

    def test_smells_detected_on_column(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Amt")
        ws.cell(row=2, column=1, value="1,000")
        ws.cell(row=3, column=1, value="2,000")
        columns = profiler.profile_columns(ws, header_row=1)
        assert columns[0].smells == ["thousands_separator"]


class TestDetectDuplicateHeaders:
    def test_finds_matching_pair(self) -> None:
        result = profiler.detect_duplicate_headers({"A": ["x", "y"], "B": ["x", "y"], "C": ["z"]})
        assert result == [("A", "B")]

    def test_no_duplicates_returns_empty(self) -> None:
        result = profiler.detect_duplicate_headers({"A": ["x"], "B": ["y"]})
        assert result == []

    def test_three_way_duplicate_reports_all_pairs(self) -> None:
        result = profiler.detect_duplicate_headers({"A": ["x"], "B": ["x"], "C": ["x"]})
        assert result == [("A", "B"), ("A", "C"), ("B", "C")]


class TestProfile:
    def test_profiles_ust_lite(self) -> None:
        result = profiler.profile(FIXTURES_DIR / "ust_lite.xlsx")
        assert result.source_file == "ust_lite.xlsx"
        assert len(result.sheets) == 1
        sheet = result.sheets[0]
        assert sheet.name == "Results"
        assert sheet.header_candidates[0].row == 3
        assert len(sheet.columns) == 7
        assert sheet.merged_cells == ["F3:G3"]
        assert result.duplicate_headers == []

    def test_profiles_h5_buried_header(self) -> None:
        result = profiler.profile(FIXTURES_DIR / "h5_buried_header.xlsx")
        sheet = result.sheets[0]
        assert sheet.header_candidates[0].row == 7
        assert [c.header_text for c in sheet.columns] == [
            "Position ID",
            "Desk",
            "Book",
            "Market Value",
        ]

    def test_sheet_with_no_detectable_header_gets_empty_columns(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=1)
        ws.cell(row=1, column=2, value=2)
        ws.cell(row=2, column=1, value=3)
        ws.cell(row=2, column=2, value=4)
        path = tmp_path / "no_header.xlsx"
        wb.save(path)
        result = profiler.profile(path)
        assert result.sheets[0].header_candidates == []
        assert result.sheets[0].columns == []
        assert result.duplicate_headers == []


class TestToJson:
    def test_output_is_valid_sorted_json(self) -> None:
        result = profiler.profile(FIXTURES_DIR / "h3_date_ambiguity.xlsx")
        text = profiler.to_json(result)
        parsed = json.loads(text)
        assert parsed["source_file"] == "h3_date_ambiguity.xlsx"
        # sorted-key requirement: re-dumping with sort_keys must be a no-op
        assert json.dumps(parsed, sort_keys=True, indent=2) == text

    @pytest.mark.parametrize("fixture_name", ALL_FIXTURE_NAMES)
    def test_two_runs_produce_byte_identical_json(self, fixture_name: str) -> None:
        path = FIXTURES_DIR / f"{fixture_name}.xlsx"
        first = profiler.to_json(profiler.profile(path))
        second = profiler.to_json(profiler.profile(path))
        assert first == second


class TestGoldenProfiles:
    @pytest.mark.parametrize("fixture_name", ALL_FIXTURE_NAMES)
    def test_profile_matches_frozen_golden(self, fixture_name: str) -> None:
        actual = profiler.to_json(profiler.profile(FIXTURES_DIR / f"{fixture_name}.xlsx"))
        expected = (GOLDEN_PROFILES_DIR / f"{fixture_name}.json").read_text()
        assert actual == expected
