"""Unit tests for src/naru/__main__.py."""

import json
import sqlite3
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from naru import __main__ as cli
from naru.fingerprint import Difference, FingerprintCheckResult
from naru.runtime import FingerprintDriftError

VALID_MANIFEST = """\
name: test_pipeline
version: v1
sheet: Sheet1
target_table: final_test
key: [id]
"""

VALID_FINGERPRINT = """\
{
  "sheet": "Sheet1",
  "header_row": 1,
  "columns": [
    {"name": "id", "type": "integer", "strictness": "strict"},
    {"name": "label", "type": "string", "strictness": "strict"}
  ]
}
"""

VALID_SCHEMA = """\
from pydantic import BaseModel


class SourceRow(BaseModel):
    id: int
    label: str


class TargetRow(BaseModel):
    id: int
    label: str
"""

VALID_TRANSFORM = """\
def transform(df):
    return ops.promote_header(df, header_row=1, column_names=["id", "label"])
"""

VALID_VALIDATIONS = "row_count:\n  min: 1\n"
VALID_CHANGELOG = "# Changelog\n\nv1: initial.\n"


def _make_input_workbook(
    rows: list[tuple[object, object]], header: tuple[str, str] = ("id", "label")
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=header[0])
    ws.cell(row=1, column=2, value=header[1])
    for i, (id_val, label_val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=id_val)
        ws.cell(row=i, column=2, value=label_val)
    return wb


def _write_artifact(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    (root / "manifest.yaml").write_text(VALID_MANIFEST)
    (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (root / "schema.py").write_text(VALID_SCHEMA)
    (root / "transform.py").write_text(VALID_TRANSFORM)
    (root / "validations.yaml").write_text(VALID_VALIDATIONS)
    (root / "CHANGELOG.md").write_text(VALID_CHANGELOG)
    golden = root / "golden"
    golden.mkdir(exist_ok=True)
    _make_input_workbook([(1, "a"), (2, "b")]).save(golden / "input_sample.xlsx")
    pd.DataFrame({"id": [1, 2], "label": ["a", "b"]}).to_parquet(
        golden / "expected_output.parquet", index=False
    )


@pytest.fixture
def artifact_dir(tmp_path: Path) -> Path:
    root = tmp_path / "artifact"
    _write_artifact(root)
    return root


@pytest.fixture(autouse=True)
def _isolated_cwd(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """main() writes naru.sqlite/.naru/raw/drift_report.json relative to
    CWD -- always run from an isolated directory, never the repo root.
    """
    work_dir = tmp_path / "cwd"
    work_dir.mkdir()
    monkeypatch.chdir(work_dir)


class TestDriftReport:
    def test_structure_names_file_and_differences(self, tmp_path: Path) -> None:
        result = FingerprintCheckResult(
            ok=False,
            differences=[
                Difference("header_text_mismatch", "Sheet1", "id", "identifier", 1),
            ],
        )
        exc = FingerprintDriftError(result)
        report = cli._drift_report(Path("some/artifact"), Path("some/input.xlsx"), exc)
        assert report["artifact"] == "some/artifact"
        assert report["input_file"] == "some/input.xlsx"
        assert report["differences"] == [
            {
                "kind": "header_text_mismatch",
                "sheet": "Sheet1",
                "column_position": 1,
                "expected": "id",
                "found": "identifier",
                "message": "sheet 'Sheet1' col 1: expected `id`, found `identifier`",
            }
        ]


class TestCmdRun:
    def test_success_returns_zero_and_loads_db(self, artifact_dir: Path, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a"), (2, "b")]).save(input_path)

        exit_code = cli.main(["run", str(artifact_dir), str(input_path)])

        assert exit_code == 0
        assert cli.DB_PATH.exists()
        conn = sqlite3.connect(cli.DB_PATH)
        rows = conn.execute("SELECT id, label FROM final_test ORDER BY id").fetchall()
        assert rows == [(1, "a"), (2, "b")]

    def test_as_of_flag_is_stored_exactly(self, artifact_dir: Path, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(input_path)

        exit_code = cli.main(["run", str(artifact_dir), str(input_path), "--as-of", "2020-06-01"])

        assert exit_code == 0
        conn = sqlite3.connect(cli.DB_PATH)
        (as_of,) = conn.execute("SELECT as_of FROM meta_runs").fetchone()
        assert as_of == "2020-06-01"

    def test_fingerprint_drift_returns_three_and_writes_report(
        self, artifact_dir: Path, tmp_path: Path
    ) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")], header=("renamed_id", "label")).save(input_path)

        exit_code = cli.main(["run", str(artifact_dir), str(input_path)])

        assert exit_code == 3
        assert cli.DRIFT_REPORT_PATH.exists()
        report = json.loads(cli.DRIFT_REPORT_PATH.read_text())
        assert report["artifact"] == str(artifact_dir)
        assert report["input_file"] == str(input_path)
        kinds = {d["kind"] for d in report["differences"]}
        assert "header_text_mismatch" in kinds
        mismatch = next(d for d in report["differences"] if d["kind"] == "header_text_mismatch")
        assert mismatch["expected"] == "id"
        assert mismatch["found"] == "renamed_id"

    def test_schema_conformance_failure_returns_four(
        self, artifact_dir: Path, tmp_path: Path
    ) -> None:
        (artifact_dir / "transform.py").write_text(
            "def transform(df):\n"
            "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
            "    return df[['id', '_src_row']]\n"
        )
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(input_path)

        exit_code = cli.main(["run", str(artifact_dir), str(input_path)])

        assert exit_code == 4
        assert not cli.DRIFT_REPORT_PATH.exists()


class TestCmdTest:
    """The naru test subcommand: its own artifact fixture, built with a
    coercing transform, so the golden's parquet-normalized dtypes (e.g.
    object -> int64 for whole-number columns) match the live transform
    output -- see tests/test_goldenharness.py for why that matters.
    """

    @staticmethod
    def _write_golden_artifact(root: Path, transform_body: str, expected: pd.DataFrame) -> None:
        root.mkdir(parents=True, exist_ok=True)
        (root / "manifest.yaml").write_text(VALID_MANIFEST)
        (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
        (root / "schema.py").write_text(VALID_SCHEMA)
        (root / "transform.py").write_text(transform_body)
        (root / "validations.yaml").write_text(VALID_VALIDATIONS)
        (root / "CHANGELOG.md").write_text(VALID_CHANGELOG)
        golden = root / "golden"
        golden.mkdir(exist_ok=True)
        _make_input_workbook([(1, "a"), (2, "b")]).save(golden / "input_sample.xlsx")
        expected.to_parquet(golden / "expected_output.parquet", index=False)

    def test_matching_golden_returns_zero(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        transform_body = (
            "def transform(df):\n"
            "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
            "    return ops.coerce_numeric(df, 'id')\n"
        )
        self._write_golden_artifact(
            root,
            transform_body,
            pd.DataFrame({"id": [1.0, 2.0], "label": ["a", "b"], "_src_row": [2, 3]}),
        )

        exit_code = cli.main(["test", str(root)])

        assert exit_code == 0

    def test_value_drift_returns_two(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        transform_body = (
            "def transform(df):\n"
            "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
            "    return ops.coerce_numeric(df, 'id')\n"
        )
        self._write_golden_artifact(
            root,
            transform_body,
            pd.DataFrame({"id": [1.0, 2.0], "label": ["a", "STALE"], "_src_row": [2, 3]}),
        )

        exit_code = cli.main(["test", str(root)])

        assert exit_code == 2

    def test_schema_drift_returns_two(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        transform_body = (
            "def transform(df):\n"
            "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
            "    return ops.coerce_numeric(df, 'id')\n"
        )
        self._write_golden_artifact(
            root,
            transform_body,
            pd.DataFrame({"id": [1.0, 2.0], "label": ["a", "b"], "extra": [1, 2]}),
        )

        exit_code = cli.main(["test", str(root)])

        assert exit_code == 2


class TestCmdLint:
    def test_clean_artifact_returns_zero(self, artifact_dir: Path) -> None:
        exit_code = cli.main(["lint", str(artifact_dir)])
        assert exit_code == 0

    def test_violation_returns_five(self, artifact_dir: Path) -> None:
        (artifact_dir / "transform.py").write_text(
            "import os\n\n\ndef transform(df):\n    return df\n"
        )
        exit_code = cli.main(["lint", str(artifact_dir)])
        assert exit_code == 5

    def test_neither_artifact_kind_returns_one(self, tmp_path: Path) -> None:
        root = tmp_path / "not_an_artifact"
        root.mkdir()
        exit_code = cli.main(["lint", str(root)])
        assert exit_code == 1


class TestMainArgParsing:
    def test_missing_subcommand_exits_nonzero(self) -> None:
        with pytest.raises(SystemExit) as exc_info:
            cli.main([])
        assert exc_info.value.code != 0

    def test_run_requires_artifact_and_input(self) -> None:
        with pytest.raises(SystemExit):
            cli.main(["run", "only-one-arg"])
