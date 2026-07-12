"""Unit tests for src/naru/cli.py."""

import json
import sqlite3
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook
from typer.testing import CliRunner

from naru import cli
from naru.mapping import ColumnMapping, ExcelTarget, Mapping, to_yaml
from naru.profiler import profile as profile_fn
from naru.profiler import to_json

runner = CliRunner()

VALID_MANIFEST = """\
name: test_pipeline
version: v1
sheet: Sheet1
target_table: final_test
key: [id]
"""

VALID_FINGERPRINT = """\
{
  "sheet": "Sheet1",
  "header_row": 1,
  "columns": [
    {"name": "id", "type": "integer", "strictness": "strict"},
    {"name": "label", "type": "string", "strictness": "strict"}
  ]
}
"""

VALID_SCHEMA = """\
from pydantic import BaseModel


class SourceRow(BaseModel):
    id: int
    label: str


class TargetRow(BaseModel):
    id: int
    label: str
"""

VALID_TRANSFORM = """\
def transform(df):
    return ops.promote_header(df, header_row=1, column_names=["id", "label"])
"""

VALID_VALIDATIONS = "row_count:\n  min: 1\n"
VALID_CHANGELOG = "# Changelog\n\nv1: initial.\n"


def _make_input_workbook(
    rows: list[tuple[object, object]], header: tuple[str, str] = ("id", "label")
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=header[0])
    ws.cell(row=1, column=2, value=header[1])
    for i, (id_val, label_val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=id_val)
        ws.cell(row=i, column=2, value=label_val)
    return wb


def _write_artifact(root: Path, transform_body: str = VALID_TRANSFORM) -> None:
    root.mkdir(parents=True, exist_ok=True)
    (root / "manifest.yaml").write_text(VALID_MANIFEST)
    (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (root / "schema.py").write_text(VALID_SCHEMA)
    (root / "transform.py").write_text(transform_body)
    (root / "validations.yaml").write_text(VALID_VALIDATIONS)
    (root / "CHANGELOG.md").write_text(VALID_CHANGELOG)
    golden = root / "golden"
    golden.mkdir(exist_ok=True)
    _make_input_workbook([(1, "a"), (2, "b")]).save(golden / "input_sample.xlsx")
    pd.DataFrame({"id": [1, 2], "label": ["a", "b"]}).to_parquet(
        golden / "expected_output.parquet", index=False
    )


class TestProfile:
    def test_prints_json_to_stdout(self, tmp_path: Path) -> None:
        source = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(source)
        result = runner.invoke(cli.app, ["profile", str(source)])
        assert result.exit_code == 0
        assert '"sheets"' in result.output

    def test_writes_to_out_file(self, tmp_path: Path) -> None:
        source = tmp_path / "input.xlsx"
        out = tmp_path / "profile.json"
        _make_input_workbook([(1, "a")]).save(source)
        result = runner.invoke(cli.app, ["profile", str(source), "--out", str(out)])
        assert result.exit_code == 0
        assert out.exists()

    def test_missing_file_returns_one(self, tmp_path: Path) -> None:
        result = runner.invoke(cli.app, ["profile", str(tmp_path / "missing.xlsx")])
        assert result.exit_code == 1


class TestRun:
    def test_success_returns_zero_and_loads_db(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("artifact")
        _write_artifact(artifact)
        input_path = Path("input.xlsx")
        _make_input_workbook([(1, "a"), (2, "b")]).save(input_path)

        result = runner.invoke(cli.app, ["run", str(artifact), str(input_path)])

        assert result.exit_code == 0
        conn = sqlite3.connect(cli.DEFAULT_DB_PATH)
        rows = conn.execute("SELECT id, label FROM final_test ORDER BY id").fetchall()
        assert rows == [(1, "a"), (2, "b")]

    def test_as_of_flag_is_stored_exactly(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("artifact")
        _write_artifact(artifact)
        input_path = Path("input.xlsx")
        _make_input_workbook([(1, "a")]).save(input_path)

        result = runner.invoke(
            cli.app, ["run", str(artifact), str(input_path), "--as-of", "2020-06-01"]
        )

        assert result.exit_code == 0
        conn = sqlite3.connect(cli.DEFAULT_DB_PATH)
        (as_of,) = conn.execute("SELECT as_of FROM meta_runs").fetchone()
        assert as_of == "2020-06-01"

    def test_fingerprint_drift_returns_three_and_writes_report(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("artifact")
        _write_artifact(artifact)
        input_path = Path("input.xlsx")
        _make_input_workbook([(1, "a")], header=("renamed_id", "label")).save(input_path)

        result = runner.invoke(cli.app, ["run", str(artifact), str(input_path)])

        assert result.exit_code == 3
        assert cli.DRIFT_REPORT_PATH.exists()
        report = json.loads(cli.DRIFT_REPORT_PATH.read_text())
        kinds = {d["kind"] for d in report["differences"]}
        assert "header_text_mismatch" in kinds

    def test_validation_failure_returns_two(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("artifact")
        _write_artifact(
            artifact,
            transform_body=(
                "def transform(df):\n"
                "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
                "    return df[['id', '_src_row']]\n"
            ),
        )
        input_path = Path("input.xlsx")
        _make_input_workbook([(1, "a")]).save(input_path)

        result = runner.invoke(cli.app, ["run", str(artifact), str(input_path)])

        assert result.exit_code == 2
        assert not cli.DRIFT_REPORT_PATH.exists()


class TestTest:
    def test_matching_golden_returns_zero(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        _write_artifact(
            root,
            transform_body=(
                "def transform(df):\n"
                "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
                "    return ops.coerce_numeric(df, 'id')\n"
            ),
        )
        pd.DataFrame({"id": [1.0, 2.0], "label": ["a", "b"], "_src_row": [2, 3]}).to_parquet(
            root / "golden" / "expected_output.parquet", index=False
        )
        result = runner.invoke(cli.app, ["test", str(root)])
        assert result.exit_code == 0

    def test_value_drift_returns_two(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        _write_artifact(
            root,
            transform_body=(
                "def transform(df):\n"
                "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
                "    return ops.coerce_numeric(df, 'id')\n"
            ),
        )
        pd.DataFrame({"id": [1.0, 2.0], "label": ["a", "STALE"], "_src_row": [2, 3]}).to_parquet(
            root / "golden" / "expected_output.parquet", index=False
        )
        result = runner.invoke(cli.app, ["test", str(root)])
        assert result.exit_code == 2


class TestLint:
    def test_clean_artifact_returns_zero(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        _write_artifact(root)
        result = runner.invoke(cli.app, ["lint", str(root)])
        assert result.exit_code == 0

    def test_violation_returns_four(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        _write_artifact(root, transform_body="import os\n\n\ndef transform(df):\n    return df\n")
        result = runner.invoke(cli.app, ["lint", str(root)])
        assert result.exit_code == 4

    def test_neither_artifact_kind_returns_one(self, tmp_path: Path) -> None:
        root = tmp_path / "not_an_artifact"
        root.mkdir()
        result = runner.invoke(cli.app, ["lint", str(root)])
        assert result.exit_code == 1


class TestMapSuggest:
    def test_suggests_and_prints_draft_yaml(self, tmp_path: Path) -> None:
        source = tmp_path / "client.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        ws.cell(row=1, column=1, value="Deal ID")
        ws.cell(row=2, column=1, value="D1")
        wb.save(source)
        profile_path = tmp_path / "profile.json"
        profile_path.write_text(to_json(profile_fn(source)))

        schema_path = tmp_path / "schema.py"
        schema_path.write_text(
            "from pydantic import BaseModel\n\n\nclass TargetRow(BaseModel):\n    deal_id: str\n"
        )

        result = runner.invoke(
            cli.app,
            [
                "map",
                "suggest",
                str(profile_path),
                str(schema_path),
                "--target",
                "warehouse.positions",
                "--key",
                "deal_id",
            ],
        )
        assert result.exit_code == 0
        assert "deal_id" in result.output

    def test_unmatched_source_column_printed_as_stub_note(self, tmp_path: Path) -> None:
        source = tmp_path / "client.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        ws.cell(row=1, column=1, value="Totally Unrelated Column")
        ws.cell(row=2, column=1, value="x")
        wb.save(source)
        profile_path = tmp_path / "profile.json"
        profile_path.write_text(to_json(profile_fn(source)))
        schema_path = tmp_path / "schema.py"
        schema_path.write_text(
            "from pydantic import BaseModel\n\n\nclass TargetRow(BaseModel):\n    deal_id: str\n"
        )

        result = runner.invoke(
            cli.app,
            [
                "map",
                "suggest",
                str(profile_path),
                str(schema_path),
                "--target",
                "t",
                "--key",
                "deal_id",
            ],
        )
        assert result.exit_code == 0
        assert "still unmapped" in result.output

    def test_named_sheet_not_found_returns_one(self, tmp_path: Path) -> None:
        source = tmp_path / "client.xlsx"
        wb = Workbook()
        wb.active.title = "One"
        wb.active.cell(row=1, column=1, value="A")
        wb.active.cell(row=2, column=1, value="1")
        wb.save(source)
        profile_path = tmp_path / "profile.json"
        profile_path.write_text(to_json(profile_fn(source)))
        schema_path = tmp_path / "schema.py"
        schema_path.write_text(
            "from pydantic import BaseModel\n\n\nclass TargetRow(BaseModel):\n    a: str\n"
        )

        result = runner.invoke(
            cli.app,
            [
                "map",
                "suggest",
                str(profile_path),
                str(schema_path),
                "--target",
                "t",
                "--key",
                "a",
                "--sheet",
                "DoesNotExist",
            ],
        )
        assert result.exit_code == 1

    def test_named_sheet_found_succeeds(self, tmp_path: Path) -> None:
        source = tmp_path / "client.xlsx"
        wb = Workbook()
        wb.active.title = "One"
        wb.active.cell(row=1, column=1, value="Deal ID")
        wb.active.cell(row=2, column=1, value="D1")
        wb.save(source)
        profile_path = tmp_path / "profile.json"
        profile_path.write_text(to_json(profile_fn(source)))
        schema_path = tmp_path / "schema.py"
        schema_path.write_text(
            "from pydantic import BaseModel\n\n\nclass TargetRow(BaseModel):\n    deal_id: str\n"
        )

        result = runner.invoke(
            cli.app,
            [
                "map",
                "suggest",
                str(profile_path),
                str(schema_path),
                "--target",
                "t",
                "--key",
                "deal_id",
                "--sheet",
                "One",
            ],
        )
        assert result.exit_code == 0
        assert "deal_id" in result.output

    def test_multiple_sheets_without_sheet_flag_returns_one(self, tmp_path: Path) -> None:
        source = tmp_path / "client.xlsx"
        wb = Workbook()
        wb.active.title = "One"
        wb.active.cell(row=1, column=1, value="A")
        wb.active.cell(row=2, column=1, value="1")
        wb.create_sheet("Two").cell(row=1, column=1, value="B")
        wb.save(source)
        profile_path = tmp_path / "profile.json"
        profile_path.write_text(to_json(profile_fn(source)))
        schema_path = tmp_path / "schema.py"
        schema_path.write_text(
            "from pydantic import BaseModel\n\n\nclass TargetRow(BaseModel):\n    a: str\n"
        )

        result = runner.invoke(
            cli.app,
            [
                "map",
                "suggest",
                str(profile_path),
                str(schema_path),
                "--target",
                "t",
                "--key",
                "a",
            ],
        )
        assert result.exit_code == 1

    def test_bad_profile_file_returns_one(self, tmp_path: Path) -> None:
        profile_path = tmp_path / "profile.json"
        profile_path.write_text("not json")
        schema_path = tmp_path / "schema.py"
        schema_path.write_text(
            "from pydantic import BaseModel\n\n\nclass TargetRow(BaseModel):\n    a: str\n"
        )
        result = runner.invoke(
            cli.app,
            ["map", "suggest", str(profile_path), str(schema_path), "--target", "t", "--key", "a"],
        )
        assert result.exit_code == 1


class TestMapLearn:
    def test_learns_and_prints_report(self, tmp_path: Path) -> None:
        mapping_path = tmp_path / "mapping.yaml"
        m = Mapping(
            target="t",
            key=["deal_id"],
            on_duplicate="fail",
            columns=[
                ColumnMapping(
                    source="Cpn (%)",
                    target="coupon_rate",
                    transform="",
                    basis="synonym",
                    approved=True,
                )
            ],
            unmapped_source_columns="warn",
        )
        mapping_path.write_text(to_yaml(m))
        synonyms_path = tmp_path / "synonyms.yaml"

        result = runner.invoke(
            cli.app,
            ["map", "learn", str(mapping_path), "--synonyms-path", str(synonyms_path)],
        )
        assert result.exit_code == 0
        assert "added" in result.output

    def test_bad_mapping_yaml_returns_one(self, tmp_path: Path) -> None:
        mapping_path = tmp_path / "mapping.yaml"
        mapping_path.write_text("target: [unclosed\n")
        result = runner.invoke(cli.app, ["map", "learn", str(mapping_path)])
        assert result.exit_code == 1


EXCEL_MIRROR_FINGERPRINT = {
    "sheet": "Statement",
    "header_row": 1,
    "columns": [
        {"name": "Deal ID", "type": "string", "strictness": "strict"},
        {"name": "As Of", "type": "string", "strictness": "strict"},
    ],
}

MIRROR_TARGET_ROW_SCHEMA = """\
from pydantic import BaseModel


class TargetRow(BaseModel):
    deal_id: str
    as_of: str
"""


def _write_sql_mirror_artifact(root: Path) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    m = Mapping(
        target="warehouse.positions",
        key=["deal_id", "as_of"],
        on_duplicate="fail",
        columns=[
            ColumnMapping(
                source="Deal ID", target="deal_id", transform="", basis="exact", approved=True
            ),
            ColumnMapping(
                source="As Of", target="as_of", transform="", basis="exact", approved=True
            ),
        ],
        unmapped_source_columns="warn",
    )
    (root / "mapping.yaml").write_text(to_yaml(m))
    (root / "fingerprint.json").write_text(json.dumps(EXCEL_MIRROR_FINGERPRINT))
    (root / "schema.py").write_text(MIRROR_TARGET_ROW_SCHEMA)
    return root


def _write_mirror_source(path: Path, rows: list[tuple[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.cell(row=1, column=1, value="Deal ID")
    ws.cell(row=1, column=2, value="As Of")
    for i, (deal_id, as_of) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=deal_id)
        ws.cell(row=i, column=2, value=as_of)
    wb.save(path)


class TestMirror:
    def test_dry_run_by_default_writes_nothing(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("mapping_artifact")
        _write_sql_mirror_artifact(artifact)
        source = Path("client.xlsx")
        _write_mirror_source(source, [("D1", "2024-01-01")])

        result = runner.invoke(cli.app, ["mirror", str(artifact), str(source)])

        assert result.exit_code == 0
        assert "DRY RUN" in result.output
        assert not cli.DEFAULT_DB_PATH.exists() or (
            sqlite3.connect(cli.DEFAULT_DB_PATH)
            .execute("SELECT COUNT(*) FROM warehouse_positions")
            .fetchone()[0]
            == 0
        )

    def test_commit_writes_rows(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("mapping_artifact")
        _write_sql_mirror_artifact(artifact)
        source = Path("client.xlsx")
        _write_mirror_source(source, [("D1", "2024-01-01")])

        result = runner.invoke(cli.app, ["mirror", str(artifact), str(source), "--commit"])

        assert result.exit_code == 0
        conn = sqlite3.connect(cli.DEFAULT_DB_PATH)
        count = conn.execute("SELECT COUNT(*) FROM warehouse_positions").fetchone()[0]
        assert count == 1

    def test_duplicate_key_returns_two(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("mapping_artifact")
        _write_sql_mirror_artifact(artifact)
        source = Path("client.xlsx")
        _write_mirror_source(source, [("D1", "2024-01-01")])

        runner.invoke(cli.app, ["mirror", str(artifact), str(source), "--commit"])
        result = runner.invoke(cli.app, ["mirror", str(artifact), str(source), "--commit"])

        assert result.exit_code == 2

    def test_bad_mapping_artifact_returns_one(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        result = runner.invoke(cli.app, ["mirror", "does_not_exist", "does_not_exist.xlsx"])
        assert result.exit_code == 1

    def test_fingerprint_drift_returns_three(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("mapping_artifact")
        _write_sql_mirror_artifact(artifact)
        source = Path("client.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        ws.cell(row=1, column=1, value="Renamed Deal ID")
        ws.cell(row=1, column=2, value="As Of")
        ws.cell(row=2, column=1, value="D1")
        ws.cell(row=2, column=2, value="2024-01-01")
        wb.save(source)

        result = runner.invoke(cli.app, ["mirror", str(artifact), str(source)])

        assert result.exit_code == 3
        assert cli.DRIFT_REPORT_PATH.exists()

    def test_excel_target_commit_prints_backup_path(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        artifact = Path("mapping_artifact")
        artifact.mkdir()
        warehouse_fixture = Path(__file__).parent / "fixtures" / "warehouse_workbook.xlsx"
        import shutil

        shutil.copy2(warehouse_fixture, artifact / "warehouse_workbook.xlsx")

        m = Mapping(
            target="warehouse.positions",
            key=["deal_id", "as_of"],
            on_duplicate="fail",
            columns=[
                ColumnMapping(
                    source="Deal ID", target="deal_id", transform="", basis="exact", approved=True
                ),
                ColumnMapping(
                    source="Cpn (%)",
                    target="coupon_rate",
                    transform="coerce_numeric(scale=0.01)",
                    basis="synonym",
                    approved=True,
                ),
                ColumnMapping(
                    source="As Of", target="as_of", transform="", basis="exact", approved=True
                ),
                ColumnMapping(
                    source="Counterparty",
                    target="counterparty",
                    transform="",
                    basis="exact",
                    approved=True,
                ),
            ],
            unmapped_source_columns="warn",
            excel_target=ExcelTarget(
                path="warehouse_workbook.xlsx",
                first_data_col="A",
                last_data_col="D",
                column_order=["deal_id", "coupon_rate", "as_of", "counterparty"],
            ),
        )
        (artifact / "mapping.yaml").write_text(to_yaml(m))
        (artifact / "fingerprint.json").write_text(
            json.dumps(
                {
                    "sheet": "Statement",
                    "header_row": 1,
                    "columns": [
                        {"name": "Deal ID", "type": "string", "strictness": "strict"},
                        {"name": "Cpn (%)", "type": "float", "strictness": "strict"},
                        {"name": "As Of", "type": "string", "strictness": "strict"},
                        {"name": "Counterparty", "type": "string", "strictness": "strict"},
                    ],
                }
            )
        )
        (artifact / "warehouse_fingerprint.json").write_text(
            json.dumps(
                {
                    "sheet": "Positions",
                    "header_row": 1,
                    "columns": [
                        {"name": "Deal ID", "type": "string", "strictness": "strict"},
                        {"name": "Coupon Rate", "type": "float", "strictness": "strict"},
                        {"name": "As Of", "type": "string", "strictness": "strict"},
                        {"name": "Counterparty", "type": "string", "strictness": "strict"},
                    ],
                }
            )
        )
        (artifact / "schema.py").write_text(
            "from pydantic import BaseModel\n\n\n"
            "class TargetRow(BaseModel):\n"
            "    deal_id: str\n"
            "    coupon_rate: float\n"
            "    as_of: str\n"
            "    counterparty: str\n"
        )

        source = Path("client.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        ws.cell(row=1, column=1, value="Deal ID")
        ws.cell(row=1, column=2, value="Cpn (%)")
        ws.cell(row=1, column=3, value="As Of")
        ws.cell(row=1, column=4, value="Counterparty")
        ws.cell(row=2, column=1, value="D900")
        ws.cell(row=2, column=2, value=2.5)
        ws.cell(row=2, column=3, value="2024-03-31")
        ws.cell(row=2, column=4, value="Delta Partners")
        wb.save(source)

        result = runner.invoke(cli.app, ["mirror", str(artifact), str(source), "--commit"])

        assert result.exit_code == 0
        assert "backup:" in result.output


class TestQuery:
    def test_runs_recipe_and_prints_rows(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        db_path = Path("naru.sqlite")
        conn = sqlite3.connect(db_path)
        conn.execute("CREATE TABLE t (id INTEGER, name TEXT)")
        conn.execute("INSERT INTO t VALUES (1, 'a')")
        conn.commit()
        conn.close()

        recipes_dir = Path("recipes")
        recipes_dir.mkdir()
        (recipes_dir / "greet.sql").write_text(
            "---\nname: greet\nparams:\n  who:\n    type: string\n"
            "expected_columns: [id, name]\n---\nSELECT id, name FROM t WHERE name = :who\n"
        )

        result = runner.invoke(
            cli.app,
            ["query", "greet", "--recipes-dir", str(recipes_dir), "--param", "who=a"],
        )
        assert result.exit_code == 0
        assert "name" in result.output

    def test_malformed_param_returns_one(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        db_path = Path("naru.sqlite")
        sqlite3.connect(db_path).close()
        recipes_dir = Path("recipes")
        recipes_dir.mkdir()
        (recipes_dir / "greet.sql").write_text(
            "---\nname: greet\nparams:\n  who:\n    type: string\n"
            "expected_columns: [id]\n---\nSELECT 1 AS id\n"
        )
        result = runner.invoke(
            cli.app,
            ["query", "greet", "--recipes-dir", str(recipes_dir), "--param", "no_equals_sign"],
        )
        assert result.exit_code == 1

    def test_missing_param_returns_one(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        db_path = Path("naru.sqlite")
        sqlite3.connect(db_path).close()
        recipes_dir = Path("recipes")
        recipes_dir.mkdir()
        (recipes_dir / "greet.sql").write_text(
            "---\nname: greet\nparams:\n  who:\n    type: string\n"
            "expected_columns: [id]\n---\nSELECT 1 AS id WHERE :who = :who\n"
        )
        result = runner.invoke(cli.app, ["query", "greet", "--recipes-dir", str(recipes_dir)])
        assert result.exit_code == 1

    def test_shape_mismatch_returns_two(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        db_path = Path("naru.sqlite")
        sqlite3.connect(db_path).close()
        recipes_dir = Path("recipes")
        recipes_dir.mkdir()
        (recipes_dir / "bad.sql").write_text(
            "---\nname: bad\nexpected_columns: [id, extra]\n---\nSELECT 1 AS id\n"
        )
        result = runner.invoke(cli.app, ["query", "bad", "--recipes-dir", str(recipes_dir)])
        assert result.exit_code == 2

    def test_recipe_not_found_returns_one(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        db_path = Path("naru.sqlite")
        sqlite3.connect(db_path).close()
        recipes_dir = Path("recipes")
        recipes_dir.mkdir()
        result = runner.invoke(cli.app, ["query", "nope", "--recipes-dir", str(recipes_dir)])
        assert result.exit_code == 1


class TestNoArgsShowsHelp:
    def test_bare_invocation_shows_help(self) -> None:
        result = runner.invoke(cli.app, [])
        assert "Commands" in result.output


class TestMainModule:
    def test_python_dash_m_naru_shim_exposes_the_same_app(self) -> None:
        import naru.__main__

        assert naru.__main__.app is cli.app
