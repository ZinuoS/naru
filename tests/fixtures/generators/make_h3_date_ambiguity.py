"""Generate h3_date_ambiguity.xlsx: a single date column mixing raw Excel
serials, mm/dd/yyyy strings, and dd/mm/yyyy strings -- including rows where
the day is <=12, so the string form is genuinely ambiguous without knowing
which convention produced it. This is deliberately not something the
profiler should try to resolve; it should be flagged as a smell.

All figures are invented. Deterministic given the fixed SEED below.
"""

import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

SEED = 303
NUM_DATA_ROWS = 24
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "h3_date_ambiguity.xlsx"

HEADER_ROW = 1
FIRST_DATA_ROW = HEADER_ROW + 1

# Each entry is (day, month, year); day <= 12 makes mm/dd vs dd/mm
# genuinely ambiguous, day > 12 makes the string unambiguously dd/mm.
DATES = [
    (3, 4, 2019),  # ambiguous: Mar 4 (mm/dd) or Apr 3 (dd/mm)
    (25, 12, 2019),  # unambiguous dd/mm: day > 12
    (7, 1, 2020),  # ambiguous
    (11, 6, 2020),  # ambiguous
    (30, 6, 2020),  # unambiguous dd/mm
    (2, 9, 2020),  # ambiguous
    (15, 8, 2020),  # unambiguous dd/mm
    (9, 10, 2020),  # ambiguous
    (1, 1, 2021),  # ambiguous
    (28, 2, 2021),  # unambiguous dd/mm
    (5, 5, 2021),  # ambiguous (same either way)
    (12, 12, 2021),  # ambiguous (same either way)
    (17, 3, 2021),  # unambiguous dd/mm
    (4, 11, 2021),  # ambiguous
    (31, 1, 2022),  # unambiguous dd/mm
    (6, 6, 2022),  # ambiguous (same either way)
    (10, 7, 2022),  # ambiguous
    (22, 9, 2022),  # unambiguous dd/mm
    (8, 2, 2023),  # ambiguous
    (19, 4, 2023),  # unambiguous dd/mm
    (12, 1, 2023),  # ambiguous
    (27, 10, 2023),  # unambiguous dd/mm
    (3, 3, 2024),  # ambiguous (same either way)
    (14, 6, 2024),  # unambiguous dd/mm
]

# How each row's date value is encoded: excel serial, mm/dd/yyyy string,
# or dd/mm/yyyy string. Cycled deterministically, not randomly, so the mix
# is exact and reviewable.
ENCODINGS = ["serial", "mm_dd", "dd_mm"]


def build_workbook(seed: int = SEED) -> Workbook:
    """Construct the mixed-date-encoding workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.cell(row=1, column=1).value
    'Trade Date'
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    ws.cell(row=HEADER_ROW, column=1, value="Trade Date")
    ws.cell(row=HEADER_ROW, column=2, value="Amount")

    for i, (day, month, year) in enumerate(DATES):
        row = FIRST_DATA_ROW + i
        the_date = dt.date(year, month, day)
        encoding = ENCODINGS[i % len(ENCODINGS)]
        if encoding == "serial":
            ws.cell(row=row, column=1, value=int(to_excel(the_date)))
        elif encoding == "mm_dd":
            ws.cell(row=row, column=1, value=the_date.strftime("%m/%d/%Y"))
        else:
            ws.cell(row=row, column=1, value=the_date.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=rng.choice(range(1000, 100_000, 500)))

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
