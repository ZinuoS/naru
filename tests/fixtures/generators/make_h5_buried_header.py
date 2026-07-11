"""Generate h5_buried_header.xlsx: six banner rows, header on row 7, data
below, then footnote rows with real (non-blank) text after the data block
-- the hardest header/data-region boundary case, since the trailing rows
aren't blank the way ust_lite's are.

All figures are invented, including the report-date banner text, which is
fixture content (not a runtime wall-clock read). Deterministic given the
fixed SEED below.
"""

import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook

SEED = 505
NUM_DATA_ROWS = 28
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "h5_buried_header.xlsx"

HEADER_ROW = 7
FIRST_DATA_ROW = HEADER_ROW + 1

BANNER_LINES = [
    "ACME CAPITAL -- MONTHLY POSITION REPORT (SYNTHETIC DATA)",
    "For illustration only. Not sourced from any real book.",
    "Report Date: 2020-01-31",
    "Prepared By: Middle Office",
    "",
    "Confidential -- internal distribution only",
]

FOOTNOTE_LINES = [
    "Note: figures preliminary, subject to revision.",
    "Source: internal risk system.",
    "End of report.",
]

DESKS = ["Rates", "Credit", "FX", "Equities"]


def build_workbook(seed: int = SEED, num_rows: int = NUM_DATA_ROWS) -> Workbook:
    """Construct the buried-header-with-footnotes workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.cell(row=7, column=1).value
    'Position ID'
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"

    for i, line in enumerate(BANNER_LINES, start=1):
        ws.cell(row=i, column=1, value=line if line else None)

    headers = ["Position ID", "Desk", "Book", "Market Value"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=header)

    for i in range(num_rows):
        row = FIRST_DATA_ROW + i
        ws.cell(row=row, column=1, value=f"P{200000 + i}")
        ws.cell(row=row, column=2, value=rng.choice(DESKS))
        ws.cell(row=row, column=3, value=f"BOOK-{rng.randint(1, 12):02d}")
        ws.cell(row=row, column=4, value=round(rng.uniform(-5_000_000, 5_000_000), 2))

    footnote_start = FIRST_DATA_ROW + num_rows + 1
    for i, line in enumerate(FOOTNOTE_LINES):
        ws.cell(row=footnote_start + i, column=1, value=line)

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
