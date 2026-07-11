"""Generate the synthetic, deliberately-messy ust_lite.xlsx fixture.

All figures are invented; nothing here is sourced from TreasuryDirect or any
real auction. Run this script to regenerate tests/fixtures/ust_lite.xlsx —
output is deterministic given the fixed SEED below.

>>> import subprocess, sys
>>> subprocess.run([sys.executable, __file__]).returncode
0
"""

import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

SEED = 42
NUM_DATA_ROWS = 40
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "ust_lite.xlsx"

HEADER_ROW = 3
FIRST_DATA_ROW = HEADER_ROW + 1
EXCEL_EPOCH = dt.date(1899, 12, 30)

SECURITY_TERMS = [
    "4-Week Bill",
    "13-Week Bill",
    "26-Week Bill",
    "52-Week Bill",
    "2-Year Note",
    "3-Year Note",
    "5-Year Note",
    "7-Year Note",
    "10-Year Note",
    "20-Year Bond",
    "30-Year Bond",
]

# Header row 3 deliberately carries a merged cell over columns F:G — a
# common real-world artifact where a desk group-labels two columns under
# one heading. The tracer's hardcoded promote_header step assigns final
# column names by position, not by reading merged text, so this messiness
# doesn't need to be "solved" generically yet.
COLUMN_HEADERS = [
    "Auction Date",
    "Security Term",
    "CUSIP",
    "High Yield",
    "Offering Amt ($MM)",
    "Settlement Detail",
    "",  # merged away into the "Settlement Detail" cell to its left
]


def _random_cusip(rng: random.Random) -> str:
    """Build a synthetic 9-character CUSIP-shaped identifier.

    >>> _random_cusip(random.Random(0))[:2]
    '91'
    """
    digits = "".join(str(rng.randint(0, 9)) for _ in range(8))
    letters = "ABCDEFGHJKLMNPQRSTUVWXYZ"
    check = rng.choice(letters + "0123456789")
    return f"91{digits[:6]}{check}"


def _format_thousands(amount: int) -> str:
    """Render an integer with comma thousands separators, e.g. 38000 -> '38,000'.

    >>> _format_thousands(38000)
    '38,000'
    """
    return f"{amount:,}"


def build_workbook(seed: int = SEED, num_rows: int = NUM_DATA_ROWS) -> Workbook:
    """Construct the messy workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.max_row
    46
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.cell(row=1, column=1, value="U.S. TREASURY -- AUCTION RESULTS SUMMARY (SYNTHETIC DATA)")
    ws.cell(
        row=2,
        column=1,
        value="For illustration only. Not sourced from TreasuryDirect or any real auction.",
    )

    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=header)
    ws.merge_cells(start_row=HEADER_ROW, start_column=6, end_row=HEADER_ROW, end_column=7)

    start_date = dt.date(2019, 1, 15)
    for i in range(num_rows):
        row = FIRST_DATA_ROW + i
        auction_date = start_date + dt.timedelta(days=7 * i)
        issue_date = auction_date + dt.timedelta(days=2)
        high_yield = round(rng.uniform(0.05, 5.75), 3)
        offering_amt = rng.choice(range(20_000, 90_000, 1_000))
        bid_to_cover = round(rng.uniform(2.1, 2.9), 2)

        ws.cell(row=row, column=1, value=auction_date.strftime("%m/%d/%Y"))
        ws.cell(row=row, column=2, value=rng.choice(SECURITY_TERMS))
        ws.cell(row=row, column=3, value=_random_cusip(rng))
        ws.cell(row=row, column=4, value=f"{high_yield}%")
        ws.cell(row=row, column=5, value=_format_thousands(offering_amt))
        ws.cell(row=row, column=6, value=bid_to_cover)
        ws.cell(row=row, column=7, value=int(to_excel(issue_date)))

    # Trailing blank rows: openpyxl drops cells whose value is None when it
    # serializes the sheet, so a row written with value=None disappears on
    # reload and max_row shrinks back to the last data row. An empty string
    # is a real (if invisible) value, so it survives the round-trip and the
    # row still reads back as blank downstream (pandas/openpyxl treat ""
    # the same as NaN when the whole row is empty).
    last_data_row = FIRST_DATA_ROW + num_rows - 1
    for offset in range(1, 4):
        ws.cell(row=last_data_row + offset, column=1, value="")

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
