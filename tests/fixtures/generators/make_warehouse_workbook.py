"""Generate warehouse_workbook.xlsx: a synthetic, formula-laden "existing
warehouse" file for the Excel append-only mirror target
(naru.mirror._mirror_excel). Deliberately carries everything an
append-only writer must never disturb:

- A formula column ("Total", =Coupon Rate * Notional) with its own
  number_format, sitting just OUTSIDE the declared mirror region
  (A:D) -- proves the writer doesn't touch columns beyond what it owns.
- Conditional formatting on the Coupon Rate column.
- A named range covering the header + existing data rows.
- A second, entirely unrelated sheet ("Notes").

The mirror region itself (Sheet "Positions", header row 1, columns A:D
-- Deal ID / Coupon Rate / As Of / Counterparty) has 2 existing data
rows; a mirror run is expected to append new rows starting at row 4,
touching only columns A:D of those new rows.

All figures are invented. Deterministic given the fixed SEED below.
"""

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook as WorkbookType

SEED = 20240229
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "warehouse_workbook.xlsx"

HEADER_ROW = 1
REGION_HEADERS = ["Deal ID", "Coupon Rate", "As Of", "Counterparty"]
EXISTING_ROWS: list[tuple[str, float, str, str]] = [
    ("W1", 0.025, "2023-12-31", "Alpha Bank"),
    ("W2", 0.031, "2023-12-31", "Beta Capital"),
]


def build_workbook() -> WorkbookType:
    """Construct the warehouse workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.title
    'Positions'
    >>> wb.active["B2"].value
    0.025
    >>> wb.active["E2"].value
    1000000
    >>> wb.active["F2"].value
    '=B2*E2'
    >>> wb.sheetnames
    ['Positions', 'Notes']
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"

    for col_idx, header in enumerate(REGION_HEADERS, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=header)
    ws.cell(row=HEADER_ROW, column=5, value="Notional")
    ws.cell(row=HEADER_ROW, column=6, value="Total")

    notionals = [1_000_000, 2_000_000]
    for offset, (row_values, notional) in enumerate(zip(EXISTING_ROWS, notionals, strict=True)):
        row = HEADER_ROW + 1 + offset
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        ws.cell(row=row, column=5, value=notional)
        ws.cell(row=row, column=6, value=f"=B{row}*E{row}")
        ws.cell(row=row, column=6).number_format = "#,##0.00"

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "B2:B3", CellIsRule(operator="greaterThan", formula=["0.03"], fill=red_fill)
    )

    last_row = HEADER_ROW + len(EXISTING_ROWS)
    wb.defined_names["PositionsData"] = DefinedName(
        "PositionsData", attr_text=f"Positions!$A$1:$F${last_row}"
    )

    notes = wb.create_sheet("Notes")
    notes["A1"] = "This sheet is not part of the mirror target and must stay untouched."
    notes["A2"] = 42

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
