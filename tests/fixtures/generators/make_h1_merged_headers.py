"""Generate h1_merged_headers.xlsx: a two-row header with grouped, merged
column labels -- e.g. "Coupon" merged over "Rate (%)" and "Type" beneath it.

All figures are invented. Deterministic given the fixed SEED below.
"""

import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook

SEED = 101
NUM_DATA_ROWS = 25
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "h1_merged_headers.xlsx"

GROUP_HEADER_ROW = 1
SUB_HEADER_ROW = 2
FIRST_DATA_ROW = SUB_HEADER_ROW + 1

COUPON_TYPES = ["Fixed", "Floating"]
CURRENCIES = ["USD", "EUR", "GBP", "JPY"]


def build_workbook(seed: int = SEED, num_rows: int = NUM_DATA_ROWS) -> Workbook:
    """Construct the two-row-merged-header workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.cell(row=1, column=2).value
    'Coupon'
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    # Group header row: two merged spans over the sub-columns they own.
    ws.cell(row=GROUP_HEADER_ROW, column=1, value="")
    ws.cell(row=GROUP_HEADER_ROW, column=2, value="Coupon")
    ws.merge_cells(
        start_row=GROUP_HEADER_ROW, start_column=2, end_row=GROUP_HEADER_ROW, end_column=3
    )
    ws.cell(row=GROUP_HEADER_ROW, column=4, value="Notional")
    ws.merge_cells(
        start_row=GROUP_HEADER_ROW, start_column=4, end_row=GROUP_HEADER_ROW, end_column=5
    )
    ws.cell(row=GROUP_HEADER_ROW, column=6, value="")

    sub_headers = [
        "Trade Date",
        "Rate (%)",
        "Type",
        "Amount ($mm)",
        "Currency",
        "Maturity Date",
    ]
    for col_idx, header in enumerate(sub_headers, start=1):
        ws.cell(row=SUB_HEADER_ROW, column=col_idx, value=header)

    start_date = dt.date(2020, 1, 6)
    for i in range(num_rows):
        row = FIRST_DATA_ROW + i
        trade_date = start_date + dt.timedelta(days=3 * i)
        maturity = trade_date + dt.timedelta(days=365 * rng.randint(2, 10))
        rate = round(rng.uniform(0.5, 6.0), 3)
        notional = rng.choice(range(1_000_000, 50_000_000, 500_000))

        ws.cell(row=row, column=1, value=trade_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=rate)
        ws.cell(row=row, column=3, value=rng.choice(COUPON_TYPES))
        ws.cell(row=row, column=4, value=notional / 1_000_000)
        ws.cell(row=row, column=5, value=rng.choice(CURRENCIES))
        ws.cell(row=row, column=6, value=maturity.strftime("%Y-%m-%d"))

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
