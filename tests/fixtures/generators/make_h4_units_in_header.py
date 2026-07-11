"""Generate h4_units_in_header.xlsx: a single, clean header row whose
labels embed units -- "Notional ($mm)", "Cpn (%)", "Px (32nds)" -- with
plain unit-free values in the data cells below. Tests unit-smell detection
that must come from header text, not cell content.

All figures are invented. Deterministic given the fixed SEED below.
"""

import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook

SEED = 404
NUM_DATA_ROWS = 22
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "h4_units_in_header.xlsx"

HEADER_ROW = 1
FIRST_DATA_ROW = HEADER_ROW + 1

HEADERS = [
    "Trade Date",
    "Counterparty",
    "Notional ($mm)",
    "Cpn (%)",
    "Px (32nds)",
]

COUNTERPARTIES = ["Alpha Bank", "Beta Capital", "Gamma Securities", "Delta Partners"]


def build_workbook(seed: int = SEED, num_rows: int = NUM_DATA_ROWS) -> Workbook:
    """Construct the units-in-header workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.cell(row=1, column=3).value
    'Notional ($mm)'
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Blotter"

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=header)

    start_date = dt.date(2022, 1, 3)
    for i in range(num_rows):
        row = FIRST_DATA_ROW + i
        trade_date = start_date + dt.timedelta(days=4 * i)
        ws.cell(row=row, column=1, value=trade_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=rng.choice(COUNTERPARTIES))
        ws.cell(row=row, column=3, value=round(rng.uniform(1.0, 250.0), 2))
        ws.cell(row=row, column=4, value=round(rng.uniform(0.25, 7.5), 3))
        ws.cell(row=row, column=5, value=rng.randint(0, 31))

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
