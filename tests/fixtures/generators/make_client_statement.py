"""Generate client_statement.xlsx: a synthetic client-provided position
statement whose column names and units deliberately differ from the
warehouse schema it will be mirrored into (docs/spec.md §2.7's own worked
example: "Cpn (%)" holds a plain percentage number, not the decimal
fraction naru.ops.coerce_numeric(scale=0.01) will produce for
coupon_rate).

Used by the naru mirror end-to-end demo/test (tests/test_end_to_end.py):
"Deal ID" and "As Of" match the warehouse schema via tier-1 exact
normalization; "Cpn (%)" is a pre-seeded tier-2 synonym; "Broker" starts
completely unmatched -- standing in for a column only a human (or, once
implemented, tier 3/4) can resolve -- and only becomes an automatic
tier-2 match after naru map learn promotes a human-approved basis: llm
match for it; "Notes" stays unmapped throughout, to demonstrate the
unmapped_source_columns: warn report.

All figures are invented. Deterministic given the fixed SEED below.
"""

import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook

SEED = 20240131
NUM_DATA_ROWS = 5
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "client_statement.xlsx"

HEADER_ROW = 1
FIRST_DATA_ROW = HEADER_ROW + 1

HEADERS = ["Deal ID", "Cpn (%)", "As Of", "Broker", "Notes"]

BROKERS = ["Alpha Bank", "Beta Capital", "Gamma Securities"]
AS_OF = "2024-01-31"


def build_workbook(seed: int = SEED, num_rows: int = NUM_DATA_ROWS) -> Workbook:
    """Construct the client-statement workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.cell(row=1, column=2).value
    'Cpn (%)'
    >>> wb.active.cell(row=2, column=1).value
    'D001'
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=header)

    for i in range(num_rows):
        row = FIRST_DATA_ROW + i
        ws.cell(row=row, column=1, value=f"D{i + 1:03d}")
        ws.cell(row=row, column=2, value=round(rng.uniform(1.5, 7.5), 3))
        ws.cell(row=row, column=3, value=AS_OF)
        ws.cell(row=row, column=4, value=rng.choice(BROKERS))
        ws.cell(row=row, column=5, value=f"note {i + 1}")

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
