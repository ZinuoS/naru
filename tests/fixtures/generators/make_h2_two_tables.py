"""Generate h2_two_tables.xlsx: two distinct, differently-shaped tables
stacked on one sheet with a blank row between them -- Positions above,
Trades below.

All figures are invented. Deterministic given the fixed SEED below.
"""

import datetime as dt
import random
import string
from pathlib import Path

from openpyxl import Workbook

SEED = 202
NUM_POSITIONS_ROWS = 14
NUM_TRADES_ROWS = 13
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "h2_two_tables.xlsx"

TABLE_A_HEADER_ROW = 1
TABLE_A_FIRST_DATA_ROW = TABLE_A_HEADER_ROW + 1
TABLE_A_LAST_DATA_ROW = TABLE_A_FIRST_DATA_ROW + NUM_POSITIONS_ROWS - 1
GAP_ROW = TABLE_A_LAST_DATA_ROW + 1
TABLE_B_HEADER_ROW = GAP_ROW + 1
TABLE_B_FIRST_DATA_ROW = TABLE_B_HEADER_ROW + 1

SIDES = ["Buy", "Sell"]


def _random_ticker(rng: random.Random) -> str:
    """Build a synthetic 3-4 letter ticker.

    >>> _random_ticker(random.Random(0))
    'YNBI'
    """
    length = rng.choice([3, 4])
    return "".join(rng.choice(string.ascii_uppercase) for _ in range(length))


def build_workbook(seed: int = SEED) -> Workbook:
    """Construct the two-tables-on-one-sheet workbook in memory.

    >>> wb = build_workbook()
    >>> wb.active.cell(row=1, column=1).value
    'Ticker'
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Book"

    # Table A: Positions
    for col_idx, header in enumerate(["Ticker", "Quantity", "Price"], start=1):
        ws.cell(row=TABLE_A_HEADER_ROW, column=col_idx, value=header)
    for i in range(NUM_POSITIONS_ROWS):
        row = TABLE_A_FIRST_DATA_ROW + i
        ws.cell(row=row, column=1, value=_random_ticker(rng))
        ws.cell(row=row, column=2, value=rng.choice(range(-5000, 5000, 100)))
        ws.cell(row=row, column=3, value=round(rng.uniform(5.0, 500.0), 2))

    # GAP_ROW is left entirely empty -- no cells written.

    # Table B: Trades (different shape/columns entirely)
    for col_idx, header in enumerate(["Trade ID", "Side", "Notional", "Trade Date"], start=1):
        ws.cell(row=TABLE_B_HEADER_ROW, column=col_idx, value=header)
    start_date = dt.date(2021, 3, 1)
    for i in range(NUM_TRADES_ROWS):
        row = TABLE_B_FIRST_DATA_ROW + i
        ws.cell(row=row, column=1, value=f"T{100000 + i}")
        ws.cell(row=row, column=2, value=rng.choice(SIDES))
        ws.cell(row=row, column=3, value=rng.choice(range(10_000, 2_000_000, 10_000)))
        ws.cell(
            row=row, column=4, value=(start_date + dt.timedelta(days=2 * i)).strftime("%Y-%m-%d")
        )

    wb.properties.created = dt.datetime(2026, 1, 1)
    wb.properties.modified = dt.datetime(2026, 1, 1)
    wb.properties.creator = "naru fixture generator"
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
