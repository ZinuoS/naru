"""Proves examples/counterparty_mirror (the second worked example, spec.md
§3) actually works, not just that it once did when hand-verified. Runs
against a temp copy for the commit-mode test so the checked-in fixture
is never mutated by the test suite.
"""

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from naru import mirror
from naru.lint import lint_artifact
from naru.runtime import FingerprintDriftError

EXAMPLE_DIR = Path(__file__).parent.parent / "examples" / "counterparty_mirror"


def test_lints_clean() -> None:
    assert lint_artifact(EXAMPLE_DIR) == []


def test_dry_run_matches_expected_summary(tmp_path: Path) -> None:
    result = mirror.mirror(
        EXAMPLE_DIR,
        EXAMPLE_DIR / "client_statement.xlsx",
        tmp_path / "naru.sqlite",
        tmp_path / "raw",
        dry_run=True,
    )
    assert result.summary.rows_in == 5
    assert result.summary.rows_out == 5
    assert result.summary.unmapped_source_columns == ["Notes"]
    assert result.summary.target_columns_not_populated == []


def test_commit_writes_rows_and_leaves_warehouse_readable(tmp_path: Path) -> None:
    working_copy = tmp_path / "counterparty_mirror"
    shutil.copytree(EXAMPLE_DIR, working_copy)

    result = mirror.mirror(
        working_copy,
        working_copy / "client_statement.xlsx",
        tmp_path / "naru.sqlite",
        tmp_path / "raw",
        dry_run=False,
    )
    assert result.backup_path is not None
    assert result.backup_path.exists()

    wb = load_workbook(working_copy / "warehouse_workbook.xlsx")
    ws = wb["Positions"]
    assert ws.max_row == 1 + 2 + 5  # header + 2 existing + 5 mirrored
    assert [ws.cell(row=r, column=1).value for r in range(4, 9)] == [
        "D001",
        "D002",
        "D003",
        "D004",
        "D005",
    ]


def test_renamed_column_variant_halts_with_drift(tmp_path: Path) -> None:
    with pytest.raises(FingerprintDriftError, match="Deal Identifier"):
        mirror.mirror(
            EXAMPLE_DIR,
            EXAMPLE_DIR / "client_statement_renamed_column.xlsx",
            tmp_path / "naru.sqlite",
            tmp_path / "raw",
            dry_run=True,
        )
