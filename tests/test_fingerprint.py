"""Unit tests for src/naru/fingerprint.py."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from naru import fingerprint as fp
from naru.artifact import Fingerprint, HeaderColumnSpec

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


def _simple_fingerprint(**overrides: object) -> Fingerprint:
    defaults: dict[str, object] = {
        "sheet": "Sheet1",
        "header_row": 1,
        "columns": [
            HeaderColumnSpec(name="id", type="integer", strictness="strict"),
            HeaderColumnSpec(name="label", type="string", strictness="strict"),
        ],
        "max_rows_from_header_to_data": 1,
    }
    defaults.update(overrides)
    return Fingerprint(**defaults)  # type: ignore[arg-type]


def _make_workbook(rows: list[tuple[object, object]], sheet_title: str = "Sheet1") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="label")
    for i, (id_val, label_val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=id_val)
        ws.cell(row=i, column=2, value=label_val)
    return wb


class TestDifferenceMessage:
    def test_renders_spec_example_format(self) -> None:
        d = fp.Difference("header_text_mismatch", "Results", "High Yield", "High Rate", 7)
        assert d.message() == "sheet 'Results' col 7: expected `High Yield`, found `High Rate`"

    def test_renders_without_column_position(self) -> None:
        d = fp.Difference("sheet_missing", "Results", "Results", "Other")
        assert d.message() == "sheet 'Results': expected `Results`, found `Other`"


class TestCheckFingerprintSheetMatching:
    def test_exact_match_ok(self) -> None:
        wb = _make_workbook([(1, "a")])
        result = fp.check_fingerprint(_simple_fingerprint(), wb)
        assert result.ok

    def test_missing_sheet_reports_difference(self) -> None:
        wb = _make_workbook([(1, "a")], sheet_title="Other")
        result = fp.check_fingerprint(_simple_fingerprint(), wb)
        assert not result.ok
        assert result.differences[0].kind == "sheet_missing"

    def test_regex_match_ok(self) -> None:
        wb = _make_workbook([(1, "a")], sheet_title="Results_2020")
        finger = _simple_fingerprint(sheet=r"Results_\d+", sheet_is_regex=True)
        result = fp.check_fingerprint(finger, wb)
        assert result.ok

    def test_regex_no_match_reports_difference(self) -> None:
        wb = _make_workbook([(1, "a")], sheet_title="Nope")
        finger = _simple_fingerprint(sheet=r"Results_\d+", sheet_is_regex=True)
        result = fp.check_fingerprint(finger, wb)
        assert not result.ok
        assert result.differences[0].kind == "sheet_missing"

    def test_ambiguous_regex_match_reports_difference(self) -> None:
        wb = Workbook()
        wb.active.title = "Results_A"
        wb.create_sheet("Results_B")
        finger = _simple_fingerprint(sheet=r"Results_.*", sheet_is_regex=True)
        result = fp.check_fingerprint(finger, wb)
        assert not result.ok
        assert result.differences[0].kind == "sheet_ambiguous"

    def test_sheet_index_mismatch_reports_difference(self) -> None:
        wb = Workbook()
        wb.active.title = "Cover"
        ws = wb.create_sheet("Sheet1")
        ws.cell(row=1, column=1, value="id")
        ws.cell(row=1, column=2, value="label")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="a")
        finger = _simple_fingerprint(sheet_index=0)
        result = fp.check_fingerprint(finger, wb)
        assert not result.ok
        assert result.differences[0].kind == "sheet_position_mismatch"
        assert result.differences[0].expected == "0"
        assert result.differences[0].found == "1"

    def test_sheet_index_none_skips_position_check(self) -> None:
        wb = Workbook()
        wb.active.title = "Cover"
        ws = wb.create_sheet("Sheet1")
        ws.cell(row=1, column=1, value="id")
        ws.cell(row=1, column=2, value="label")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="a")
        result = fp.check_fingerprint(_simple_fingerprint(sheet_index=None), wb)
        assert result.ok


class TestCheckFingerprintHeaderSignature:
    def test_strict_column_mismatch_reports_expected_and_found(self) -> None:
        wb = _make_workbook([(1, "a")])
        wb.active.cell(row=1, column=2, value="renamed_label")
        result = fp.check_fingerprint(_simple_fingerprint(), wb)
        assert not result.ok
        diff = next(d for d in result.differences if d.kind == "header_text_mismatch")
        assert diff.column_position == 2
        assert diff.expected == "label"
        assert diff.found == "renamed_label"

    def test_position_only_column_text_never_checked(self) -> None:
        wb = _make_workbook([(1, "a")])
        wb.active.cell(row=1, column=2, value="anything at all")
        finger = _simple_fingerprint(
            columns=[
                HeaderColumnSpec(name="id", type="integer", strictness="strict"),
                HeaderColumnSpec(name="label", type="string", strictness="position_only"),
            ]
        )
        result = fp.check_fingerprint(finger, wb)
        assert result.ok

    def test_blank_strict_header_reports_blank(self) -> None:
        wb = _make_workbook([(1, "a")])
        wb.active.cell(row=1, column=2).value = None
        result = fp.check_fingerprint(_simple_fingerprint(), wb)
        diff = next(d for d in result.differences if d.kind == "header_text_mismatch")
        assert diff.found == "(blank)"


class TestCheckFingerprintDataStartsNearHeader:
    def test_data_immediately_below_header_ok(self) -> None:
        wb = _make_workbook([(1, "a")])
        result = fp.check_fingerprint(_simple_fingerprint(max_rows_from_header_to_data=1), wb)
        assert result.ok

    def test_data_too_far_below_header_reports_difference(self) -> None:
        wb = _make_workbook([])  # header only -- row 2 stays genuinely blank
        wb.active.cell(row=5, column=1, value=1)
        wb.active.cell(row=5, column=2, value="a")
        result = fp.check_fingerprint(_simple_fingerprint(max_rows_from_header_to_data=1), wb)
        assert not result.ok
        assert result.differences[0].kind == "data_start_too_far"


class TestCheckFingerprintColumnTypes:
    def test_matching_types_ok(self) -> None:
        wb = _make_workbook([(1, "a"), (2, "b")])
        result = fp.check_fingerprint(_simple_fingerprint(), wb)
        assert result.ok

    def test_mismatched_type_reports_expected_and_found(self) -> None:
        wb = _make_workbook([("not-an-int", "a")])
        result = fp.check_fingerprint(_simple_fingerprint(), wb)
        assert not result.ok
        diff = next(d for d in result.differences if d.kind == "column_type_mismatch")
        assert diff.column_position == 1
        assert diff.expected == "integer"
        assert diff.found == "string"

    def test_optional_column_type_never_checked(self) -> None:
        wb = _make_workbook([(1, "a")])
        wb.active.cell(row=2, column=2, value=999)  # int instead of declared string
        finger = _simple_fingerprint(
            columns=[
                HeaderColumnSpec(name="id", type="integer", strictness="strict"),
                HeaderColumnSpec(name="label", type="string", strictness="optional"),
            ]
        )
        result = fp.check_fingerprint(finger, wb)
        assert result.ok

    def test_empty_column_skips_type_check(self) -> None:
        wb = _make_workbook([(1, None)])
        result = fp.check_fingerprint(_simple_fingerprint(), wb)
        assert result.ok


class TestGenerateFingerprint:
    def test_strict_columns_get_actual_header_text(self) -> None:
        generated = fp.generate_fingerprint(FIXTURES_DIR / "ust_lite.xlsx", "Results", header_row=3)
        assert generated.columns[0].strictness == "strict"
        assert generated.columns[0].name == "Auction Date"
        assert generated.columns[0].type == "string"

    def test_merged_columns_get_position_only(self) -> None:
        generated = fp.generate_fingerprint(FIXTURES_DIR / "ust_lite.xlsx", "Results", header_row=3)
        assert generated.columns[5].strictness == "position_only"
        assert generated.columns[6].strictness == "position_only"

    def test_merge_outside_header_row_does_not_affect_strictness(self, tmp_path: Path) -> None:
        wb = _make_workbook([(1, "a"), (2, "b")])
        wb.active.merge_cells("A2:B2")  # merged in a data row, not the header row
        path = tmp_path / "merged_elsewhere.xlsx"
        wb.save(path)
        generated = fp.generate_fingerprint(path, "Sheet1", header_row=1)
        assert all(c.strictness == "strict" for c in generated.columns)

    def test_covered_merge_cell_gets_placeholder_name(self) -> None:
        generated = fp.generate_fingerprint(FIXTURES_DIR / "ust_lite.xlsx", "Results", header_row=3)
        assert generated.columns[6].name == "column_7"

    def test_sheet_index_recorded(self) -> None:
        generated = fp.generate_fingerprint(FIXTURES_DIR / "ust_lite.xlsx", "Results", header_row=3)
        assert generated.sheet_index == 0

    def test_types_inferred_from_data_sample(self) -> None:
        generated = fp.generate_fingerprint(FIXTURES_DIR / "ust_lite.xlsx", "Results", header_row=3)
        assert generated.columns[5].type == "float"  # bid_to_cover
        assert generated.columns[6].type == "integer"  # issue_date serial

    def test_generated_fingerprint_passes_its_own_check(self) -> None:
        generated = fp.generate_fingerprint(FIXTURES_DIR / "ust_lite.xlsx", "Results", header_row=3)
        wb = load_workbook(FIXTURES_DIR / "ust_lite.xlsx")
        result = fp.check_fingerprint(generated, wb)
        assert result.ok, result.differences
