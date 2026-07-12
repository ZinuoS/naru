"""Unit tests for src/naru/runtime.py."""

import datetime as dt
import sqlite3
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from naru import runtime, store

VALID_MANIFEST = """\
name: test_pipeline
version: v1
sheet: Sheet1
target_table: final_test
key: [id]
"""

VALID_FINGERPRINT = """\
{
  "sheet": "Sheet1",
  "header_row": 1,
  "columns": [
    {"name": "id", "type": "integer", "strictness": "position_only"},
    {"name": "label", "type": "string", "strictness": "position_only"}
  ]
}
"""

VALID_SCHEMA = """\
from pydantic import BaseModel


class SourceRow(BaseModel):
    id: int
    label: str


class TargetRow(BaseModel):
    id: int
    label: str
"""

VALID_TRANSFORM = """\
def transform(df):
    return ops.promote_header(df, header_row=1, column_names=["id", "label"])
"""

VALID_VALIDATIONS = """\
row_count:
  min: 1
"""

VALID_CHANGELOG = "# Changelog\n\nv1: initial.\n"


def _write_artifact(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    (root / "manifest.yaml").write_text(VALID_MANIFEST)
    (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (root / "schema.py").write_text(VALID_SCHEMA)
    (root / "transform.py").write_text(VALID_TRANSFORM)
    (root / "validations.yaml").write_text(VALID_VALIDATIONS)
    (root / "CHANGELOG.md").write_text(VALID_CHANGELOG)

    golden = root / "golden"
    golden.mkdir(exist_ok=True)
    wb = _make_input_workbook([(1, "a"), (2, "b")])
    wb.save(golden / "input_sample.xlsx")
    pd.DataFrame({"id": [1, 2], "label": ["a", "b"]}).to_parquet(
        golden / "expected_output.parquet", index=False
    )


def _make_input_workbook(rows: list[tuple[int, str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="label")
    for i, (id_val, label_val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=id_val)
        ws.cell(row=i, column=2, value=label_val)
    return wb


@pytest.fixture
def artifact_dir(tmp_path: Path) -> Path:
    root = tmp_path / "artifact"
    _write_artifact(root)
    return root


class TestRun:
    def test_loads_rows_into_sqlite(self, artifact_dir: Path, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a"), (2, "b")]).save(input_path)

        result = runtime.run(
            artifact_path=artifact_dir,
            input_path=input_path,
            db_path=tmp_path / "naru.sqlite",
            raw_dir=tmp_path / "raw",
        )

        assert result.fingerprint_check.ok
        assert result.output_check.ok
        assert len(result.row_ids) == 2

        conn = sqlite3.connect(tmp_path / "naru.sqlite")
        rows = conn.execute("SELECT id, label FROM final_test ORDER BY id").fetchall()
        assert rows == [(1, "a"), (2, "b")]

    def test_fingerprint_drift_raises_before_any_writes(
        self, artifact_dir: Path, tmp_path: Path
    ) -> None:
        input_path = tmp_path / "input.xlsx"
        wb = _make_input_workbook([(1, "a")])
        wb.active.cell(row=1, column=1, value="renamed_id_column")
        wb.save(input_path)
        # This fixture's fingerprint declares id/label as position_only, so
        # rename the header via a strict override to actually trigger drift.
        (artifact_dir / "fingerprint.json").write_text(
            VALID_FINGERPRINT.replace('"strictness": "position_only"', '"strictness": "strict"')
        )

        db_path = tmp_path / "naru.sqlite"
        with pytest.raises(runtime.FingerprintDriftError) as exc_info:
            runtime.run(
                artifact_path=artifact_dir,
                input_path=input_path,
                db_path=db_path,
                raw_dir=tmp_path / "raw",
            )
        assert exc_info.value.result.differences
        # The final table structure may exist (idempotent DDL), but no
        # rows were ever written -- that's the atomicity guarantee.
        conn = sqlite3.connect(db_path)
        row_count = conn.execute("SELECT COUNT(*) FROM final_test").fetchone()[0]
        assert row_count == 0

    def test_writes_lineage_rows_joinable_to_final_rows(
        self, artifact_dir: Path, tmp_path: Path
    ) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(input_path)

        runtime.run(
            artifact_path=artifact_dir,
            input_path=input_path,
            db_path=tmp_path / "naru.sqlite",
            raw_dir=tmp_path / "raw",
        )

        conn = sqlite3.connect(tmp_path / "naru.sqlite")
        joined = conn.execute(
            """
            SELECT f.id, f.label, l.sheet, l.source_row_start, l.source_row_end
            FROM final_test AS f
            JOIN meta_lineage AS l
                ON l.final_table = 'final_test' AND l.row_id = f.row_id
            """
        ).fetchall()
        assert joined == [(1, "a", "Sheet1", 2, 2)]

    def test_registers_raw_file_with_correct_hash(self, artifact_dir: Path, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(input_path)
        raw_dir = tmp_path / "raw"

        runtime.run(
            artifact_path=artifact_dir,
            input_path=input_path,
            db_path=tmp_path / "naru.sqlite",
            raw_dir=raw_dir,
        )

        import hashlib

        expected_sha256 = hashlib.sha256(input_path.read_bytes()).hexdigest()
        assert (raw_dir / expected_sha256).exists()
        conn = sqlite3.connect(tmp_path / "naru.sqlite")
        row = conn.execute(
            "SELECT original_name FROM raw_files WHERE sha256 = ?", (expected_sha256,)
        ).fetchone()
        assert row == ("input.xlsx",)

    def test_as_of_none_is_never_backfilled(self, artifact_dir: Path, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(input_path)

        result = runtime.run(
            artifact_path=artifact_dir,
            input_path=input_path,
            db_path=tmp_path / "naru.sqlite",
            raw_dir=tmp_path / "raw",
        )

        conn = sqlite3.connect(tmp_path / "naru.sqlite")
        (as_of,) = conn.execute(
            "SELECT as_of FROM meta_runs WHERE run_id = ?", (result.run_id,)
        ).fetchone()
        assert as_of is None

    def test_as_of_given_is_stored_exactly(self, artifact_dir: Path, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(input_path)

        result = runtime.run(
            artifact_path=artifact_dir,
            input_path=input_path,
            db_path=tmp_path / "naru.sqlite",
            raw_dir=tmp_path / "raw",
            as_of=dt.date(2020, 6, 1),
        )

        conn = sqlite3.connect(tmp_path / "naru.sqlite")
        (as_of,) = conn.execute(
            "SELECT as_of FROM meta_runs WHERE run_id = ?", (result.run_id,)
        ).fetchone()
        assert as_of == "2020-06-01"

    def test_rerun_supersedes_via_store(self, artifact_dir: Path, tmp_path: Path) -> None:
        db_path = tmp_path / "naru.sqlite"
        raw_dir = tmp_path / "raw"

        first_input = tmp_path / "first.xlsx"
        _make_input_workbook([(1, "a")]).save(first_input)
        runtime.run(
            artifact_path=artifact_dir, input_path=first_input, db_path=db_path, raw_dir=raw_dir
        )

        second_input = tmp_path / "second.xlsx"
        _make_input_workbook([(1, "a-corrected")]).save(second_input)
        runtime.run(
            artifact_path=artifact_dir, input_path=second_input, db_path=db_path, raw_dir=raw_dir
        )

        conn = sqlite3.connect(db_path)
        assert conn.execute("SELECT COUNT(*) FROM final_test").fetchone()[0] == 2
        active = store.active_rows(conn, "final_test")
        assert len(active) == 1
        assert active[0]["label"] == "a-corrected"

    def test_transform_output_mismatched_with_target_row_raises(
        self, artifact_dir: Path, tmp_path: Path
    ) -> None:
        (artifact_dir / "transform.py").write_text(
            "def transform(df):\n"
            "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
            "    return df[['id', '_src_row']]\n"
        )
        input_path = tmp_path / "input.xlsx"
        _make_input_workbook([(1, "a")]).save(input_path)

        with pytest.raises(runtime.RuntimeCheckError, match="don't match TargetRow fields"):
            runtime.run(
                artifact_path=artifact_dir,
                input_path=input_path,
                db_path=tmp_path / "naru.sqlite",
                raw_dir=tmp_path / "raw",
            )
