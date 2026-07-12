"""Mutation tests: programmatically corrupt copies of the real
ust_auction_results golden input and assert each specific corruption
produces exit-behavior 3 with a drift report naming that SPECIFIC change,
not just any failure.
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from naru import cli

ARTIFACT_PATH = Path(__file__).resolve().parent.parent / "pipelines" / "ust_auction_results" / "v1"
GOLDEN_INPUT = ARTIFACT_PATH / "golden" / "input_sample.xlsx"

runner = CliRunner()


@pytest.fixture(autouse=True)
def _isolated_cwd(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """main() writes naru.sqlite/.naru/raw/drift_report.json relative to
    CWD -- always run from an isolated directory, never the repo root.
    """
    work_dir = tmp_path / "cwd"
    work_dir.mkdir()
    monkeypatch.chdir(work_dir)


def _run_cli(input_path: Path) -> int:
    result = runner.invoke(cli.app, ["run", str(ARTIFACT_PATH), str(input_path)])
    return int(result.exit_code)


def _read_drift_report() -> dict[str, Any]:
    result: dict[str, Any] = json.loads(cli.DRIFT_REPORT_PATH.read_text())
    return result


class TestRenamedColumn:
    """Mutation (a): rename one strict column's header text."""

    def test_exit_code_three(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb["Results"].cell(row=3, column=2, value="Sec Term (renamed)")
        mutated = tmp_path / "renamed_column.xlsx"
        wb.save(mutated)

        exit_code = _run_cli(mutated)

        assert exit_code == 3

    def test_report_names_the_specific_column_and_values(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb["Results"].cell(row=3, column=2, value="Sec Term (renamed)")
        mutated = tmp_path / "renamed_column.xlsx"
        wb.save(mutated)

        _run_cli(mutated)
        report = _read_drift_report()

        mismatches = [d for d in report["differences"] if d["kind"] == "header_text_mismatch"]
        assert len(mismatches) == 1
        assert mismatches[0]["column_position"] == 2
        assert mismatches[0]["expected"] == "Security Term"
        assert mismatches[0]["found"] == "Sec Term (renamed)"
        assert mismatches[0]["sheet"] == "Results"
        assert (
            mismatches[0]["message"]
            == "sheet 'Results' col 2: expected `Security Term`, found `Sec Term (renamed)`"
        )

    def test_untouched_columns_report_no_mismatch(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb["Results"].cell(row=3, column=2, value="Sec Term (renamed)")
        mutated = tmp_path / "renamed_column.xlsx"
        wb.save(mutated)

        _run_cli(mutated)
        report = _read_drift_report()

        mismatched_positions = {
            d["column_position"]
            for d in report["differences"]
            if d["kind"] == "header_text_mismatch"
        }
        assert mismatched_positions == {2}


class TestInsertedSheet:
    """Mutation (b): insert a new sheet before the data sheet."""

    def test_exit_code_three(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb.create_sheet("Cover Page", 0)
        mutated = tmp_path / "inserted_sheet.xlsx"
        wb.save(mutated)

        exit_code = _run_cli(mutated)

        assert exit_code == 3

    def test_report_names_the_position_shift(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb.create_sheet("Cover Page", 0)
        mutated = tmp_path / "inserted_sheet.xlsx"
        wb.save(mutated)

        _run_cli(mutated)
        report = _read_drift_report()

        position_diffs = [
            d for d in report["differences"] if d["kind"] == "sheet_position_mismatch"
        ]
        assert len(position_diffs) == 1
        assert position_diffs[0]["sheet"] == "Results"
        assert position_diffs[0]["expected"] == "0"
        assert position_diffs[0]["found"] == "1"

    def test_sheet_still_found_by_name_only_position_flagged(self, tmp_path: Path) -> None:
        # Confirms this is specifically a structural-position check, not a
        # "sheet missing" false positive -- name-based lookup still works.
        wb = load_workbook(GOLDEN_INPUT)
        wb.create_sheet("Cover Page", 0)
        mutated = tmp_path / "inserted_sheet.xlsx"
        wb.save(mutated)

        _run_cli(mutated)
        report = _read_drift_report()

        kinds = {d["kind"] for d in report["differences"]}
        assert "sheet_missing" not in kinds
        assert "sheet_position_mismatch" in kinds


class TestShiftedHeader:
    """Mutation (c): shift the header (and everything below it) down two rows."""

    def test_exit_code_three(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb["Results"].insert_rows(3, amount=2)
        mutated = tmp_path / "shifted_header.xlsx"
        wb.save(mutated)

        exit_code = _run_cli(mutated)

        assert exit_code == 3

    def test_report_shows_blank_header_at_declared_row(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb["Results"].insert_rows(3, amount=2)
        mutated = tmp_path / "shifted_header.xlsx"
        wb.save(mutated)

        _run_cli(mutated)
        report = _read_drift_report()

        mismatches = {
            d["column_position"]: d
            for d in report["differences"]
            if d["kind"] == "header_text_mismatch"
        }
        # All 5 strict columns now read blank at the declared header_row=3,
        # since the real header text moved down to row 5.
        assert mismatches[1]["expected"] == "Auction Date"
        assert mismatches[1]["found"] == "(blank)"
        assert mismatches[2]["expected"] == "Security Term"
        assert mismatches[2]["found"] == "(blank)"

    def test_report_also_flags_data_too_far_from_header(self, tmp_path: Path) -> None:
        wb = load_workbook(GOLDEN_INPUT)
        wb["Results"].insert_rows(3, amount=2)
        mutated = tmp_path / "shifted_header.xlsx"
        wb.save(mutated)

        _run_cli(mutated)
        report = _read_drift_report()

        kinds = {d["kind"] for d in report["differences"]}
        assert "data_start_too_far" in kinds
