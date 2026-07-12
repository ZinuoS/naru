"""Unit tests for src/naru/lint.py."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from naru import lint

VALID_MANIFEST = """\
name: test_pipeline
version: v1
sheet: Sheet1
target_table: final_test
key: [id]
"""

VALID_FINGERPRINT = """\
{
  "sheet": "Sheet1",
  "header_row": 1,
  "columns": [
    {"name": "id", "type": "integer", "strictness": "strict"}
  ]
}
"""

VALID_SCHEMA = """\
from pydantic import BaseModel


class SourceRow(BaseModel):
    id: int


class TargetRow(BaseModel):
    id: int
"""

VALID_TRANSFORM = """\
def transform(df):
    df = ops.promote_header(df, header_row=1, column_names=["id"])
    return ops.coerce_numeric(df, "id")
"""

VALID_VALIDATIONS = "row_count:\n  min: 1\n"
VALID_CHANGELOG = "# Changelog\n\nv1: initial.\n"


def _write_pipeline_artifact(root: Path, transform_body: str = VALID_TRANSFORM) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    (root / "manifest.yaml").write_text(VALID_MANIFEST)
    (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (root / "schema.py").write_text(VALID_SCHEMA)
    (root / "transform.py").write_text(transform_body)
    (root / "validations.yaml").write_text(VALID_VALIDATIONS)
    (root / "CHANGELOG.md").write_text(VALID_CHANGELOG)

    golden = root / "golden"
    golden.mkdir(exist_ok=True)
    wb = Workbook()
    wb.active.cell(row=1, column=1, value="id")
    wb.active.cell(row=2, column=1, value=1)
    wb.save(golden / "input_sample.xlsx")
    pd.DataFrame({"id": [1]}).to_parquet(golden / "expected_output.parquet", index=False)
    return root


VALID_MAPPING_YAML = """\
target: warehouse.positions
key: [deal_id]
on_duplicate: fail
columns:
  - source: "Cpn (%)"
    target: coupon_rate
    transform: coerce_numeric(scale=0.01)
    basis: synonym
    approved: true
unmapped_source_columns: warn
"""

MAPPING_TARGET_ROW_SCHEMA = """\
from pydantic import BaseModel


class TargetRow(BaseModel):
    coupon_rate: float
"""


def _write_mapping_artifact(root: Path, mapping_yaml: str = VALID_MAPPING_YAML) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    (root / "mapping.yaml").write_text(mapping_yaml)
    (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (root / "schema.py").write_text(MAPPING_TARGET_ROW_SCHEMA)
    return root


class TestPublicOpsNames:
    def test_excludes_incidental_imports(self) -> None:
        names = lint._public_ops_names()
        assert "pd" not in names
        assert "re" not in names
        assert "Literal" not in names
        assert "from_excel" not in names
        assert "DEFAULT_ROW_MARKER" not in names

    def test_includes_real_ops(self) -> None:
        names = lint._public_ops_names()
        assert {"promote_header", "coerce_numeric", "map_values", "derive"} <= names


class TestLintPipelineArtifact:
    def test_clean_artifact_has_no_findings(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        assert lint.lint_pipeline_artifact(root) == []

    def test_missing_required_file_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        (root / "validations.yaml").unlink()
        findings = lint.lint_pipeline_artifact(root)
        assert any("validations.yaml" in str(f.file) and "missing" in f.message for f in findings)

    def test_missing_golden_dir_reported(self, tmp_path: Path) -> None:
        import shutil

        root = _write_pipeline_artifact(tmp_path / "artifact")
        shutil.rmtree(root / "golden")
        findings = lint.lint_pipeline_artifact(root)
        assert any("golden" in str(f.file) and "directory missing" in f.message for f in findings)

    def test_missing_golden_file_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        (root / "golden" / "expected_output.parquet").unlink()
        findings = lint.lint_pipeline_artifact(root)
        assert any("expected_output.parquet" in str(f.file) for f in findings)

    def test_empty_changelog_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        (root / "CHANGELOG.md").write_text("   \n")
        findings = lint.lint_pipeline_artifact(root)
        assert any("CHANGELOG.md is empty" in f.message for f in findings)

    def test_nonempty_changelog_not_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        findings = lint.lint_pipeline_artifact(root)
        assert not any("CHANGELOG" in f.message for f in findings)

    def test_module_level_import_reported_with_line(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(
            tmp_path / "artifact",
            transform_body="import os\n\n\ndef transform(df):\n    return df\n",
        )
        findings = lint.lint_pipeline_artifact(root)
        matches = [f for f in findings if "import 'os'" in f.message]
        assert len(matches) == 1
        assert matches[0].line == 1

    def test_from_import_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(
            tmp_path / "artifact",
            transform_body="from os import path\n\n\ndef transform(df):\n    return df\n",
        )
        findings = lint.lint_pipeline_artifact(root)
        assert any("from os import" in f.message for f in findings)

    def test_import_buried_in_function_body_is_still_caught(self, tmp_path: Path) -> None:
        # The exact gap docs/adr/0003-transform-loading.md documents the
        # runtime exec()-time check cannot catch until the function is
        # called -- lint catches it via AST without ever running the code.
        root = _write_pipeline_artifact(
            tmp_path / "artifact",
            transform_body="def transform(df):\n    import os\n    return df\n",
        )
        findings = lint.lint_pipeline_artifact(root)
        matches = [f for f in findings if "import 'os'" in f.message]
        assert len(matches) == 1
        assert matches[0].line == 2

    def test_unknown_ops_call_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(
            tmp_path / "artifact",
            transform_body="def transform(df):\n    return ops.totally_made_up(df)\n",
        )
        findings = lint.lint_pipeline_artifact(root)
        assert any("ops.totally_made_up is not a naru.ops function" in f.message for f in findings)

    def test_valid_ops_calls_not_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        findings = lint.lint_pipeline_artifact(root)
        assert not any("is not a naru.ops function" in f.message for f in findings)

    def test_transform_syntax_error_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(
            tmp_path / "artifact", transform_body="def transform(df:\n    return df\n"
        )
        findings = lint.lint_pipeline_artifact(root)
        assert any("syntax error" in f.message for f in findings)

    def test_transform_unreadable_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        transform_dir_masquerading_as_file = root / "transform.py"
        transform_dir_masquerading_as_file.unlink()
        transform_dir_masquerading_as_file.mkdir()
        findings = lint.lint_pipeline_artifact(root)
        assert any("could not read file" in f.message for f in findings)

    def test_invalid_fingerprint_reported(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        (root / "fingerprint.json").write_text("{not valid json")
        findings = lint.lint_pipeline_artifact(root)
        assert any("invalid JSON" in f.message for f in findings)

    def test_transform_missing_is_not_ast_checked(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        (root / "transform.py").unlink()
        findings = lint.lint_pipeline_artifact(root)
        assert any("transform.py" in str(f.file) and "missing" in f.message for f in findings)
        assert not any("syntax error" in f.message for f in findings)

    def test_missing_fingerprint_skips_fingerprint_check(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        (root / "fingerprint.json").unlink()
        findings = lint.lint_pipeline_artifact(root)
        assert any("fingerprint.json" in str(f.file) and "missing" in f.message for f in findings)
        assert not any("invalid JSON" in f.message for f in findings)


class TestFindSourceLine:
    def test_returns_none_when_not_found(self) -> None:
        assert lint._find_source_line("target: t\nkey: [id]\n", "Nonexistent Column") is None


class TestLintMappingArtifact:
    def test_clean_artifact_has_no_findings(self, tmp_path: Path) -> None:
        root = _write_mapping_artifact(tmp_path / "mapping_artifact")
        assert lint.lint_mapping_artifact(root) == []

    def test_missing_required_file_reported(self, tmp_path: Path) -> None:
        root = _write_mapping_artifact(tmp_path / "mapping_artifact")
        (root / "schema.py").unlink()
        findings = lint.lint_mapping_artifact(root)
        assert any("schema.py" in str(f.file) and "missing" in f.message for f in findings)

    def test_unapproved_column_reported_with_line(self, tmp_path: Path) -> None:
        text = VALID_MAPPING_YAML.replace("approved: true", "approved: false")
        root = _write_mapping_artifact(tmp_path / "mapping_artifact", mapping_yaml=text)
        findings = lint.lint_mapping_artifact(root)
        matches = [f for f in findings if "not approved" in f.message]
        assert len(matches) == 1
        assert "Cpn (%)" in matches[0].message
        # line 5 is `- source: "Cpn (%)"` in VALID_MAPPING_YAML
        assert matches[0].line == 5

    def test_unparseable_transform_reported(self, tmp_path: Path) -> None:
        text = VALID_MAPPING_YAML.replace(
            "transform: coerce_numeric(scale=0.01)", "transform: os.system('boom')"
        )
        root = _write_mapping_artifact(tmp_path / "mapping_artifact", mapping_yaml=text)
        findings = lint.lint_mapping_artifact(root)
        assert any("bare name" in f.message for f in findings)

    def test_invalid_yaml_reported_as_single_finding(self, tmp_path: Path) -> None:
        root = _write_mapping_artifact(
            tmp_path / "mapping_artifact", mapping_yaml="target: [unclosed\n"
        )
        findings = lint.lint_mapping_artifact(root)
        assert any("invalid YAML" in f.message for f in findings)

    def test_approved_column_with_empty_transform_is_not_reported(self, tmp_path: Path) -> None:
        text = """\
target: t
key: [id]
on_duplicate: fail
columns:
  - source: "id"
    target: id
    transform: ""
    basis: exact
    approved: true
unmapped_source_columns: warn
"""
        root = _write_mapping_artifact(tmp_path / "mapping_artifact", mapping_yaml=text)
        findings = lint.lint_mapping_artifact(root)
        assert findings == []

    def test_missing_mapping_yaml_skips_approval_check(self, tmp_path: Path) -> None:
        root = _write_mapping_artifact(tmp_path / "mapping_artifact")
        (root / "mapping.yaml").unlink()
        findings = lint.lint_mapping_artifact(root)
        assert any("mapping.yaml" in str(f.file) and "missing" in f.message for f in findings)
        assert not any("not approved" in f.message for f in findings)

    def test_missing_fingerprint_skips_fingerprint_check(self, tmp_path: Path) -> None:
        root = _write_mapping_artifact(tmp_path / "mapping_artifact")
        (root / "fingerprint.json").unlink()
        findings = lint.lint_mapping_artifact(root)
        assert any("fingerprint.json" in str(f.file) and "missing" in f.message for f in findings)
        assert not any("invalid JSON" in f.message for f in findings)

    def test_invalid_fingerprint_reported(self, tmp_path: Path) -> None:
        root = _write_mapping_artifact(tmp_path / "mapping_artifact")
        (root / "fingerprint.json").write_text("{not valid json")
        findings = lint.lint_mapping_artifact(root)
        assert any("invalid JSON" in f.message for f in findings)


class TestLintArtifactDispatch:
    def test_manifest_yaml_dispatches_to_pipeline(self, tmp_path: Path) -> None:
        root = _write_pipeline_artifact(tmp_path / "artifact")
        assert lint.lint_artifact(root) == []

    def test_mapping_yaml_dispatches_to_mapping(self, tmp_path: Path) -> None:
        root = _write_mapping_artifact(tmp_path / "mapping_artifact")
        assert lint.lint_artifact(root) == []

    def test_neither_raises_lint_error(self, tmp_path: Path) -> None:
        root = tmp_path / "not_an_artifact"
        root.mkdir()
        with pytest.raises(lint.LintError, match="not an artifact directory"):
            lint.lint_artifact(root)


class TestLintFindingRender:
    def test_render_with_line(self) -> None:
        finding = lint.LintFinding(Path("x/transform.py"), 12, "boom")
        assert finding.render() == "x/transform.py:12: boom"

    def test_render_without_line(self) -> None:
        finding = lint.LintFinding(Path("x/CHANGELOG.md"), None, "empty")
        assert finding.render() == "x/CHANGELOG.md: empty"
