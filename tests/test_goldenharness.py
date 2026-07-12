"""Unit tests for src/naru/goldenharness.py."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from naru import ops
from naru.goldenharness import run_golden_test
from naru.runtime import read_raw_grid

VALID_MANIFEST = """\
name: test_pipeline
version: v1
sheet: Sheet1
target_table: final_test
key: [id]
"""

VALID_FINGERPRINT = """\
{
  "sheet": "Sheet1",
  "header_row": 1,
  "columns": [
    {"name": "id", "type": "integer", "strictness": "strict"},
    {"name": "label", "type": "string", "strictness": "strict"}
  ]
}
"""

VALID_SCHEMA = """\
from pydantic import BaseModel


class SourceRow(BaseModel):
    id: int
    label: str


class TargetRow(BaseModel):
    id: int
    label: str
"""

VALID_VALIDATIONS = "row_count:\n  min: 1\n"
VALID_CHANGELOG = "# Changelog\n\nv1: initial.\n"

PASSTHROUGH_TRANSFORM = (
    "def transform(df):\n"
    "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
    "    return ops.coerce_numeric(df, 'id')\n"
)


def _make_input_workbook(rows: list[tuple[int, str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="label")
    for i, (id_val, label_val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=id_val)
        ws.cell(row=i, column=2, value=label_val)
    return wb


def _natural_output(golden_rows: list[tuple[int, str]], tmp_path: Path) -> pd.DataFrame:
    """The real, natural shape produced by PASSTHROUGH_TRANSFORM for these
    rows -- used as the golden baseline so tests don't have to guess
    dtypes (and don't fight parquet's own dtype normalization on
    round-trip, e.g. object -> int64 for whole-number id columns).
    """
    path = tmp_path / "_natural.xlsx"
    _make_input_workbook(golden_rows).save(path)
    wb = load_workbook(path, data_only=True)
    raw_grid = read_raw_grid(wb, "Sheet1")
    df = ops.promote_header(raw_grid, header_row=1, column_names=["id", "label"])
    return ops.coerce_numeric(df, "id")


def _write_artifact(
    root: Path,
    transform_body: str,
    golden_rows: list[tuple[int, str]],
    expected_df: pd.DataFrame,
) -> None:
    root.mkdir(parents=True, exist_ok=True)
    (root / "manifest.yaml").write_text(VALID_MANIFEST)
    (root / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (root / "schema.py").write_text(VALID_SCHEMA)
    (root / "transform.py").write_text(transform_body)
    (root / "validations.yaml").write_text(VALID_VALIDATIONS)
    (root / "CHANGELOG.md").write_text(VALID_CHANGELOG)

    golden = root / "golden"
    golden.mkdir(exist_ok=True)
    _make_input_workbook(golden_rows).save(golden / "input_sample.xlsx")
    expected_df.to_parquet(golden / "expected_output.parquet", index=False)


class TestRunGoldenTest:
    def test_matching_golden_passes(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        golden_rows = [(1, "a"), (2, "b")]
        expected = _natural_output(golden_rows, tmp_path)
        _write_artifact(root, PASSTHROUGH_TRANSFORM, golden_rows, expected)

        result = run_golden_test(root)
        assert result.ok
        assert result.schema_differences == []
        assert result.value_differences == []

    def test_null_in_both_actual_and_expected_is_not_a_difference(self, tmp_path: Path) -> None:
        # A blank label cell -> None/NaN on both sides, since actual and
        # expected are both derived from the same golden input file here.
        root = tmp_path / "artifact"
        golden_rows: list[tuple[int, str]] = [(1, None)]  # type: ignore[list-item]
        expected = _natural_output(golden_rows, tmp_path)
        _write_artifact(root, PASSTHROUGH_TRANSFORM, golden_rows, expected)

        result = run_golden_test(root)
        assert result.ok
        assert result.value_differences == []

    def test_value_drift_reports_row_and_column(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        golden_rows = [(1, "a"), (2, "b")]
        expected = _natural_output(golden_rows, tmp_path).copy()
        expected.loc[1, "label"] = "CHANGED"  # stale golden expectation
        _write_artifact(root, PASSTHROUGH_TRANSFORM, golden_rows, expected)

        result = run_golden_test(root)
        assert not result.ok
        assert result.schema_differences == []
        assert len(result.value_differences) == 1
        assert "row 1" in result.value_differences[0]
        assert "'label'" in result.value_differences[0]
        assert "CHANGED" in result.value_differences[0]

    def test_row_count_mismatch_reported_as_single_value_difference(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        golden_rows = [(1, "a"), (2, "b")]
        expected = _natural_output(golden_rows, tmp_path).iloc[:1]
        _write_artifact(root, PASSTHROUGH_TRANSFORM, golden_rows, expected)

        result = run_golden_test(root)
        assert not result.ok
        assert len(result.value_differences) == 1
        assert "row count mismatch" in result.value_differences[0]

    def test_schema_drift_missing_column_reported(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        golden_rows = [(1, "a")]
        expected = _natural_output(golden_rows, tmp_path).copy()
        expected["extra"] = True
        _write_artifact(root, PASSTHROUGH_TRANSFORM, golden_rows, expected)

        result = run_golden_test(root)
        assert not result.ok
        assert result.value_differences == []  # short-circuited before value comparison
        assert any("'extra' missing" in d for d in result.schema_differences)

    def test_schema_drift_extra_column_reported(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        golden_rows = [(1, "a")]
        expected = _natural_output(golden_rows, tmp_path)
        _write_artifact(
            root,
            "def transform(df):\n"
            "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
            "    df['bonus'] = 'x'\n"
            "    return df\n",
            golden_rows,
            expected,
        )

        result = run_golden_test(root)
        assert not result.ok
        assert any("'bonus'" in d and "not in golden" in d for d in result.schema_differences)

    def test_dtype_mismatch_reported(self, tmp_path: Path) -> None:
        root = tmp_path / "artifact"
        golden_rows = [(1, "a")]
        expected = _natural_output(golden_rows, tmp_path).copy()
        # PASSTHROUGH_TRANSFORM's coerce_numeric produces float64 for "id" --
        # declare the golden as int64 instead to force a real mismatch.
        expected["id"] = expected["id"].astype("int64")
        _write_artifact(root, PASSTHROUGH_TRANSFORM, golden_rows, expected)

        result = run_golden_test(root)
        assert not result.ok
        assert any("dtype mismatch" in d for d in result.schema_differences)
