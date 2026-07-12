"""Unit tests for scripts/refreeze.py.

Uses real, isolated git repos in tmp_path -- the CHANGELOG-modified gate
shells out to `git status`, so it needs real git state to test honestly.
"""

import importlib.util
import subprocess
import sys
from pathlib import Path
from types import ModuleType

import pandas as pd
import pytest
from openpyxl import Workbook


def _load_refreeze_module() -> ModuleType:
    script_path = Path(__file__).resolve().parent.parent / "scripts" / "refreeze.py"
    spec = importlib.util.spec_from_file_location("refreeze_under_test", script_path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


refreeze_module = _load_refreeze_module()

VALID_MANIFEST = """\
name: test_pipeline
version: v1
sheet: Sheet1
target_table: final_test
key: [id]
"""

VALID_FINGERPRINT = """\
{
  "sheet": "Sheet1",
  "header_row": 1,
  "columns": [
    {"name": "id", "type": "integer", "strictness": "strict"},
    {"name": "label", "type": "string", "strictness": "strict"}
  ]
}
"""

VALID_SCHEMA = """\
from pydantic import BaseModel


class SourceRow(BaseModel):
    id: int
    label: str


class TargetRow(BaseModel):
    id: int
    label: str
"""

VALID_TRANSFORM = (
    "def transform(df):\n"
    "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
    "    return ops.coerce_numeric(df, 'id')\n"
)

VALID_VALIDATIONS = "row_count:\n  min: 1\n"


def _git(*args: str, cwd: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(["git", *args], cwd=cwd, capture_output=True, text=True, check=True)


def _make_input_workbook(rows: list[tuple[int, str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="label")
    for i, (id_val, label_val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=id_val)
        ws.cell(row=i, column=2, value=label_val)
    return wb


@pytest.fixture
def git_repo_with_artifact(tmp_path: Path) -> Path:
    """A tmp git repo, committed, containing one valid artifact at
    'artifact/' -- CHANGELOG.md untouched relative to HEAD to start.
    """
    repo = tmp_path / "repo"
    repo.mkdir()
    _git("init", cwd=repo)
    _git("config", "user.email", "test@example.com", cwd=repo)
    _git("config", "user.name", "Test", cwd=repo)

    artifact = repo / "artifact"
    artifact.mkdir()
    (artifact / "manifest.yaml").write_text(VALID_MANIFEST)
    (artifact / "fingerprint.json").write_text(VALID_FINGERPRINT)
    (artifact / "schema.py").write_text(VALID_SCHEMA)
    (artifact / "transform.py").write_text(VALID_TRANSFORM)
    (artifact / "validations.yaml").write_text(VALID_VALIDATIONS)
    (artifact / "CHANGELOG.md").write_text("# Changelog\n\nv1: initial.\n")

    golden = artifact / "golden"
    golden.mkdir()
    _make_input_workbook([(1, "a"), (2, "b")]).save(golden / "input_sample.xlsx")
    pd.DataFrame({"id": [1.0, 2.0], "label": ["a", "b"], "_src_row": [2, 3]}).to_parquet(
        golden / "expected_output.parquet", index=False
    )

    _git("add", "-A", cwd=repo)
    _git("commit", "-m", "initial", cwd=repo)
    return repo


class TestChangelogModifiedInWorkingTree:
    def test_unmodified_returns_false(self, git_repo_with_artifact: Path) -> None:
        changelog = git_repo_with_artifact / "artifact" / "CHANGELOG.md"
        assert refreeze_module._changelog_modified_in_working_tree(changelog) is False

    def test_modified_returns_true(self, git_repo_with_artifact: Path) -> None:
        changelog = git_repo_with_artifact / "artifact" / "CHANGELOG.md"
        changelog.write_text("# Changelog\n\nv1: initial.\n\nv2: a change.\n")
        assert refreeze_module._changelog_modified_in_working_tree(changelog) is True

    def test_other_file_modified_does_not_count(self, git_repo_with_artifact: Path) -> None:
        (git_repo_with_artifact / "artifact" / "transform.py").write_text(
            VALID_TRANSFORM + "\n# a comment\n"
        )
        changelog = git_repo_with_artifact / "artifact" / "CHANGELOG.md"
        assert refreeze_module._changelog_modified_in_working_tree(changelog) is False


class TestRefreeze:
    def test_refuses_without_changelog_edit(self, git_repo_with_artifact: Path) -> None:
        artifact_path = git_repo_with_artifact / "artifact"
        with pytest.raises(refreeze_module.RefreezeRefusedError, match="CHANGELOG.md"):
            refreeze_module.refreeze(artifact_path)

    def test_succeeds_with_changelog_edit(self, git_repo_with_artifact: Path) -> None:
        artifact_path = git_repo_with_artifact / "artifact"
        (artifact_path / "CHANGELOG.md").write_text("# Changelog\n\nv1: initial.\n\nv2: fix.\n")
        (artifact_path / "transform.py").write_text(
            "def transform(df):\n"
            "    df = ops.promote_header(df, header_row=1, column_names=['id', 'label'])\n"
            "    df = ops.coerce_numeric(df, 'id')\n"
            "    df['label'] = 'CHANGED'\n"
            "    return df\n"
        )

        output_path = refreeze_module.refreeze(artifact_path)

        assert output_path == artifact_path / "golden" / "expected_output.parquet"
        refrozen = pd.read_parquet(output_path)
        assert refrozen["label"].tolist() == ["CHANGED", "CHANGED"]

    def test_main_returns_one_on_refusal(
        self, git_repo_with_artifact: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        artifact_path = git_repo_with_artifact / "artifact"
        # main() reads sys.argv directly; simulate it rather than calling
        # refreeze() itself, so this exercises main()'s own error handling.
        old_argv = sys.argv
        try:
            sys.argv = ["refreeze.py", str(artifact_path)]
            exit_code = refreeze_module.main()
        finally:
            sys.argv = old_argv
        assert exit_code == 1
        assert "refusing to refreeze" in capsys.readouterr().err

    def test_main_returns_zero_on_success(
        self, git_repo_with_artifact: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        artifact_path = git_repo_with_artifact / "artifact"
        (artifact_path / "CHANGELOG.md").write_text("# Changelog\n\nv1: initial.\n\nv2: fix.\n")

        old_argv = sys.argv
        try:
            sys.argv = ["refreeze.py", str(artifact_path)]
            exit_code = refreeze_module.main()
        finally:
            sys.argv = old_argv
        assert exit_code == 0
        assert "refrozen" in capsys.readouterr().out

    def test_main_requires_exactly_one_argument(self, capsys: pytest.CaptureFixture[str]) -> None:
        old_argv = sys.argv
        try:
            sys.argv = ["refreeze.py"]
            exit_code = refreeze_module.main()
        finally:
            sys.argv = old_argv
        assert exit_code == 1
        assert "usage:" in capsys.readouterr().err
