"""Mirror a client file into a SQL or Excel target through a frozen
Mapping Artifact: `naru mirror`, per docs/spec.md §2.7.

Sequence:
  1. Fingerprint check on the source file (naru.fingerprint, reusing the
     same Week 4 machinery naru.runtime.run uses -- see
     naru.runtime.FingerprintDriftError). For an Excel target, the
     warehouse side gets the *same* check against its own declared
     header region (naru.mapping.ExcelTarget/warehouse_fingerprint.json)
     -- spec.md §2.7: "fingerprint check on both sides... each halt with
     a report."
  2. Apply the frozen crosswalk: promote the file's own header row to
     column names, then run each column's transform (naru.mapping.
     apply_transform -- ast-parsed, never eval()).
  3. Key-based duplicate check against the target (a SQLite table, or the
     existing rows in an Excel region), checking both the batch against
     existing rows AND the batch against itself (two source rows sharing
     a key is just as much a collision as one colliding with an existing
     row). Any collision aborts before anything is written -- on_duplicate
     is always "fail" in v0.1 (naru.mapping.Mapping rejects "skip"
     outright), so there is no partial/upsert path to fall into.
  4. Build the reconciliation summary: rows in/out, per-numeric-column
     sums pre/post transform, unmapped source columns (and, since
     unmapped_source_columns: fail aborts, only "warn" columns ever reach
     this point), and target columns this mapping doesn't populate.
  5. dry_run=True (the default) writes nothing. dry_run=False:
     - SQL target: the standard store path (store.create_mirror_table,
       store.load_mirror_rows), with lineage tracing every mirrored row
       to its source file hash and row span, exactly as naru.runtime.
       run's own final-table loads do.
     - Excel target: a timestamped backup of the warehouse file is
       written first, then new rows are appended strictly below the
       declared region's last existing data row -- no existing cell,
       formula, formatting, named range, or other sheet is ever touched.
       There is no SQLite lineage table for an Excel target (nothing
       here writes to SQLite at all); the backup file plus the new
       rows' position (append-only, first new row = old last row + 1)
       is the audit trail.

mirror() itself does no printing (CLAUDE.md directive 5: pure functions
in src/naru/, I/O at the edges) -- ReconciliationSummary.render() builds
the human-readable trust document; the CLI/demo layer decides when to
print it.
"""

import datetime as dt
import io
import shutil
import sqlite3
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from pydantic import BaseModel

from naru import ops, store
from naru.fingerprint import FingerprintCheckResult, check_fingerprint
from naru.mapping import Mapping, MappingArtifact, load_mapping_artifact
from naru.mapping import apply_transform as apply_column_transform
from naru.runtime import FingerprintDriftError, read_raw_grid


class MirrorError(Exception):
    """A mirror run can't proceed: a mapped source column is missing from
    the live file's header row, or unmapped_source_columns: fail found
    unmapped columns. Never raised for a duplicate key -- see
    MirrorDuplicateKeyError for that.
    """


class MirrorDuplicateKeyError(Exception):
    """The batch would collide on the mapping's natural key -- either with
    an existing row in the target (a SQL table's active rows, or an
    Excel region's existing data rows), or with another row in this same
    batch. Names every colliding key so the operator can investigate
    before deciding to re-run; nothing is written.
    """

    def __init__(
        self,
        colliding_with_existing: list[tuple[object, ...]],
        colliding_within_batch: list[tuple[object, ...]],
    ) -> None:
        self.colliding_with_existing = colliding_with_existing
        self.colliding_within_batch = colliding_within_batch
        parts = []
        if colliding_with_existing:
            parts.append(
                f"{len(colliding_with_existing)} already in target: {colliding_with_existing}"
            )
        if colliding_within_batch:
            parts.append(
                f"{len(colliding_within_batch)} duplicated within this file: "
                f"{colliding_within_batch}"
            )
        super().__init__("on_duplicate: fail -- " + "; ".join(parts))


class NumericColumnCheck(BaseModel):
    """Sum of one column's values before and after its transform ran --
    the reconciliation summary shows these side by side so a human can
    sanity-check a unit conversion (e.g. a %-to-decimal scale) rather than
    asserting the two must be numerically equal, which they usually
    shouldn't be.
    """

    source_column: str
    target_column: str
    source_sum: float
    target_sum: float


class ReconciliationSummary(BaseModel):
    """The user-facing trust document a dry (or committed) mirror run
    produces: what came in, what would go out (or did), and everything
    that didn't make it into the target by name. `target` is a SQLite
    table name for a SQL target, or the warehouse file's path for an
    Excel target -- generic on purpose, since a mapping's own `target`
    field (spec.md §2.7) is "SQL table or Excel region ref" either way.
    """

    source_file: str
    target: str
    dry_run: bool
    rows_in: int
    rows_out: int
    numeric_checks: list[NumericColumnCheck]
    unmapped_source_columns: list[str]
    unmapped_source_columns_action: Literal["warn", "fail"]
    target_columns_not_populated: list[str]

    def render(self) -> str:
        """Aligned, plain-text rendering -- format it like something you'd
        show an MD: no unexplained jargon, numbers lined up in columns.
        """
        lines = ["Mirror reconciliation summary", "=" * 30, ""]
        lines.append(f"Source file    : {self.source_file}")
        lines.append(f"Target         : {self.target}")
        mode = "DRY RUN -- nothing written" if self.dry_run else "COMMITTED"
        lines.append(f"Mode           : {mode}")
        lines.append(f"Rows in file   : {self.rows_in}")
        lines.append(f"Rows written   : {self.rows_out}")
        lines.append("")

        lines.append("Numeric column checks (sum of source values -> sum of target values)")
        if self.numeric_checks:
            source_width = max(len(c.source_column) for c in self.numeric_checks)
            target_width = max(len(c.target_column) for c in self.numeric_checks)
            for check in self.numeric_checks:
                lines.append(
                    f"  {check.source_column.ljust(source_width)} -> "
                    f"{check.target_column.ljust(target_width)} : "
                    f"{check.source_sum:>14,.4f}  ->  {check.target_sum:>14,.4f}"
                )
        else:
            lines.append("  (no numeric target columns in this mapping)")
        lines.append("")

        lines.append(f"Unmapped source columns (policy: {self.unmapped_source_columns_action})")
        if self.unmapped_source_columns:
            for name in self.unmapped_source_columns:
                lines.append(f'  - "{name}" -- present in the file, not in the mapping')
        else:
            lines.append("  (none -- every source column is mapped)")
        lines.append("")

        lines.append("Target columns not populated by this mapping")
        if self.target_columns_not_populated:
            for name in self.target_columns_not_populated:
                lines.append(f"  - {name}")
        else:
            lines.append("  (none -- every target column is populated)")

        return "\n".join(lines)


class MirrorResult(BaseModel):
    dry_run: bool
    run_id: int | None
    row_ids: list[int]
    summary: ReconciliationSummary
    backup_path: Path | None = None


def _table_name_for_target(target: str) -> str:
    """SQLite table name for a mapping's `target` (e.g.
    "warehouse.positions" -> "warehouse_positions").
    """
    return target.replace(".", "_")


def _crosswalk_source_columns(raw_grid: pd.DataFrame, header_row: int, n_cols: int) -> pd.DataFrame:
    """Promote the file's OWN header row text to column names (unlike
    naru.ops.promote_header's usual caller-supplied names, a Mapping
    Artifact's `columns[].source` values must match the file's literal
    header text, since that's what naru map suggest read them from).
    """
    header_texts = [str(v) for v in raw_grid.iloc[header_row - 1, :n_cols]]
    return ops.promote_header(raw_grid, header_row=header_row, column_names=header_texts)


def _apply_crosswalk(crosswalked: pd.DataFrame, mapping: Mapping) -> pd.DataFrame:
    working = crosswalked
    for column in mapping.columns:
        if column.source not in working.columns:
            raise MirrorError(
                f"mapping column source {column.source!r} not found in the file's "
                f"header row -- found {sorted(c for c in working.columns if c != '_src_row')}"
            )
        if column.transform:
            working = apply_column_transform(working, column.source, column.transform)
    return working


def _numeric_checks(
    crosswalked: pd.DataFrame, transformed: pd.DataFrame, output: pd.DataFrame, mapping: Mapping
) -> list[NumericColumnCheck]:
    checks = []
    for column in mapping.columns:
        if not pd.api.types.is_numeric_dtype(output[column.target]):
            continue
        source_sum = float(
            ops.coerce_numeric(crosswalked, column.source, allow_null=True)[column.source].sum(
                skipna=True
            )
        )
        target_sum = float(transformed[column.source].sum(skipna=True))
        checks.append(
            NumericColumnCheck(
                source_column=column.source,
                target_column=column.target,
                source_sum=source_sum,
                target_sum=target_sum,
            )
        )
    return checks


@dataclass
class _PreparedMirror:
    """Everything the SQL and Excel target paths share: the crosswalked
    and transformed output rows, plus every reconciliation-summary fact
    that doesn't depend on which kind of target this mapping has.
    """

    output: pd.DataFrame
    rows_in: int
    unmapped_source_columns: list[str]
    target_columns_not_populated: list[str]
    numeric_checks: list[NumericColumnCheck]


def _prepare(
    artifact: MappingArtifact, raw_bytes: bytes, fingerprint_check: FingerprintCheckResult
) -> _PreparedMirror:
    mapping = artifact.mapping
    assert fingerprint_check.matched_sheet is not None  # ok=True guarantees this

    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
    raw_grid = read_raw_grid(wb, fingerprint_check.matched_sheet)
    crosswalked = _crosswalk_source_columns(
        raw_grid, artifact.fingerprint.header_row, len(artifact.fingerprint.columns)
    )

    unmapped_source_columns = sorted(
        set(crosswalked.columns) - {"_src_row"} - {column.source for column in mapping.columns}
    )
    if unmapped_source_columns and mapping.unmapped_source_columns == "fail":
        raise MirrorError(
            f"unmapped_source_columns: fail -- {unmapped_source_columns} present in the "
            "source file but not in the mapping"
        )

    transformed = _apply_crosswalk(crosswalked, mapping)
    output = pd.DataFrame({column.target: transformed[column.source] for column in mapping.columns})
    output["_src_row"] = transformed["_src_row"]

    target_columns_not_populated = sorted(
        set(artifact.target_row.model_fields) - {column.target for column in mapping.columns}
    )

    return _PreparedMirror(
        output=output,
        rows_in=len(crosswalked),
        unmapped_source_columns=unmapped_source_columns,
        target_columns_not_populated=target_columns_not_populated,
        numeric_checks=_numeric_checks(crosswalked, transformed, output, mapping),
    )


def _check_duplicate_keys(
    mapping: Mapping, output: pd.DataFrame, existing_keys: set[tuple[object, ...]]
) -> None:
    new_keys = [tuple(row) for row in output[mapping.key].itertuples(index=False, name=None)]
    colliding_with_existing = sorted(set(new_keys) & existing_keys)
    key_counts = Counter(new_keys)
    colliding_within_batch = sorted(key for key, count in key_counts.items() if count > 1)
    if colliding_with_existing or colliding_within_batch:
        raise MirrorDuplicateKeyError(colliding_with_existing, colliding_within_batch)


def _mirror_sql(
    artifact: MappingArtifact,
    source_file: Path,
    raw_bytes: bytes,
    db_path: Path,
    raw_dir: Path,
    fingerprint_check: FingerprintCheckResult,
    prepared: _PreparedMirror,
    dry_run: bool,
) -> MirrorResult:
    assert fingerprint_check.matched_sheet is not None  # ok=True guarantees this
    mapping = artifact.mapping
    table_name = _table_name_for_target(mapping.target)
    output = prepared.output

    conn = sqlite3.connect(db_path)
    store.init_db(conn)
    store.create_mirror_table(conn, table_name, artifact.target_row)

    existing_keys = store.mirror_table_keys(conn, table_name, mapping.key)
    try:
        _check_duplicate_keys(mapping, output, existing_keys)
    except MirrorDuplicateKeyError:
        conn.close()
        raise

    summary = ReconciliationSummary(
        source_file=source_file.name,
        target=table_name,
        dry_run=dry_run,
        rows_in=prepared.rows_in,
        rows_out=len(output),
        numeric_checks=prepared.numeric_checks,
        unmapped_source_columns=prepared.unmapped_source_columns,
        unmapped_source_columns_action=mapping.unmapped_source_columns,
        target_columns_not_populated=prepared.target_columns_not_populated,
    )

    if dry_run:
        conn.close()
        return MirrorResult(dry_run=True, run_id=None, row_ids=[], summary=summary)

    run_id = store.register_run(conn, mapping.target, "mirror")
    file_sha256 = store.register_raw_file(conn, raw_dir, raw_bytes, source_file.name, run_id)
    target_columns = [column.target for column in mapping.columns]
    rows = [
        {col: record[col] for col in target_columns} for record in output.to_dict(orient="records")
    ]
    row_markers = output["_src_row"].tolist()
    row_ids = store.load_mirror_rows(
        conn,
        table_name,
        rows=rows,
        row_markers=row_markers,
        run_id=run_id,
        file_sha256=file_sha256,
        sheet=fingerprint_check.matched_sheet,
        lineage_version="mirror",
    )
    conn.close()
    return MirrorResult(dry_run=False, run_id=run_id, row_ids=row_ids, summary=summary)


def _write_backup(target_path: Path, timestamp: dt.datetime) -> Path:
    """Timestamped copy of the warehouse file, written before any
    modification (spec.md §2.7).

    `timestamp` is always caller-supplied (see mirror()'s `now`
    parameter) -- this function itself never reads the wall clock, so
    it's trivially deterministic to test. The clock read that produces a
    real timestamp when the caller doesn't supply one lives in mirror(),
    once, and only ever labels a backup FILENAME -- it never reaches a
    mirrored row's data, so it isn't the kind of wall-clock dependence
    CLAUDE.md prime directive 2 is about (output determinism); it's
    closer to a log line's timestamp than a runtime decision.
    """
    stamp = timestamp.strftime("%Y%m%dT%H%M%SZ")
    backup_path = target_path.with_name(f"{target_path.stem}.{stamp}.bak{target_path.suffix}")
    shutil.copy2(target_path, backup_path)
    return backup_path


def _mirror_excel(
    artifact: MappingArtifact,
    source_file: Path,
    prepared: _PreparedMirror,
    dry_run: bool,
    now: dt.datetime,
) -> MirrorResult:
    mapping = artifact.mapping
    excel_target = mapping.excel_target
    warehouse_fingerprint = artifact.warehouse_fingerprint
    assert excel_target is not None  # guaranteed by the caller's dispatch
    assert warehouse_fingerprint is not None  # guaranteed by load_mapping_artifact
    output = prepared.output
    target_path = artifact.root / excel_target.path

    # Warehouse-side fingerprint check -- the same Fingerprint/
    # check_fingerprint machinery as the source side (spec.md §2.7).
    # Loaded data_only=True, and never touched again: same reason as
    # naru.runtime.run and mirror()'s own source-side check -- openpyxl
    # silently materializes (grows max_row/max_column on) any cell it
    # touches, and fingerprint checking's type-sampling probes cells
    # past the region's real extent.
    warehouse_check = check_fingerprint(
        warehouse_fingerprint, load_workbook(target_path, data_only=True)
    )
    if not warehouse_check.ok:
        raise FingerprintDriftError(warehouse_check)

    # A second, independent load, WITHOUT data_only -- formula cells'
    # `.value` is the formula string, not its last-computed value, so a
    # later save() re-emits the formula unchanged rather than baking in
    # a snapshot.
    live_wb = load_workbook(target_path)
    ws = live_wb[warehouse_fingerprint.sheet]
    header_row = warehouse_fingerprint.header_row
    first_col = column_index_from_string(excel_target.first_data_col)
    last_col = column_index_from_string(excel_target.last_data_col)

    last_data_row = header_row
    for row in range(header_row + 1, ws.max_row + 1):
        if any(
            ws.cell(row=row, column=col).value is not None for col in range(first_col, last_col + 1)
        ):
            last_data_row = row

    key_positions = {
        field: first_col + excel_target.column_order.index(field) for field in mapping.key
    }
    existing_keys: set[tuple[object, ...]] = set()
    for row in range(header_row + 1, last_data_row + 1):
        existing_keys.add(
            tuple(ws.cell(row=row, column=key_positions[k]).value for k in mapping.key)
        )
    _check_duplicate_keys(mapping, output, existing_keys)

    summary = ReconciliationSummary(
        source_file=source_file.name,
        target=excel_target.path,
        dry_run=dry_run,
        rows_in=prepared.rows_in,
        rows_out=len(output),
        numeric_checks=prepared.numeric_checks,
        unmapped_source_columns=prepared.unmapped_source_columns,
        unmapped_source_columns_action=mapping.unmapped_source_columns,
        target_columns_not_populated=prepared.target_columns_not_populated,
    )

    if dry_run:
        return MirrorResult(dry_run=True, run_id=None, row_ids=[], summary=summary)

    backup_path = _write_backup(target_path, now)

    for offset, record in enumerate(output.to_dict(orient="records")):
        row_num = last_data_row + 1 + offset
        for position, field in enumerate(excel_target.column_order):
            ws.cell(row=row_num, column=first_col + position, value=record.get(field))
    live_wb.save(target_path)

    return MirrorResult(
        dry_run=False, run_id=None, row_ids=[], summary=summary, backup_path=backup_path
    )


def mirror(
    mapping_artifact: Path,
    source_file: Path,
    db_path: Path,
    raw_dir: Path,
    dry_run: bool = True,
    now: dt.datetime | None = None,
) -> MirrorResult:
    """Mirror `source_file` into the target described by a Mapping
    Artifact directory at `mapping_artifact` -- a SQLite table by
    default, or the Excel region declared by `mapping.excel_target` if
    present. See this module's docstring for the full sequence.

    `now` only ever labels an Excel target's backup filename (see
    _write_backup); it's an explicit parameter, not a bare
    dt.datetime.now() call buried in the function body, so tests can
    pass a fixed value and the SQL path is entirely unaffected by it.
    """
    artifact = load_mapping_artifact(mapping_artifact)
    raw_bytes = source_file.read_bytes()
    # Two independent workbook loads, same reason as naru.runtime.run:
    # fingerprint checking's type-sampling probes cells past the sheet's
    # real extent, and openpyxl silently materializes (grows max_row/
    # max_column on) any cell it touches, read or write.
    fingerprint_check = check_fingerprint(
        artifact.fingerprint, load_workbook(io.BytesIO(raw_bytes), data_only=True)
    )
    if not fingerprint_check.ok:
        raise FingerprintDriftError(fingerprint_check)

    prepared = _prepare(artifact, raw_bytes, fingerprint_check)

    if artifact.mapping.excel_target is not None:
        return _mirror_excel(
            artifact, source_file, prepared, dry_run, now or dt.datetime.now(dt.UTC)
        )
    return _mirror_sql(
        artifact, source_file, raw_bytes, db_path, raw_dir, fingerprint_check, prepared, dry_run
    )
