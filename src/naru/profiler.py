"""Deterministic, rule-based Excel profiler.

Per docs/spec.md §2.2: emits, per sheet, dimensions, merged-cell map,
detected header row(s) with confidence, per-column type/null/cardinality/
samples, format smells, and cross-sheet duplicate-header detection. No LLM,
no ML, no network, no wall clock: this is design-time evidence a human or
LLM reasons over, not a runtime decision.

`profile(path)` is a thin I/O shim (opens the workbook) -- every actual
computation lives in pure functions below that take already-loaded
openpyxl/pandas objects, so the analysis logic stays independently testable
per CLAUDE.md directive 5 ("pure functions in src/naru/, I/O at the edges").
"""

import datetime as dt
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel


class HeaderCandidate(BaseModel):
    row: int
    confidence: float


class ColumnProfile(BaseModel):
    position: int
    header_text: str | None
    inferred_type: str
    null_rate: float
    cardinality: int
    samples: list[str]
    smells: list[str]


class SheetProfile(BaseModel):
    name: str
    dimensions: str
    n_rows: int
    n_cols: int
    merged_cells: list[str]
    header_candidates: list[HeaderCandidate]
    columns: list[ColumnProfile]


class Profile(BaseModel):
    source_file: str
    sheets: list[SheetProfile]
    duplicate_headers: list[tuple[str, str]]


def sheet_dimensions(ws: Worksheet) -> tuple[str, int, int]:
    """Return (dimensions string, row count, column count) for a sheet.

    >>> from openpyxl import Workbook
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> _ = ws.cell(row=1, column=1, value="a")
    >>> _ = ws.cell(row=3, column=2, value="b")
    >>> sheet_dimensions(ws)
    ('A1:B3', 3, 2)
    """
    dimensions = ws.dimensions
    return dimensions, ws.max_row, ws.max_column


def merged_cell_ranges(ws: Worksheet) -> list[str]:
    """Return every merged-cell range on a sheet as sorted 'A1:B2' strings.

    >>> from openpyxl import Workbook
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws.merge_cells("B1:C1")
    >>> merged_cell_ranges(ws)
    ['B1:C1']
    """
    return sorted(str(r) for r in ws.merged_cells.ranges)


def _cell_type(value: object) -> str:
    """Classify one cell's Python value into a coarse type label.

    >>> _cell_type(None)
    'empty'
    >>> _cell_type(True)
    'boolean'
    >>> _cell_type(5)
    'integer'
    >>> _cell_type(5.5)
    'float'
    >>> _cell_type(dt.date(2020, 1, 1))
    'date'
    >>> _cell_type("x")
    'string'
    """
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "float"
    if isinstance(value, dt.date | dt.datetime):
        return "date"
    if isinstance(value, str):
        return "string"
    return "other"


def _row_type_counts(ws: Worksheet, row: int, n_cols: int) -> dict[str, int]:
    """Count each coarse type across a row's cells, 1..n_cols."""
    counts: dict[str, int] = {}
    for col in range(1, n_cols + 1):
        cell_type = _cell_type(ws.cell(row=row, column=col).value)
        counts[cell_type] = counts.get(cell_type, 0) + 1
    return counts


def _density_and_text_fraction(counts: dict[str, int], n_cols: int) -> tuple[float, float]:
    """From a row's type counts: (fraction of columns populated, fraction
    of populated cells that are strings).
    """
    non_empty = n_cols - counts.get("empty", 0)
    density = non_empty / n_cols if n_cols else 0.0
    text_fraction = counts.get("string", 0) / non_empty if non_empty else 0.0
    return density, text_fraction


def _non_string_fraction(counts: dict[str, int], n_cols: int) -> float:
    """Fraction of a row's populated cells that are NOT strings -- the
    signal used to detect a transition into a data region.
    """
    non_empty = n_cols - counts.get("empty", 0)
    if non_empty == 0:
        return 0.0
    return (non_empty - counts.get("string", 0)) / non_empty


def detect_header_rows(
    ws: Worksheet,
    density_threshold: float = 0.5,
    text_threshold: float = 0.5,
) -> list[HeaderCandidate]:
    """Scan top-down (no fixed window) for the first row that is dense and
    mostly text, immediately followed by a row with a distinctly different
    (more numeric/date) type profile. See docs/adr/0002-header-detection.md
    for why this algorithm was chosen over a fixed-window alternative.

    Confidence is a simple average of the candidate row's own density and
    the strength of the type transition below it -- it is evidence for a
    human/LLM to weigh, not a calibrated probability (see the ADR's
    Consequences section on over-interpreting it).

    Returns an empty list if no row in the sheet qualifies -- this is
    profiling evidence, not a runtime pipeline step, so an inconclusive
    scan is reported, not raised.

    >>> from openpyxl import Workbook
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> _ = ws.cell(row=1, column=1, value="Name")
    >>> _ = ws.cell(row=1, column=2, value="Amount")
    >>> _ = ws.cell(row=2, column=1, value="a")
    >>> _ = ws.cell(row=2, column=2, value=1.5)
    >>> detect_header_rows(ws)
    [HeaderCandidate(row=1, confidence=0.75)]
    """
    n_cols = ws.max_column
    max_row = ws.max_row
    if n_cols == 0 or max_row < 2:
        return []
    for row in range(1, max_row):
        counts = _row_type_counts(ws, row, n_cols)
        density, text_fraction = _density_and_text_fraction(counts, n_cols)
        if density < density_threshold or text_fraction < text_threshold:
            continue
        next_counts = _row_type_counts(ws, row + 1, n_cols)
        transition = _non_string_fraction(next_counts, n_cols)
        confidence = round(0.5 * density + 0.5 * transition, 4)
        return [HeaderCandidate(row=row, confidence=confidence)]
    return []


_PERCENT_STRING_RE = re.compile(r"^\s*-?\d[\d,]*(\.\d+)?\s*%\s*$")
_THOUSANDS_STRING_RE = re.compile(r"^\s*-?\d{1,3}(,\d{3})+(\.\d+)?\s*$")
_PARENS_NEGATIVE_RE = re.compile(r"^\s*\(\s*[\d,]+(\.\d+)?\s*%?\s*\)\s*$")
_DATE_SERIAL_MIN = 20000
_DATE_SERIAL_MAX = 60000


def _detect_column_smells(non_null_values: list[object]) -> list[str]:
    """Flag format smells among a column's non-null values: '%' strings,
    comma-thousands strings, parens-as-negative strings, and integers in a
    plausible Excel date-serial range.

    >>> _detect_column_smells(["2.5%", "38,000", "(1,234)", 43480])
    ['date_serial_suspect', 'parens_negative', 'percent_string', 'thousands_separator']
    """
    smells: set[str] = set()
    for value in non_null_values:
        if isinstance(value, str):
            if _PERCENT_STRING_RE.match(value):
                smells.add("percent_string")
            if _THOUSANDS_STRING_RE.match(value):
                smells.add("thousands_separator")
            if _PARENS_NEGATIVE_RE.match(value):
                smells.add("parens_negative")
        elif isinstance(value, int) and not isinstance(value, bool):
            if _DATE_SERIAL_MIN <= value <= _DATE_SERIAL_MAX:
                smells.add("date_serial_suspect")
    return sorted(smells)


def _infer_column_type(values: list[object]) -> str:
    """Infer a column's dominant type: the single coarse type shared by all
    non-null values, 'mixed' if they disagree, or 'empty' if all are null.

    >>> _infer_column_type([1, 2, None])
    'integer'
    >>> _infer_column_type([1, "a"])
    'mixed'
    >>> _infer_column_type([None, None])
    'empty'
    """
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "empty"
    types = {_cell_type(v) for v in non_null}
    if len(types) == 1:
        return types.pop()
    return "mixed"


def profile_columns(ws: Worksheet, header_row: int) -> list[ColumnProfile]:
    """Profile every column's data region: all rows strictly below `header_row`.

    >>> from openpyxl import Workbook
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> _ = ws.cell(row=1, column=1, value="Amt")
    >>> _ = ws.cell(row=2, column=1, value="1,000")
    >>> _ = ws.cell(row=3, column=1, value="2,000")
    >>> profile_columns(ws, header_row=1)[0].smells
    ['thousands_separator']
    """
    n_cols = ws.max_column
    max_row = ws.max_row
    columns: list[ColumnProfile] = []
    for col in range(1, n_cols + 1):
        header_value = ws.cell(row=header_row, column=col).value
        header_text = str(header_value) if header_value is not None else None
        values = [ws.cell(row=r, column=col).value for r in range(header_row + 1, max_row + 1)]
        non_null = [v for v in values if v is not None]
        null_rate = round((len(values) - len(non_null)) / len(values), 4) if values else 0.0
        distinct: list[object] = []
        for value in non_null:
            if value not in distinct:
                distinct.append(value)
        columns.append(
            ColumnProfile(
                position=col,
                header_text=header_text,
                inferred_type=_infer_column_type(values),
                null_rate=null_rate,
                cardinality=len(distinct),
                samples=[str(v) for v in distinct[:5]],
                smells=_detect_column_smells(non_null),
            )
        )
    return columns


def detect_duplicate_headers(
    sheet_headers: dict[str, list[str | None]],
) -> list[tuple[str, str]]:
    """Find pairs of sheets (sorted, deduplicated) whose header text is
    identical in order -- a signal that a table shape was copy-pasted or
    reused across sheets.

    >>> detect_duplicate_headers({"A": ["x", "y"], "B": ["x", "y"], "C": ["z"]})
    [('A', 'B')]
    """
    pairs: list[tuple[str, str]] = []
    names = sorted(sheet_headers)
    for i, name_a in enumerate(names):
        for name_b in names[i + 1 :]:
            if sheet_headers[name_a] == sheet_headers[name_b]:
                pairs.append((name_a, name_b))
    return pairs


def profile(path: Path) -> Profile:
    """Profile every sheet in an Excel workbook.

    Thin I/O shim: opens the workbook, then delegates all analysis to the
    pure functions above. No LLM, no network, no wall clock.

    >>> p = profile(Path("tests/fixtures/ust_lite.xlsx"))
    >>> p.sheets[0].header_candidates[0].row
    3
    """
    wb = load_workbook(path, data_only=True)
    sheet_profiles: list[SheetProfile] = []
    sheet_headers: dict[str, list[str | None]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        dimensions, n_rows, n_cols = sheet_dimensions(ws)
        candidates = detect_header_rows(ws)
        columns = profile_columns(ws, candidates[0].row) if candidates else []
        if candidates:
            sheet_headers[name] = [c.header_text for c in columns]
        sheet_profiles.append(
            SheetProfile(
                name=name,
                dimensions=dimensions,
                n_rows=n_rows,
                n_cols=n_cols,
                merged_cells=merged_cell_ranges(ws),
                header_candidates=candidates,
                columns=columns,
            )
        )
    return Profile(
        source_file=path.name,
        sheets=sheet_profiles,
        duplicate_headers=detect_duplicate_headers(sheet_headers),
    )


def to_json(profile_obj: Profile) -> str:
    """Serialize a Profile to deterministic, sorted-key JSON.

    Sorted keys and no timestamps, per CLAUDE.md determinism directive:
    the same input bytes must produce byte-identical profile JSON forever.
    """
    return json.dumps(profile_obj.model_dump(mode="json"), sort_keys=True, indent=2)


if __name__ == "__main__":  # pragma: no cover
    import sys

    for arg in sys.argv[1:]:
        print(to_json(profile(Path(arg))))
