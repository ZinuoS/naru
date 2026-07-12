"""Fingerprint drift detection, per docs/spec.md §2.3.

`check_fingerprint` compares a live workbook against a Fingerprint spec
and reports every difference with sheet/column coordinates, rendered in
the exact format spec.md §2.3 itself uses as an example
("sheet 'Results' col 7: expected `High Yield`, found `High Rate`") --
designed to be read by a human and pasted back into a design-time
recompilation session.

`generate_fingerprint` mechanically derives a Fingerprint from a golden
input sample, so pipeline authors don't hand-write it: a header column
whose cell is covered by a merge gets `position_only` strictness (its
text is unreliable, not absent by mistake); every other column gets
`strict`. Types are inferred from a sample of the data rows below the
header.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from naru.artifact import Fingerprint, HeaderColumnSpec
from naru.profiler import cell_type


@dataclass
class Difference:
    """One concrete way a live file differs from what a fingerprint expects."""

    kind: str
    sheet: str
    expected: str
    found: str
    column_position: int | None = None

    def message(self) -> str:
        """Render in spec.md §2.3's own example format.

        >>> Difference("header_text_mismatch", "Results", "High Yield", "High Rate", 7).message()
        "sheet 'Results' col 7: expected `High Yield`, found `High Rate`"
        """
        if self.column_position is not None:
            return (
                f"sheet {self.sheet!r} col {self.column_position}: "
                f"expected `{self.expected}`, found `{self.found}`"
            )
        return f"sheet {self.sheet!r}: expected `{self.expected}`, found `{self.found}`"


@dataclass
class FingerprintCheckResult:
    ok: bool
    differences: list[Difference] = field(default_factory=list)
    matched_sheet: str | None = None


def _match_sheet(
    fingerprint: Fingerprint, sheet_names: list[str]
) -> tuple[str | None, list[Difference]]:
    """Find the one sheet the fingerprint's `sheet`/`sheet_is_regex` refers to."""
    if fingerprint.sheet_is_regex:
        pattern = re.compile(fingerprint.sheet)
        matches = [name for name in sheet_names if pattern.fullmatch(name)]
    else:
        matches = [name for name in sheet_names if name == fingerprint.sheet]

    if not matches:
        return None, [
            Difference(
                kind="sheet_missing",
                sheet=fingerprint.sheet,
                expected=fingerprint.sheet,
                found=", ".join(sheet_names) if sheet_names else "(no sheets)",
            )
        ]
    if len(matches) > 1:
        return None, [
            Difference(
                kind="sheet_ambiguous",
                sheet=fingerprint.sheet,
                expected="exactly one matching sheet",
                found=f"{len(matches)} matches: {', '.join(matches)}",
            )
        ]

    matched = matches[0]
    differences: list[Difference] = []
    if fingerprint.sheet_index is not None:
        actual_index = sheet_names.index(matched)
        if actual_index != fingerprint.sheet_index:
            differences.append(
                Difference(
                    kind="sheet_position_mismatch",
                    sheet=matched,
                    expected=str(fingerprint.sheet_index),
                    found=str(actual_index),
                )
            )
    return matched, differences


def _check_header_signature(
    fingerprint: Fingerprint, ws: Worksheet, sheet_name: str
) -> list[Difference]:
    """Check header text at `header_row` for every `strict` column.

    `position_only`/`optional` columns are skipped entirely -- a column
    covered by a merged cell reads back as the same blank/duplicated text
    in a correct file as in a drifted one, so text isn't a usable signal
    for them (that's the whole reason they're marked non-strict).
    """
    differences = []
    for position, col_spec in enumerate(fingerprint.columns, start=1):
        if col_spec.strictness != "strict":
            continue
        cell_value = ws.cell(row=fingerprint.header_row, column=position).value
        if cell_value != col_spec.name:
            differences.append(
                Difference(
                    kind="header_text_mismatch",
                    sheet=sheet_name,
                    column_position=position,
                    expected=col_spec.name,
                    found=str(cell_value) if cell_value is not None else "(blank)",
                )
            )
    return differences


def _check_data_starts_near_header(
    fingerprint: Fingerprint, ws: Worksheet, sheet_name: str
) -> list[Difference]:
    """Structural invariant: a non-blank data row within N rows of the header."""
    n_cols = len(fingerprint.columns)
    for offset in range(1, fingerprint.max_rows_from_header_to_data + 1):
        row = fingerprint.header_row + offset
        if any(ws.cell(row=row, column=col).value is not None for col in range(1, n_cols + 1)):
            return []
    max_rows = fingerprint.max_rows_from_header_to_data
    return [
        Difference(
            kind="data_start_too_far",
            sheet=sheet_name,
            expected=f"non-blank data within {max_rows} row(s) of header",
            found="no non-blank row found in that range",
        )
    ]


def _sample_column_types(
    ws: Worksheet, first_row: int, position: int, sample_size: int
) -> list[str]:
    types = []
    for row in range(first_row, first_row + sample_size):
        value = ws.cell(row=row, column=position).value
        if value is not None:
            types.append(cell_type(value))
    return types


def _check_column_types(
    fingerprint: Fingerprint, ws: Worksheet, sheet_name: str, sample_size: int = 20
) -> list[Difference]:
    """Check each column's dominant sampled type against its declared type.

    Applies to `strict` and `position_only` columns alike: type is a
    separate signal from name, and position_only columns rely on it since
    their header text isn't checked at all.
    """
    differences = []
    first_data_row = fingerprint.header_row + 1
    for position, col_spec in enumerate(fingerprint.columns, start=1):
        if col_spec.strictness == "optional":
            continue
        samples = _sample_column_types(ws, first_data_row, position, sample_size)
        if not samples:
            continue
        dominant = max(set(samples), key=samples.count)
        if dominant != col_spec.type:
            differences.append(
                Difference(
                    kind="column_type_mismatch",
                    sheet=sheet_name,
                    column_position=position,
                    expected=col_spec.type,
                    found=dominant,
                )
            )
    return differences


def check_fingerprint(fingerprint: Fingerprint, wb: Workbook) -> FingerprintCheckResult:
    """Compare a live workbook against a Fingerprint spec.

    Checks, in order: sheet presence (name or regex) and position, header
    text for strict columns, the header-to-data structural invariant, and
    column types. Sheet-matching failures short-circuit the rest (there's
    nothing sensible to check inside a sheet that wasn't found).
    """
    matched_sheet, differences = _match_sheet(fingerprint, wb.sheetnames)
    if matched_sheet is None:
        return FingerprintCheckResult(ok=False, differences=differences)

    ws = wb[matched_sheet]
    differences = list(differences)
    differences += _check_header_signature(fingerprint, ws, matched_sheet)
    differences += _check_data_starts_near_header(fingerprint, ws, matched_sheet)
    differences += _check_column_types(fingerprint, ws, matched_sheet)
    return FingerprintCheckResult(
        ok=not differences, differences=differences, matched_sheet=matched_sheet
    )


def _merged_header_column_positions(ws: Worksheet, header_row: int) -> set[int]:
    """Column positions at `header_row` covered by a merged range (anchor
    or covered cell) -- these columns' header text is unreliable.
    """
    positions: set[int] = set()
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= header_row <= merged_range.max_row:
            positions.update(range(merged_range.min_col, merged_range.max_col + 1))
    return positions


def generate_fingerprint(
    path: Path,
    sheet: str,
    header_row: int,
    max_rows_from_header_to_data: int = 5,
    sample_size: int = 20,
) -> Fingerprint:
    """Mechanically derive a Fingerprint from a golden input sample.

    >>> fp = generate_fingerprint(
    ...     Path("tests/fixtures/ust_lite.xlsx"), "Results", header_row=3
    ... )
    >>> fp.columns[0].strictness, fp.columns[0].name
    ('strict', 'Auction Date')
    >>> fp.columns[6].strictness
    'position_only'
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    sheet_index = wb.sheetnames.index(sheet)
    merged_positions = _merged_header_column_positions(ws, header_row)

    columns = []
    for position in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=position).value
        samples = _sample_column_types(ws, header_row + 1, position, sample_size)
        inferred_type = max(set(samples), key=samples.count) if samples else "string"
        strictness = "position_only" if position in merged_positions else "strict"
        columns.append(
            HeaderColumnSpec(
                name=str(header_value) if header_value is not None else f"column_{position}",
                type=inferred_type,  # type: ignore[arg-type]
                strictness=strictness,  # type: ignore[arg-type]
            )
        )

    return Fingerprint(
        sheet=sheet,
        sheet_index=sheet_index,
        header_row=header_row,
        columns=columns,
        max_rows_from_header_to_data=max_rows_from_header_to_data,
    )
