"""Runner: naru run <artifact> <input> -- spec.md §2's runtime sequence.

Deterministic, offline: no network, no wall-clock reads except the
explicit `as_of` parameter, which is stored exactly as given and never
auto-filled from the clock (CLAUDE.md prime directive 1).

Fingerprint checking (step a) is real (src/naru/fingerprint.py) -- a
mismatch raises FingerprintDriftError, which the CLI (src/naru/__main__.py)
turns into exit code 3 and a drift_report.json, per spec.md §2.3.

Output-contract validation (step d) is also real (src/naru/validations.py).
Every check's outcome is persisted to meta_validation_results before any
pass/fail decision is made, so the audit trail survives even a failed run;
any FAIL raises before store.load_final_rows() is ever called, so a failed
run never writes a partial final table -- atomicity by never starting,
not by rollback. Per-row TargetRow schema conformance is a separate,
narrower check (does this row match its declared shape?) from the
validations.yaml business-rule engine (row-count bounds, sum
preservation, etc.) -- both run, but they're checking different things.
"""

import datetime as dt
import io
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from naru import store
from naru.artifact import load_artifact
from naru.fingerprint import FingerprintCheckResult, check_fingerprint
from naru.validations import ValidationOutcome, run_validations


class RuntimeCheckError(Exception):
    """A runtime-sequence check failed: output-contract validation or
    per-row schema conformance. (Fingerprint drift raises
    FingerprintDriftError instead -- it needs its own exit code and
    drift_report.json, so it isn't just another instance of this.)
    """


class FingerprintDriftError(Exception):
    """The source file doesn't match what the pipeline was compiled
    against. Carries the full FingerprintCheckResult so the caller (the
    CLI) can render drift_report.json -- see spec.md §2.3.
    """

    def __init__(self, result: FingerprintCheckResult) -> None:
        self.result = result
        summary = "; ".join(d.message() for d in result.differences)
        super().__init__(f"fingerprint drift: {summary}")


def read_raw_grid(wb: Workbook, sheet_name: str) -> pd.DataFrame:
    """Read one already-loaded sheet's raw cell grid into a DataFrame,
    tagging each row with its 1-indexed source row number in `_src_row`.
    """
    ws = wb[sheet_name]
    records: list[dict[int | str, Any]] = []
    for src_row, row in enumerate(ws.iter_rows(), start=1):
        record: dict[int | str, Any] = {i: cell.value for i, cell in enumerate(row)}
        record["_src_row"] = src_row
        records.append(record)
    return pd.DataFrame.from_records(records)


@dataclass
class RunResult:
    run_id: int
    row_ids: list[int]
    fingerprint_check: FingerprintCheckResult
    validation_outcomes: list[ValidationOutcome]


def run(
    artifact_path: Path,
    input_path: Path,
    db_path: Path,
    raw_dir: Path,
    as_of: dt.date | None = None,
) -> RunResult:
    """Execute a pipeline artifact against an input file.

    Sequence (spec.md §2):
      a. fingerprint check -- raises FingerprintDriftError on mismatch
      b. raw zone: store bytes + SHA256
      c. apply frozen transforms
      d. output contract validation -- every outcome persisted, then any
         FAIL raises before touching the final table
      e. load SQLite + lineage rows
      f. register the run

    The run is registered first here, ahead of steps b-e, not last: every
    later step needs a run_id to attach records to. spec.md's a-f
    lettering reads as a narrative walkthrough of what happens, not an
    enforced literal order.
    """
    artifact = load_artifact(artifact_path)

    conn = sqlite3.connect(db_path)
    store.init_db(conn)
    store.create_final_table(conn, artifact.manifest.target_table, artifact.target_row)
    run_id = store.register_run(conn, artifact.manifest.name, artifact.manifest.version, as_of)

    raw_bytes = input_path.read_bytes()

    # Fingerprint checking deliberately probes cells past the sheet's real
    # extent (structural-invariant and type-sampling checks). openpyxl
    # materializes a cell -- silently growing ws.max_row -- on any access,
    # read or write. Loading a fresh workbook for the raw-grid read avoids
    # that probing corrupting what iter_rows() later considers the sheet's
    # true bounds.
    fingerprint_check = check_fingerprint(
        artifact.fingerprint, load_workbook(io.BytesIO(raw_bytes), data_only=True)
    )
    if not fingerprint_check.ok:
        conn.close()
        raise FingerprintDriftError(fingerprint_check)

    file_sha256 = store.register_raw_file(conn, raw_dir, raw_bytes, input_path.name, run_id)

    assert fingerprint_check.matched_sheet is not None  # ok=True guarantees this
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
    raw_grid = read_raw_grid(wb, fingerprint_check.matched_sheet)
    transformed = artifact.transform(raw_grid)

    business_columns = [c for c in transformed.columns if c != "_src_row"]
    target_fields = set(artifact.target_row.model_fields)
    if set(business_columns) != target_fields:
        conn.close()
        raise RuntimeCheckError(
            f"transform output columns {sorted(business_columns)} don't match "
            f"TargetRow fields {sorted(target_fields)}"
        )

    validation_outcomes = run_validations(artifact.validations, raw_grid, transformed)
    store.record_validation_results(
        conn, run_id, [(o.check_name, o.status, o.detail) for o in validation_outcomes]
    )
    failures = [o for o in validation_outcomes if o.status == "FAIL"]
    if failures:
        conn.close()
        detail = "; ".join(f"{o.check_name}: {o.detail}" for o in failures)
        raise RuntimeCheckError(f"{len(failures)} validation(s) failed: {detail}")

    rows: list[dict[str, object]] = []
    row_markers: list[int] = []
    for record in transformed.to_dict(orient="records"):
        validated = artifact.target_row(**{k: record[k] for k in business_columns})
        rows.append(validated.model_dump(mode="json"))
        row_markers.append(record["_src_row"])

    row_ids = store.load_final_rows(
        conn,
        table_name=artifact.manifest.target_table,
        key_columns=artifact.manifest.key,
        rows=rows,
        row_markers=row_markers,
        run_id=run_id,
        verification="TO_VERIFY",
        file_sha256=file_sha256,
        sheet=fingerprint_check.matched_sheet,
        pipeline_version=artifact.manifest.version,
    )

    conn.close()
    return RunResult(
        run_id=run_id,
        row_ids=row_ids,
        fingerprint_check=fingerprint_check,
        validation_outcomes=validation_outcomes,
    )
