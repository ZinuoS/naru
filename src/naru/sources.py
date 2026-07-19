"""Source readers: turn raw input bytes into an openpyxl Workbook.

The runtime, fingerprint engine, golden harness, and raw-grid reader are
all written against an openpyxl ``Workbook`` (docs/spec.md §2.2, §2.3).
v0.1 could only produce one from an ``.xlsx`` file. This module adds
delimited-text (CSV/TSV) sources by parsing their bytes into an in-memory
single-sheet workbook, so every downstream stage -- fingerprint check,
raw-grid read, transform, validation, lineage -- works unchanged.

Design choices, all in service of run-time determinism (spec.md §0.2):

* **Every cell is a string** (empty field -> ``None``). Delimited text
  carries no type metadata, so naru does not guess one: the pipeline's
  ``transform.py`` declares types explicitly via ``coerce_numeric`` /
  ``coerce_date``, exactly as it already must for messy Excel string
  cells. The fingerprint therefore sees ``string`` columns and guards
  header *names*; the ``TargetRow`` contract guards the real types.
* **No sniffing.** The delimiter is fixed by ``source_format`` (``csv``
  -> ``,``, ``tsv`` -> tab) unless the artifact overrides it in
  ``source_options``. csv.Sniffer is deliberately not used -- a source's
  shape is a declared, reviewable fact, never a run-time inference.
* **BOM-tolerant.** Default decoding is ``utf-8-sig`` so a leading BOM
  (common in government CSV exports) never contaminates the first header
  name; overridable via ``source_options.encoding``.
"""

import csv
import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType

SUPPORTED_FORMATS = ("xlsx", "csv", "tsv")

_DEFAULT_DELIMITER = {"csv": ",", "tsv": "\t"}


def source_workbook_from_bytes(
    raw_bytes: bytes,
    source_format: str,
    sheet_name: str,
    options: dict[str, Any] | None = None,
) -> WorkbookType:
    """Build an openpyxl Workbook from raw input bytes.

    ``source_format`` is one of SUPPORTED_FORMATS. ``sheet_name`` names the
    synthetic worksheet for delimited-text sources (ignored for ``xlsx``,
    which carries its own sheet names); it must match the artifact's
    ``manifest.sheet`` / ``fingerprint.sheet`` so the fingerprint engine
    finds it. ``options`` (``manifest.source_options``) may carry
    ``delimiter`` and ``encoding`` overrides for delimited text.
    """
    options = options or {}
    if source_format == "xlsx":
        return load_workbook(io.BytesIO(raw_bytes), data_only=True)
    if source_format in ("csv", "tsv"):
        return _delimited_to_workbook(raw_bytes, source_format, sheet_name, options)
    raise ValueError(
        f"unsupported source_format {source_format!r}; expected one of {SUPPORTED_FORMATS}"
    )


def _delimited_to_workbook(
    raw_bytes: bytes, source_format: str, sheet_name: str, options: dict[str, Any]
) -> WorkbookType:
    encoding = options.get("encoding", "utf-8-sig")
    delimiter = options.get("delimiter", _DEFAULT_DELIMITER[source_format])
    text = raw_bytes.decode(encoding)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    for record in reader:
        # Empty field -> None so blank cells read as "empty" (matching an
        # unset Excel cell), which drop_empty and the null-policy check rely
        # on. Every other value stays a verbatim string.
        ws.append([(value if value != "" else None) for value in record])
    return wb
