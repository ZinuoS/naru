"""Tracer bullet: ust_lite.xlsx -> naru.sqlite, end to end, provenance intact.

Deliberately outside src/naru/ — this is throwaway scaffolding proving the
architecture works before Phase 1 extracts real, tested, typed ops. Every
transform here is hardcoded to this one fixture; nothing here is the
constrained op API described in docs/spec.md §2.4.

Per docs/adr/0001-lineage-carrier.md: provenance rides as an ordinary
`_src_row` column through every transform, not a wrapper type.

PROOF QUERY -- a final output row, joined back to where it came from:

    SELECT
        f.row_id, f.auction_date, f.security_term, f.cusip,
        f.high_yield, f.offering_amt, f.bid_to_cover, f.issue_date,
        l.file_sha256, l.sheet, l.source_row_start, l.source_row_end
    FROM final_auction_results AS f
    JOIN meta_lineage AS l
        ON l.final_table = 'final_auction_results' AND l.row_id = f.row_id
    ORDER BY f.row_id;

Everything below is built backwards from making that query return correctly.
"""

import hashlib
import sqlite3
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

REPO_ROOT = Path(__file__).resolve().parent
FIXTURE_PATH = REPO_ROOT / "tests" / "fixtures" / "ust_lite.xlsx"
RAW_DIR = REPO_ROOT / ".naru" / "raw"
DB_PATH = REPO_ROOT / "naru.sqlite"

SHEET_NAME = "Results"
HEADER_ROW = 3
PIPELINE_VERSION = "tracer-0.0.1"

# Positional column names for the fixture's 7 data columns. The header row
# (row 3) is not read for its text: one of its cells is merged (see the
# fixture generator), so blindly promoting header text would produce a
# null/duplicate column name. See docs/adr/0001-lineage-carrier.md.
COLUMN_NAMES = [
    "auction_date",
    "security_term",
    "cusip",
    "high_yield",
    "offering_amt",
    "bid_to_cover",
    "issue_date",
]

PROOF_QUERY = """
    SELECT
        f.row_id, f.auction_date, f.security_term, f.cusip,
        f.high_yield, f.offering_amt, f.bid_to_cover, f.issue_date,
        l.file_sha256, l.sheet, l.source_row_start, l.source_row_end
    FROM final_auction_results AS f
    JOIN meta_lineage AS l
        ON l.final_table = 'final_auction_results' AND l.row_id = f.row_id
    ORDER BY f.row_id;
"""

DDL = """
CREATE TABLE IF NOT EXISTS meta_runs (
    run_id INTEGER PRIMARY KEY AUTOINCREMENT,
    pipeline_version TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS raw_files (
    sha256 TEXT PRIMARY KEY,
    original_name TEXT NOT NULL,
    ingested_run_id INTEGER NOT NULL REFERENCES meta_runs (run_id)
);

CREATE TABLE IF NOT EXISTS final_auction_results (
    row_id INTEGER PRIMARY KEY,
    auction_date TEXT NOT NULL,
    security_term TEXT NOT NULL,
    cusip TEXT NOT NULL,
    high_yield REAL NOT NULL,
    offering_amt REAL NOT NULL,
    bid_to_cover REAL NOT NULL,
    issue_date TEXT NOT NULL,
    _run_id INTEGER NOT NULL REFERENCES meta_runs (run_id),
    _verification TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS meta_lineage (
    final_table TEXT NOT NULL,
    row_id INTEGER NOT NULL,
    file_sha256 TEXT NOT NULL,
    sheet TEXT NOT NULL,
    source_row_start INTEGER NOT NULL,
    source_row_end INTEGER NOT NULL,
    pipeline_version TEXT NOT NULL,
    run_id INTEGER NOT NULL REFERENCES meta_runs (run_id),
    PRIMARY KEY (final_table, row_id)
);
"""


def sha256_bytes(data: bytes) -> str:
    """Hex-digest SHA-256 of raw file bytes.

    >>> sha256_bytes(b"")
    'e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855'
    """
    return hashlib.sha256(data).hexdigest()


def store_raw_bytes(raw_dir: Path, data: bytes, sha256: str) -> Path:
    """Copy raw bytes into the hash-addressed raw zone, idempotently."""
    raw_dir.mkdir(parents=True, exist_ok=True)
    dest = raw_dir / f"{sha256}.xlsx"
    if not dest.exists():
        dest.write_bytes(data)
    return dest


def read_raw_grid(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    """Read every cell of a sheet into a DataFrame, tagging each row with
    its 1-indexed source row number in a `_src_row` column.

    This bypasses pandas.read_excel deliberately: pandas silently drops
    trailing all-blank rows, which would erase exactly the messiness the
    'drop trailing blanks' transform below is supposed to handle explicitly.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    records: list[dict[int | str, Any]] = []
    for src_row, row in enumerate(ws.iter_rows(), start=1):
        record: dict[int | str, Any] = {i: cell.value for i, cell in enumerate(row)}
        record["_src_row"] = src_row
        records.append(record)
    return pd.DataFrame.from_records(records)


def promote_header(df: pd.DataFrame) -> pd.DataFrame:
    """Drop the two banner rows and the header row; assign column names by
    position (see COLUMN_NAMES docstring above for why not by header text).
    """
    data_rows = df[df["_src_row"] > HEADER_ROW].copy()
    data_rows.columns = pd.Index([*COLUMN_NAMES, "_src_row"])
    return data_rows.reset_index(drop=True)


def drop_trailing_blanks(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows where every business column is null, keeping `_src_row`."""
    business_cols = [c for c in df.columns if c != "_src_row"]
    return df.dropna(subset=business_cols, how="all").reset_index(drop=True)


def coerce_thousands_column(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """Coerce a comma-thousands numeric string column to float, e.g. '38,000' -> 38000.0."""
    df = df.copy()
    df[column] = df[column].astype(str).str.replace(",", "", regex=False).astype(float)
    return df


def coerce_percent_column(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """Coerce a '%'-suffixed string column to a decimal fraction, e.g. '2.747%' -> 0.02747."""
    df = df.copy()
    df[column] = df[column].astype(str).str.rstrip("%").astype(float) / 100.0
    return df


def parse_date_string_column(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """Parse an mm/dd/yyyy string column into ISO-8601 date strings."""
    df = df.copy()
    df[column] = pd.to_datetime(df[column], format="%m/%d/%Y").dt.date.astype(str)
    return df


def parse_excel_serial_column(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """Parse an Excel date-serial numeric column into ISO-8601 date strings."""
    df = df.copy()
    df[column] = df[column].apply(lambda v: from_excel(v).date().isoformat())
    return df


def transform(raw_grid: pd.DataFrame) -> pd.DataFrame:
    """Apply the full hardcoded transform chain, preserving `_src_row` throughout."""
    df = promote_header(raw_grid)
    df = drop_trailing_blanks(df)
    df = coerce_thousands_column(df, "offering_amt")
    df = coerce_percent_column(df, "high_yield")
    df = parse_date_string_column(df, "auction_date")
    df = parse_excel_serial_column(df, "issue_date")
    return df[[*COLUMN_NAMES, "_src_row"]]


def load(
    conn: sqlite3.Connection,
    df: pd.DataFrame,
    file_sha256: str,
    sheet_name: str,
    run_id: int,
) -> None:
    """Load transformed rows into final_auction_results and write matching
    meta_lineage rows, one row of lineage per output row.
    """
    cur = conn.cursor()
    (max_row_id,) = cur.execute(
        "SELECT COALESCE(MAX(row_id), 0) FROM final_auction_results"
    ).fetchone()
    next_row_id = 1 + max_row_id

    for offset, record in enumerate(df.to_dict(orient="records")):
        row_id = next_row_id + offset
        cur.execute(
            """
            INSERT INTO final_auction_results
                (row_id, auction_date, security_term, cusip, high_yield,
                 offering_amt, bid_to_cover, issue_date, _run_id, _verification)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'TO_VERIFY')
            """,
            (
                row_id,
                record["auction_date"],
                record["security_term"],
                record["cusip"],
                record["high_yield"],
                record["offering_amt"],
                record["bid_to_cover"],
                record["issue_date"],
                run_id,
            ),
        )
        src_row = record["_src_row"]
        cur.execute(
            """
            INSERT INTO meta_lineage
                (final_table, row_id, file_sha256, sheet,
                 source_row_start, source_row_end, pipeline_version, run_id)
            VALUES ('final_auction_results', ?, ?, ?, ?, ?, ?, ?)
            """,
            (row_id, file_sha256, sheet_name, src_row, src_row, PIPELINE_VERSION, run_id),
        )
    conn.commit()


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(DDL)

    cur = conn.cursor()
    cur.execute("INSERT INTO meta_runs (pipeline_version) VALUES (?)", (PIPELINE_VERSION,))
    run_id = cur.lastrowid
    assert run_id is not None
    conn.commit()

    raw_bytes = FIXTURE_PATH.read_bytes()
    file_sha256 = sha256_bytes(raw_bytes)
    store_raw_bytes(RAW_DIR, raw_bytes, file_sha256)
    cur.execute(
        "INSERT OR REPLACE INTO raw_files (sha256, original_name, ingested_run_id) "
        "VALUES (?, ?, ?)",
        (file_sha256, FIXTURE_PATH.name, run_id),
    )
    conn.commit()

    raw_grid = read_raw_grid(FIXTURE_PATH, SHEET_NAME)
    final_df = transform(raw_grid)
    load(conn, final_df, file_sha256, SHEET_NAME, run_id)

    print(f"loaded {len(final_df)} rows from run_id={run_id}, file_sha256={file_sha256}")
    print()
    print("--- proof query ---")
    result = conn.execute(PROOF_QUERY).fetchall()
    columns = [d[0] for d in conn.execute(PROOF_QUERY).description]
    print(" | ".join(columns))
    for row in result:
        print(" | ".join(str(v) for v in row))

    conn.close()


if __name__ == "__main__":
    main()
